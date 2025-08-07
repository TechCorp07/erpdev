from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from openpyxl import Workbook
from django.http import JsonResponse, HttpResponse
from django.db.models import Q, Count, Sum, Avg
from django.utils import timezone
from django.core.paginator import Paginator
from django.db import transaction
from django.template.loader import render_to_string, get_template
from django.contrib.auth.models import User
from django.urls import reverse
from decimal import Decimal
from django.conf import settings
from collections import OrderedDict
from weasyprint import HTML
from io import BytesIO
import pandas as pd
import os
import csv
import json
import logging

from quotes import models
from django.contrib.auth.decorators import login_required
from core.decorators import ajax_required, password_expiration_check
from core.utils import create_notification
from .models import Quote, QuoteItem, QuoteRevision, QuoteTemplate
from .forms import (
    QuoteForm, QuoteItemForm, QuickQuoteForm, 
    BulkQuoteUpdateForm, QuoteSearchForm
)
from .email_utils import send_quote_email, send_quote_notification
from crm.models import Client, CustomerInteraction
from inventory.models import Product, Supplier

logger = logging.getLogger(__name__)

# =====================================
# QUOTE DASHBOARD & OVERVIEW
# =====================================

@password_expiration_check
def quote_dashboard(request):
    """
    The main quote dashboard - this is mission control for your quote operations.
    It provides an at-a-glance view of quote pipeline, performance metrics,
    and priority actions needed.
    """
    
    # Get current user's quotes if they're not an admin
    user_filter = Q()
    if not request.user.profile.is_admin:
        user_filter = Q(assigned_to=request.user) | Q(created_by=request.user)
    
    # Core metrics that matter to business
    total_quotes = Quote.objects.filter(user_filter).count()
    active_quotes = Quote.objects.filter(
        user_filter, 
        status__in=['draft', 'sent', 'viewed', 'under_review']
    ).count()
    
    # Financial metrics
    this_month = timezone.now().replace(day=1)
    month_quotes_value = Quote.objects.filter(
        user_filter,
        created_at__gte=this_month,
        status__in=['accepted', 'converted']
    ).aggregate(total=Sum('total_amount'))['total'] or Decimal('0.00')
    
    # Pipeline analysis
    pending_quotes = Quote.objects.filter(
        user_filter,
        status__in=['sent', 'viewed', 'under_review']
    ).count()
    
    quotes_needing_attention = Quote.objects.filter(
        user_filter,
        status='draft'
    ).count()
    
    # Expiring soon (next 7 days)
    expiring_soon = Quote.objects.filter(
        user_filter,
        status__in=['sent', 'viewed', 'under_review'],
        validity_date__lte=timezone.now().date() + timezone.timedelta(days=7)
    ).count()
    
    # Recent activity
    recent_quotes = Quote.objects.filter(user_filter).order_by('-created_at')[:5]
    
    # Quotes by status for chart
    status_distribution = Quote.objects.filter(user_filter).values('status').annotate(
        count=Count('id')
    ).order_by('-count')
    
    # Monthly trend (last 6 months)
    monthly_stats = []
    for i in range(6):
        month_start = (timezone.now().replace(day=1) - 
                      timezone.timedelta(days=32*i)).replace(day=1)
        month_end = (month_start.replace(month=month_start.month+1) 
                    if month_start.month < 12 
                    else month_start.replace(year=month_start.year+1, month=1))
        
        month_count = Quote.objects.filter(
            user_filter,
            created_at__gte=month_start,
            created_at__lt=month_end
        ).count()
        
        monthly_stats.append({
            'month': month_start.strftime('%b %Y'),
            'count': month_count
        })
    
    monthly_stats.reverse()  # Show oldest to newest
    
    context = {
        'total_quotes': total_quotes,
        'active_quotes': active_quotes,
        'month_quotes_value': month_quotes_value,
        'pending_quotes': pending_quotes,
        'quotes_needing_attention': quotes_needing_attention,
        'expiring_soon': expiring_soon,
        'recent_quotes': recent_quotes,
        'status_distribution': status_distribution,
        'monthly_stats': monthly_stats,
    }
    
    return render(request, 'quotes/dashboard.html', context)


@password_expiration_check
@login_required
def quote_analytics(request):
    """
    Enterprise-grade analytics dashboard for quotes.
    - Multiple filter dimensions (date, status, client, assigned user)
    - KPI cards (volume, value, conversion rate)
    - Status distribution & monthly trend charts (Chart.js)
    - Top-client leaderboard
    - Export to Excel & PDF
    """

    # ------------- 1. GET QUERY PARAMS --------------------------------
    start_date    = request.GET.get('start_date')
    end_date      = request.GET.get('end_date')
    status_param  = request.GET.get('status')
    client_id     = request.GET.get('client')
    assigned_id   = request.GET.get('assigned')
    export_format = request.GET.get('export')          # excel | pdf | None

    # ------------- 2. BASE QUERY (user permissions respected) ---------
    quotes = Quote.objects.all().select_related('client', 'assigned_to')
    if not request.user.profile.is_admin:
        quotes = quotes.filter(Q(created_by=request.user) | Q(assigned_to=request.user))

    # ------------- 3. APPLY FILTERS -----------------------------------
    if start_date:
        quotes = quotes.filter(created_at__date__gte=start_date)
    if end_date:
        quotes = quotes.filter(created_at__date__lte=end_date)
    if status_param:
        quotes = quotes.filter(status=status_param)
    if client_id:
        quotes = quotes.filter(client_id=client_id)
    if assigned_id:
        quotes = quotes.filter(assigned_to_id=assigned_id)

    # ------------- 4. METRICS -----------------------------------------
    total_quotes   = quotes.count()
    total_value    = quotes.aggregate(total=Sum('total_amount'))['total'] or 0
    avg_value      = quotes.aggregate(avg=Avg('total_amount'))['avg'] or 0
    converted      = quotes.filter(status__in=['accepted', 'converted']).count()
    conversion_pct = (converted / total_quotes * 100) if total_quotes else 0

    # status distribution
    status_counts = quotes.values('status').annotate(count=Count('id')).order_by('-count')
    status_labels = [dict(Quote.STATUS_CHOICES)[row['status']] for row in status_counts]
    status_data   = [row['count'] for row in status_counts]

    # monthly trend (last 12 months visible in current query set)
    monthly_totals = OrderedDict()
    for q in quotes:
        label = q.created_at.strftime("%b %Y")
        monthly_totals.setdefault(label, 0)
        monthly_totals[label] += float(q.total_amount)

    mon_labels = list(monthly_totals.keys())
    mon_data   = list(monthly_totals.values())

    # top clients
    top_clients = quotes.values('client__name')\
                        .annotate(total=Sum('total_amount'))\
                        .order_by('-total')[:10]

    # ---------- Excel export ----------
    if export_format == "excel":
        qs = quotes.values(
            "quote_number",           # ①
            "client__name",           # ②
            "assigned_to__username",  # ③
            "status",                 # ④
            "total_amount",           # ⑤
            "created_at",             # ⑥
        )

        # Build DataFrame (works even if qs is empty)
        df = pd.DataFrame(list(qs))

        # Ensure the right column order & headers
        headers = [
            "Quote #", "Client", "Owner",
            "Status", "Total Amount", "Created At"
        ]

        if df.empty:
            # Create an empty frame with headers
            df = pd.DataFrame(columns=headers)
        else:
            df.columns = headers            # safe now – axes match

        # ---- Workbook creation continues exactly as before ----
        wb = Workbook()
        ws = wb.active
        ws.title = "Quotes"

        # Write header + data
        for row in pd.concat([df.columns.to_frame().T, df]).itertuples(index=False):
            ws.append(list(row))

        # KPI summary sheet
        summary = wb.create_sheet("Summary")
        summary.append(["Total Quotes", total_quotes])
        summary.append(["Total Value", total_value])
        summary.append(["Average Deal Size", avg_value])
        summary.append(["Conversion %", f"{conversion_pct:.1f}%"])

        # Stream workbook
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
        )
        response["Content-Disposition"] = "attachment; filename=quote_analytics.xlsx"
        return response

    if export_format == 'pdf':
        context = {
            'quotes': quotes,
            'total_quotes': total_quotes,
            'total_value': total_value,
            'avg_value': avg_value,
            'conversion_pct': conversion_pct,
            'status_labels': status_labels,
            'status_data': status_data,
            'mon_labels': mon_labels,
            'mon_data': mon_data,
            'top_clients': top_clients,
        }
        
        html = render_to_string("quotes/quote_analytics_pdf.html", context)
        pdf  = HTML(string=html, base_url=request.build_absolute_uri()).write_pdf()
        resp = HttpResponse(pdf, content_type='application/pdf')
        resp['Content-Disposition'] = 'attachment; filename=quote_analytics.pdf'
        return resp

    # ------------- 6. NORMAL RENDER -----------------------------------
    clients   = Client.objects.filter(status__in=['prospect', 'client'])
    assignees = User.objects.filter(assigned_quotes__isnull=False).distinct()

    return render(request, 'quotes/quote_analytics.html', {
        'quotes': quotes,
        'total_quotes': total_quotes,
        'total_value': total_value,
        'avg_value': avg_value,
        'conversion_pct': conversion_pct,
        'status_labels': status_labels,
        'status_data': status_data,
        'mon_labels': mon_labels,
        'mon_data': mon_data,
        'top_clients': top_clients,
        'clients': clients,
        'assignees': assignees,
    })


def sales_report(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    client_id = request.GET.get('client')
    export_format = request.GET.get('export')

    quotes = Quote.objects.filter(status__in=['accepted', 'converted']).select_related('client')

    if start_date:
        quotes = quotes.filter(created_at__date__gte=start_date)
    if end_date:
        quotes = quotes.filter(created_at__date__lte=end_date)
    if client_id:
        quotes = quotes.filter(client_id=client_id)

    total_quotes = quotes.count()
    total_sales = quotes.aggregate(total=Sum('total_amount'))['total'] or 0
    avg_sale = quotes.aggregate(avg=Avg('total_amount'))['avg'] or 0

    # Monthly trend data
    from collections import OrderedDict
    monthly = OrderedDict()
    for quote in quotes:
        month = quote.created_at.strftime("%b %Y")
        monthly.setdefault(month, 0)
        monthly[month] += float(quote.total_amount)

    chart_labels = list(monthly.keys())
    chart_data = list(monthly.values())
    
    total_quotes = quotes.count()
    total_sales  = quotes.aggregate(total=Sum('total_amount'))['total'] or 0
    avg_sale     = quotes.aggregate(avg=Avg('total_amount'))['avg'] or 0

    # ---------- NEW: conversion % ----------
    converted = quotes.filter(status__in=['accepted', 'converted']).count()
    conversion_pct = (converted / total_quotes * 100) if total_quotes else 0
    # ---------------------------------------

    # --------- Excel export ---------------
    if export_format == "excel":
        qs = quotes.values(
            "quote_number", "client__name", "status", "total_amount", "created_at"
        )

        # build a DataFrame even if qs is empty
        df = pd.DataFrame(list(qs))
        if not df.empty:
            df.columns = [
                "Quote #", "Client", "Status", "Total Amount", "Created At"
            ]
        else:
            df = pd.DataFrame(
                columns=[
                    "Quote #", "Client", "Status", "Total Amount", "Created At"
                ]
            )

        # ------- Create workbook -------
        wb = Workbook()
        ws = wb.active
        ws.title = "Quotes"

        # Header + data
        for row in pd.concat([df.columns.to_frame().T, df]).itertuples(index=False):
            ws.append(list(row))

        # KPI summary sheet
        summary = wb.create_sheet("Summary")
        summary.append(["Total Quotes", total_quotes])
        summary.append(["Total Sales", total_sales])
        summary.append(["Average Deal Size", avg_sale])
        summary.append(["Conversion %", f"{conversion_pct:.1f}%"])

        # In-memory file
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
        )
        response["Content-Disposition"] = "attachment; filename=sales_report.xlsx"
        return response

    elif export_format == 'pdf':
        monthly_pairs = list(zip(chart_labels, chart_data))
        
        return generate_pdf_from_template('quotes/sales_report_pdf.html', {
            'quotes': quotes,
            'total_sales': total_sales,
            'total_quotes': total_quotes,
            'avg_sale': avg_sale,
            'chart_labels': chart_labels,
            'chart_data': chart_data,
            'monthly_pairs': monthly_pairs,
        }, filename='sales_report.pdf')

    clients = Client.objects.filter(status__in=['prospect', 'client'])

    return render(request, 'quotes/sales_report.html', {
        'quotes': quotes,
        'total_sales': total_sales,
        'total_quotes': total_quotes,
        'avg_sale': avg_sale,
        'clients': clients,
        'chart_labels': chart_labels,
        'chart_data': chart_data,
    })

def generate_pdf_from_template(template_path, context, filename='report.pdf'):
    from weasyprint import HTML
    from django.template.loader import render_to_string

    html = render_to_string(template_path, context)
    pdf = HTML(string=html).write_pdf()
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

# =====================================
# QUOTE CRUD OPERATIONS
# =====================================


@password_expiration_check
def quote_list(request):
    """
    List all quotes with advanced filtering and search capabilities.
    This view is designed to handle large datasets efficiently while
    providing powerful search and filtering options.
    """
    
    quotes = Quote.objects.all().select_related(
        'client', 'created_by', 'assigned_to'
    ).prefetch_related('items')
    
    # Apply permission filtering
    if not request.user.profile.is_admin:
        quotes = quotes.filter(
            Q(assigned_to=request.user) | Q(created_by=request.user)
        )
    
    # Handle search and filtering
    search_form = QuoteSearchForm(request.GET or None)
    if search_form.is_valid():
        cleaned_data = search_form.cleaned_data
        
        # Text search across multiple fields
        search_query = cleaned_data.get('search')
        if search_query:
            quotes = quotes.filter(
                Q(quote_number__icontains=search_query) |
                Q(title__icontains=search_query) |
                Q(client__name__icontains=search_query) |
                Q(client__company__icontains=search_query) |
                Q(description__icontains=search_query)
            )
        
        # Filter by specific fields
        if cleaned_data.get('status'):
            quotes = quotes.filter(status=cleaned_data['status'])
        
        if cleaned_data.get('client'):
            quotes = quotes.filter(client=cleaned_data['client'])
        
        if cleaned_data.get('assigned_to'):
            quotes = quotes.filter(assigned_to=cleaned_data['assigned_to'])
        
        # Date range filtering
        if cleaned_data.get('date_from'):
            quotes = quotes.filter(created_at__date__gte=cleaned_data['date_from'])
        
        if cleaned_data.get('date_to'):
            quotes = quotes.filter(created_at__date__lte=cleaned_data['date_to'])
        
        # Amount range filtering
        if cleaned_data.get('amount_min'):
            quotes = quotes.filter(total_amount__gte=cleaned_data['amount_min'])
        
        if cleaned_data.get('amount_max'):
            quotes = quotes.filter(total_amount__lte=cleaned_data['amount_max'])
        
        # Sorting
        sort_by = cleaned_data.get('sort_by') or '-created_at'
        quotes = quotes.order_by(sort_by)
    else:
        quotes = quotes.order_by('-created_at')
    
    # Pagination
    paginator = Paginator(quotes, 25)  # Show 25 quotes per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'search_form': search_form,
        'total_quotes': quotes.count(),
    }
    
    return render(request, 'quotes/quote_list.html', context)


@password_expiration_check
def quote_create(request):
    """
    Create a new quote. This view handles both regular quote creation
    and quick quotes from templates or client requests.
    """
    
    # Check if client is pre-selected (e.g., from CRM client detail page)
    client_id = request.GET.get('client')
    client = None
    if client_id:
        client = get_object_or_404(Client, id=client_id)
    
    if request.method == 'POST':
        form = QuoteForm(request.POST, user=request.user, client=client)
        
        if form.is_valid():
            with transaction.atomic():
                quote = form.save(commit=False)
                
                # Generate quote number
                quote.quote_number = generate_quote_number()
                
                # Set creator and initial assignee
                quote.created_by = request.user
                if not quote.assigned_to:
                    quote.assigned_to = request.user
                
                quote.save()
                
                # Create initial CRM interaction
                CustomerInteraction.objects.create(
                    client=quote.client,
                    interaction_type='quote_draft',
                    subject=f'Quote {quote.quote_number} created',
                    notes=f'New quote created: {quote.title}',
                    created_by=request.user
                )
                
                # Send notification if assigned to someone else
                if quote.assigned_to != request.user:
                    create_notification(
                        user=quote.assigned_to,
                        title="New Quote Assigned",
                        message=f"Quote {quote.quote_number} for {quote.client.name} has been assigned to you.",
                        notification_type="info"
                    )
                
                logger.info(f"Quote {quote.quote_number} created by {request.user.username}")
                messages.success(request, f'Quote {quote.quote_number} created successfully.')
                
                # Redirect to quote builder to add items
                return redirect('quotes:quote_builder', quote_id=quote.id)
    else:
        form = QuoteForm(user=request.user, client=client)
    
    context = {
        'form': form,
        'client': client,
    }
    
    return render(request, 'quotes/quote_form.html', context)


@password_expiration_check
def quote_detail(request, quote_id):
    """
    Detailed view of a specific quote. This is the comprehensive view that shows
    all quote information, history, and provides action buttons for workflow management.
    """
    
    quote = get_object_or_404(Quote, id=quote_id)
    
    # Permission check - users can only view quotes they're involved with
    if not request.user.profile.is_admin:
        if quote.assigned_to != request.user and quote.created_by != request.user:
            messages.error(request, "You don't have permission to view this quote.")
            return redirect('quotes:quote_list')
    
    # Get quote items with related data
    quote_items = quote.items.all().select_related('product', 'supplier')
    
    # Get quote revision history
    revisions = quote.revisions.all()[:5]  # Show last 5 revisions
    
    # Get related CRM interactions
    related_interactions = CustomerInteraction.objects.filter(
        client=quote.client,
        created_at__gte=quote.created_at - timezone.timedelta(days=30)
    ).order_by('-created_at')[:10]
    
    # Calculate profit analysis
    total_cost = sum(
        item.unit_cost * item.quantity for item in quote_items 
        if item.unit_cost
    )
    profit_amount = quote.subtotal - total_cost
    profit_percentage = (
        (profit_amount / total_cost * 100) if total_cost > 0 else Decimal('0.00')
    )
    
    context = {
        'quote': quote,
        'quote_items': quote_items,
        'revisions': revisions,
        'related_interactions': related_interactions,
        'total_cost': total_cost,
        'profit_amount': profit_amount,
        'profit_percentage': profit_percentage,
        'can_edit': quote.status in ['draft', 'sent'] and 
                   (request.user.profile.is_admin or 
                    quote.assigned_to == request.user),
    }
    
    return render(request, 'quotes/quote_detail.html', context)


@password_expiration_check
def quote_builder(request, quote_id):
    """
    Interactive quote builder interface. This is where the magic happens -
    a dynamic interface for adding/removing items, calculating prices,
    and building comprehensive quotes.
    """
    
    quote = get_object_or_404(Quote, id=quote_id)
    
    # Permission check
    if not request.user.profile.is_admin:
        if quote.assigned_to != request.user and quote.created_by != request.user:
            messages.error(request, "You don't have permission to edit this quote.")
            return redirect('quotes:quote_detail', quote_id=quote.id)
    
    # Can't edit sent or accepted quotes without creating revision
    if quote.status in ['accepted', 'converted', 'rejected', 'expired']:
        messages.warning(request, "This quote cannot be edited in its current status.")
        return redirect('quotes:quote_detail', quote_id=quote.id)
    
    # Get existing items
    quote_items = quote.items.all().order_by('sort_order')
    
    # Get available products for adding
    available_products = Product.objects.filter(is_quotable=True).order_by('name')
    
    # Get quote templates for quick adding
    quote_templates = QuoteTemplate.objects.filter(is_active=True)
    
    context = {
        'quote': quote,
        'quote_items': quote_items,
        'available_products': available_products,
        'quote_templates': quote_templates,
        'item_form': QuoteItemForm(),  # Empty form for adding items
    }
    
    return render(request, 'quotes/quote_builder.html', context)

def generate_quote_number():
    """
    Generate a unique quote number in format: QUO-YYYY-NNNN
    This follows a business-friendly numbering scheme that's easy to reference.
    """
    from django.db.models import Max
    
    current_year = timezone.now().year
    year_prefix = f"QUO-{current_year}-"
    
    # Find the highest number for this year
    latest_quote = Quote.objects.filter(
        quote_number__startswith=year_prefix
    ).aggregate(
        max_number=Max('quote_number')
    )['max_number']
    
    if latest_quote:
        # Extract the number part and increment
        number_part = int(latest_quote.split('-')[-1])
        new_number = number_part + 1
    else:
        # First quote of the year
        new_number = 1
    
    return f"{year_prefix}{new_number:04d}"

@ajax_required

def add_quote_item(request, quote_id):
    """
    AJAX endpoint for adding items to quotes in real-time. This endpoint handles
    the complex logic of adding products to quotes while maintaining data integrity
    and providing immediate feedback to users.
    
    Think of this as the 'smart assistant' that helps users add items by:
    - Validating product availability
    - Calculating pricing based on business rules
    - Updating quote totals instantly
    - Providing detailed feedback about any issues
    """
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'})
    
    quote = get_object_or_404(Quote, id=quote_id)
    
    # Security check - ensure user can edit this quote
    if not request.user.profile.is_admin:
        if quote.assigned_to != request.user and quote.created_by != request.user:
            return JsonResponse({'success': False, 'error': 'Permission denied'})
    
    # Parse the incoming data
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON data'})
    
    # Create the quote item with intelligent defaults
    try:
        with transaction.atomic():
            # Get or create the product reference
            product = None
            if data.get('product_id'):
                product = Product.objects.get(id=data['product_id'])
            
            # Calculate the next sort order
            max_sort = quote.items.aggregate(
                max_sort = models.Max('sort_order')
            )['max_sort'] or 0
            
            # Create the quote item
            quote_item = QuoteItem.objects.create(
                quote=quote,
                product=product,
                description=data.get('description', product.name if product else ''),
                quantity=int(data.get('quantity', 1)),
                unit_price=Decimal(str(data.get('unit_price', '0.00'))),
                source_type=data.get('source_type', 'stock'),
                sort_order=max_sort + 1
            )
            
            # If this is a product from inventory, set intelligent defaults
            if product:
                quote_item.unit_cost = product.cost_price or Decimal('0.00')
                quote_item.markup_percentage = calculate_markup_percentage(
                    quote_item.unit_cost, quote_item.unit_price
                )
                
                # Set delivery estimates based on source type
                if quote_item.source_type == 'stock':
                    quote_item.estimated_delivery = timezone.now().date() + timezone.timedelta(days=1)
                elif quote_item.source_type == 'order':
                    lead_time = product.supplier.lead_time_days if product.supplier else 30
                    quote_item.estimated_delivery = timezone.now().date() + timezone.timedelta(days=lead_time)
                
                quote_item.save()
            
            # Recalculate quote totals
            quote.calculate_totals()
            
            # Prepare the response with updated quote information
            response_data = {
                'success': True,
                'item': {
                    'id': quote_item.id,
                    'description': quote_item.description,
                    'quantity': quote_item.quantity,
                    'unit_price': float(quote_item.unit_price),
                    'total_price': float(quote_item.total_price),
                    'source_type': quote_item.source_type,
                    'estimated_delivery': quote_item.estimated_delivery.isoformat() if quote_item.estimated_delivery else None,
                },
                'quote_totals': {
                    'subtotal': float(quote.subtotal),
                    'tax_amount': float(quote.tax_amount),
                    'total_amount': float(quote.total_amount),
                    'item_count': quote.items.count(),
                }
            }
            
            # Log this action for audit trail
            logger.info(f"Item added to quote {quote.quote_number} by {request.user.username}")
            
            return JsonResponse(response_data)
    
    except Product.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Product not found'})
    except ValueError as e:
        return JsonResponse({'success': False, 'error': f'Invalid data: {str(e)}'})
    except Exception as e:
        logger.error(f"Error adding item to quote {quote_id}: {str(e)}")
        return JsonResponse({'success': False, 'error': 'An unexpected error occurred'})

@ajax_required

def update_quote_item(request, quote_id, item_id):
    """
    AJAX endpoint for updating quote items in real-time. This handles quantity changes,
    price adjustments, and other modifications while maintaining business rule integrity.
    
    This endpoint is designed to handle rapid updates (like when users are adjusting
    quantities with spinners or making price changes) without overwhelming the server.
    """
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'})
    
    quote = get_object_or_404(Quote, id=quote_id)
    quote_item = get_object_or_404(QuoteItem, id=item_id, quote=quote)
    
    # Security check
    if not request.user.profile.is_admin:
        if quote.assigned_to != request.user and quote.created_by != request.user:
            return JsonResponse({'success': False, 'error': 'Permission denied'})
    
    try:
        data = json.loads(request.body)
        
        with transaction.atomic():
            # Update the fields that were provided
            updated_fields = []
            
            if 'quantity' in data:
                new_quantity = int(data['quantity'])
                if new_quantity < 1:
                    return JsonResponse({'success': False, 'error': 'Quantity must be at least 1'})
                quote_item.quantity = new_quantity
                updated_fields.append('quantity')
            
            if 'unit_price' in data:
                new_price = Decimal(str(data['unit_price']))
                if new_price < 0:
                    return JsonResponse({'success': False, 'error': 'Price cannot be negative'})
                
                # Check if price is below cost (warning, not error)
                warning = None
                if quote_item.unit_cost and new_price < quote_item.unit_cost:
                    warning = f'Price is below cost (${quote_item.unit_cost})'
                
                quote_item.unit_price = new_price
                updated_fields.append('unit_price')
            
            if 'description' in data:
                quote_item.description = data['description'][:500]  # Limit length
                updated_fields.append('description')
            
            # Save the changes
            quote_item.save(update_fields=updated_fields + ['total_price'])
            
            # Recalculate quote totals
            quote.calculate_totals()
            
            response_data = {
                'success': True,
                'item': {
                    'id': quote_item.id,
                    'quantity': quote_item.quantity,
                    'unit_price': float(quote_item.unit_price),
                    'total_price': float(quote_item.total_price),
                    'description': quote_item.description,
                },
                'quote_totals': {
                    'subtotal': float(quote.subtotal),
                    'tax_amount': float(quote.tax_amount),
                    'total_amount': float(quote.total_amount),
                },
                'warning': warning,
            }
            
            return JsonResponse(response_data)
    
    except (ValueError, json.JSONDecodeError) as e:
        return JsonResponse({'success': False, 'error': f'Invalid data: {str(e)}'})
    except Exception as e:
        logger.error(f"Error updating quote item {item_id}: {str(e)}")
        return JsonResponse({'success': False, 'error': 'An unexpected error occurred'})

@ajax_required

def remove_quote_item(request, quote_id, item_id):
    """
    AJAX endpoint for removing items from quotes. This includes safety checks
    to prevent accidental deletions and maintains quote total integrity.
    """
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'})
    
    quote = get_object_or_404(Quote, id=quote_id)
    quote_item = get_object_or_404(QuoteItem, id=item_id, quote=quote)
    
    # Security check
    if not request.user.profile.is_admin:
        if quote.assigned_to != request.user and quote.created_by != request.user:
            return JsonResponse({'success': False, 'error': 'Permission denied'})
    
    try:
        with transaction.atomic():
            # Store item info for response
            item_description = quote_item.description
            
            # Remove the item
            quote_item.delete()
            
            # Recalculate quote totals
            quote.calculate_totals()
            
            response_data = {
                'success': True,
                'message': f'Removed "{item_description}" from quote',
                'quote_totals': {
                    'subtotal': float(quote.subtotal),
                    'tax_amount': float(quote.tax_amount),
                    'total_amount': float(quote.total_amount),
                    'item_count': quote.items.count(),
                }
            }
            
            return JsonResponse(response_data)
    
    except Exception as e:
        logger.error(f"Error removing quote item {item_id}: {str(e)}")
        return JsonResponse({'success': False, 'error': 'An unexpected error occurred'})

@ajax_required

def search_products(request):
    """
    AJAX endpoint for real-time product search. This powers the dynamic product
    search functionality in the quote builder, providing instant results as users type.
    
    The search algorithm is designed to return the most relevant results by searching
    across multiple fields and ranking results by relevance.
    """
    
    query = request.GET.get('q', '').strip()
    
    if len(query) < 2:
        return JsonResponse({'products': []})
    
    # Build a comprehensive search across multiple fields
    products = Product.objects.filter(
        is_quotable=True
    ).filter(
        Q(name__icontains=query) |
        Q(sku__icontains=query) |
        Q(description__icontains=query) |
        Q(category__name__icontains=query)
    ).select_related('category', 'supplier')[:20]  # Limit to 20 results for performance
    
    # Build the response with all information needed for quote building
    product_data = []
    for product in products:
        # Calculate suggested pricing based on business rules
        suggested_price = calculate_suggested_price(product)
        
        product_info = {
            'id': product.id,
            'sku': product.sku,
            'name': product.name,
            'description': product.description,
            'category': product.category.name if product.category else '',
            'current_stock': product.current_stock if hasattr(product, 'current_stock') else 0,
            'cost_price': float(product.cost_price) if product.cost_price else 0,
            'suggested_price': float(suggested_price),
            'supplier': product.supplier.name if product.supplier else '',
            'lead_time': product.supplier.lead_time_days if product.supplier else 0,
            'minimum_quantity': product.minimum_quote_quantity if hasattr(product, 'minimum_quote_quantity') else 1,
        }
        product_data.append(product_info)
    
    return JsonResponse({'products': product_data})

@ajax_required

def get_product_details(request, product_id):
    """
    AJAX endpoint to get detailed product information for quote building.
    This provides comprehensive product data including pricing suggestions,
    availability, and sourcing options.
    """
    
    try:
        product = Product.objects.select_related('category', 'supplier').get(
            id=product_id, is_quotable=True
        )
        
        # Calculate pricing options based on different markup strategies
        pricing_options = calculate_pricing_options(product)
        
        # Get availability information
        availability = get_product_availability(product)
        
        product_data = {
            'id': product.id,
            'sku': product.sku,
            'name': product.name,
            'description': product.description,
            'detailed_specs': product.detailed_specs if hasattr(product, 'detailed_specs') else '',
            'category': product.category.name if product.category else '',
            'cost_price': float(product.cost_price) if product.cost_price else 0,
            'pricing_options': pricing_options,
            'availability': availability,
            'supplier_info': {
                'name': product.supplier.name if product.supplier else '',
                'lead_time': product.supplier.lead_time_days if product.supplier else 0,
                'currency': product.supplier.currency if product.supplier else 'USD',
            },
            'minimum_quantity': product.minimum_quote_quantity if hasattr(product, 'minimum_quote_quantity') else 1,
            'bulk_discounts': get_bulk_discount_info(product),
        }
        
        return JsonResponse({'success': True, 'product': product_data})
    
    except Product.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Product not found'})
    except Exception as e:
        logger.error(f"Error getting product details for {product_id}: {str(e)}")
        return JsonResponse({'success': False, 'error': 'An unexpected error occurred'})

def calculate_suggested_price(product):
    """
    Calculate intelligent pricing suggestions based on business rules.
    This function encapsulates your pricing strategy and can be easily modified
    as your business rules evolve.
    """
    if not product.cost_price:
        return Decimal('0.00')
    
    # Default markup strategy - this can be made more sophisticated
    base_markup = Decimal('30.00')  # 30% default markup
    
    # Adjust markup based on product category or other factors
    if hasattr(product, 'category') and product.category:
        category_name = product.category.name.lower()
        if 'laptop' in category_name or 'computer' in category_name:
            base_markup = Decimal('25.00')  # Lower margin for computers
        elif 'accessory' in category_name:
            base_markup = Decimal('50.00')  # Higher margin for accessories
    
    # Calculate the suggested price
    suggested_price = product.cost_price * (1 + base_markup / 100)
    
    # Round to nearest cent
    return suggested_price.quantize(Decimal('0.01'))

def calculate_pricing_options(product):
    """
    Generate multiple pricing options for different markup levels.
    This gives sales teams flexibility in pricing while maintaining profitability.
    """
    if not product.cost_price:
        return []
    
    markup_levels = [
        ('conservative', 20, 'Conservative (20%)'),
        ('standard', 30, 'Standard (30%)'),
        ('premium', 50, 'Premium (50%)'),
        ('maximum', 100, 'Maximum (100%)'),
    ]
    
    options = []
    for level_key, markup_percent, level_name in markup_levels:
        price = product.cost_price * (1 + Decimal(str(markup_percent)) / 100)
        options.append({
            'level': level_key,
            'name': level_name,
            'markup_percent': markup_percent,
            'price': float(price.quantize(Decimal('0.01'))),
        })
    
    return options

def get_product_availability(product):
    """
    Determine product availability and sourcing options.
    This helps sales teams set accurate delivery expectations.
    """
    availability = {
        'in_stock': False,
        'stock_quantity': 0,
        'can_order': False,
        'estimated_delivery': None,
        'source_options': [],
    }
    
    if hasattr(product, 'current_stock'):
        availability['stock_quantity'] = product.current_stock
        availability['in_stock'] = product.current_stock > 0
    
    # Add sourcing options
    if availability['in_stock']:
        availability['source_options'].append({
            'type': 'stock',
            'name': 'From Stock',
            'delivery_days': 1,
            'notes': 'Ships immediately'
        })
    
    if product.supplier:
        availability['can_order'] = True
        availability['source_options'].append({
            'type': 'order',
            'name': f'Order from {product.supplier.name}',
            'delivery_days': product.supplier.lead_time_days,
            'notes': f'{product.supplier.lead_time_days} day lead time'
        })
    
    return availability

def get_bulk_discount_info(product):
    """
    Calculate bulk discount information for quantity-based pricing.
    """
    if not hasattr(product, 'bulk_discount_threshold'):
        return None
    
    if not product.bulk_discount_threshold or not product.bulk_discount_percentage:
        return None
    
    return {
        'threshold': product.bulk_discount_threshold,
        'discount_percent': float(product.bulk_discount_percentage),
        'message': f'{product.bulk_discount_percentage}% discount for {product.bulk_discount_threshold}+ units'
    }

def calculate_markup_percentage(cost_price, selling_price):
    """
    Calculate the markup percentage between cost and selling price.
    This is used for profit analysis and reporting.
    """
    if not cost_price or cost_price == 0:
        return Decimal('0.00')
    
    markup = ((selling_price - cost_price) / cost_price) * 100
    return markup.quantize(Decimal('0.01'))

try:
    from weasyprint import HTML, CSS
    from weasyprint.text.fonts import FontConfiguration
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False
    logger.warning("WeasyPrint not available. PDF generation will be disabled.")


def generate_quote_pdf(request, quote_id):
    """
    Generate a professional PDF version of the quote. This function creates
    a beautifully formatted document that clients can save, print, and share.
    
    The PDF generation system is designed to be:
    - Professional and branded
    - Consistent across all quotes  
    - Customizable for different client types
    - Optimized for both screen and print viewing
    """
    
    if not PDF_AVAILABLE:
        messages.error(request, "PDF generation is not available. Please contact administrator.")
        return redirect('quotes:quote_detail', quote_id=quote_id)
    
    quote = get_object_or_404(Quote, id=quote_id)
    
    # Permission check - users can only view quotes they're involved with
    if not request.user.profile.is_admin:
        if quote.assigned_to != request.user and quote.created_by != request.user:
            messages.error(request, "You don't have permission to view this quote.")
            return redirect('quotes:quote_list')
    
    try:
        # Get quote items with related data for efficient rendering
        quote_items = quote.items.all().select_related('product', 'supplier')
        
        # Calculate additional metrics for the PDF
        total_cost = sum(
            item.unit_cost * item.quantity for item in quote_items 
            if item.unit_cost
        )
        profit_amount = quote.subtotal - total_cost
        profit_percentage = (
            (profit_amount / total_cost * 100) if total_cost > 0 else Decimal('0.00')
        )
        
        # Prepare context for the PDF template
        context = {
            'quote': quote,
            'quote_items': quote_items,
            'company_info': get_company_info(),
            'total_cost': total_cost,
            'profit_amount': profit_amount,
            'profit_percentage': profit_percentage,
            'generated_date': timezone.now(),
            'generated_by': request.user,
            # Add any custom branding or styling preferences
            'pdf_settings': get_pdf_settings(),
        }
        
        # Render the PDF template
        template = get_template('quotes/quote_pdf.html')
        html_content = template.render(context)
        
        # Configure fonts and styling for professional appearance
        font_config = FontConfiguration()
        
        # Generate the PDF with custom CSS for professional styling
        css_path = os.path.join(settings.STATIC_ROOT or settings.STATICFILES_DIRS[0], 'quotes/css/quote_pdf.css')
        css = CSS(filename=css_path, font_config=font_config) if os.path.exists(css_path) else None
        
        # Create the PDF document
        html_doc = HTML(string=html_content, base_url=request.build_absolute_uri())
        pdf_buffer = BytesIO()
        
        if css:
            html_doc.write_pdf(pdf_buffer, stylesheets=[css], font_config=font_config)
        else:
            html_doc.write_pdf(pdf_buffer, font_config=font_config)
        
        pdf_buffer.seek(0)
        
        # Create the HTTP response with appropriate headers
        filename = f"Quote_{quote.quote_number}_{quote.client.name.replace(' ', '_')}.pdf"
        response = HttpResponse(pdf_buffer.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response['Content-Length'] = pdf_buffer.tell()
        
        # Log the PDF generation for audit trail
        logger.info(f"PDF generated for quote {quote.quote_number} by {request.user.username}")
        
        # Create CRM interaction for tracking
        CustomerInteraction.objects.create(
            client=quote.client,
            interaction_type='quote',
            subject=f'Quote {quote.quote_number} PDF generated',
            notes=f'PDF document generated and downloaded by {request.user.get_full_name()}',
            created_by=request.user
        )
        
        return response
        
    except Exception as e:
        logger.error(f"Error generating PDF for quote {quote_id}: {str(e)}")
        messages.error(request, "Failed to generate PDF. Please try again or contact support.")
        return redirect('quotes:quote_detail', quote_id=quote_id)

def get_company_info():
    """
    Retrieve company information for PDF header/footer.
    This centralizes company branding information for consistency.
    """
    return {
        'name': getattr(settings, 'COMPANY_NAME', 'BlitzTech Electronics'),
        'address': getattr(settings, 'COMPANY_ADDRESS', 'Harare, Zimbabwe'),
        'phone': getattr(settings, 'COMPANY_PHONE', '+263 XX XXX XXXX'),
        'email': getattr(settings, 'COMPANY_EMAIL', 'info@blitztechelectronics.co.zw'),
        'website': getattr(settings, 'COMPANY_WEBSITE', 'www.blitztechelectronics.co.zw'),
        'logo_path': getattr(settings, 'COMPANY_LOGO_PATH', '/static/images/logo.png'),
        'tax_number': getattr(settings, 'COMPANY_TAX_NUMBER', 'TAX123456'),
    }

def get_pdf_settings():
    """
    Retrieve PDF-specific settings for customization.
    This allows for easy modification of PDF appearance and behavior.
    """
    return {
        'show_profit_analysis': getattr(settings, 'PDF_SHOW_PROFIT_ANALYSIS', False),
        'show_terms_and_conditions': getattr(settings, 'PDF_SHOW_TERMS', True),
        'watermark_text': getattr(settings, 'PDF_WATERMARK', None),
        'footer_text': getattr(settings, 'PDF_FOOTER_TEXT', 'Thank you for your business!'),
        'color_scheme': getattr(settings, 'PDF_COLOR_SCHEME', 'blue'),
    }


def email_quote_to_client(request, quote_id):
    """
    Email the quote directly to the client with a professional message.
    This function handles the complete email workflow including PDF attachment
    and client notification tracking.
    """
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'})
    
    quote = get_object_or_404(Quote, id=quote_id)
    
    # Permission check
    if not request.user.profile.is_admin:
        if quote.assigned_to != request.user and quote.created_by != request.user:
            return JsonResponse({'success': False, 'error': 'Permission denied'})
    
    try:
        # Generate PDF attachment
        pdf_buffer = generate_quote_pdf_buffer(quote, request)
        
        # Send email with PDF attachment
        email_success = send_quote_email(quote, pdf_buffer, request.user)
        
        if email_success:
            # Update quote status and tracking
            quote.status = 'sent'
            quote.sent_date = timezone.now()
            quote.save()
            
            # Create CRM interaction
            CustomerInteraction.objects.create(
                client=quote.client,
                interaction_type='quote_sent',
                subject=f'Quote {quote.quote_number} emailed to client',
                notes=f'Quote emailed to {quote.client.email} by {request.user.get_full_name()}',
                next_followup=timezone.now() + timezone.timedelta(days=3),
                created_by=request.user
            )
            
            # Send notification to assigned user if different
            if quote.assigned_to and quote.assigned_to != request.user:
                create_notification(
                    user=quote.assigned_to,
                    title="Quote Sent to Client",
                    message=f"Quote {quote.quote_number} has been sent to {quote.client.name}",
                    notification_type="info"
                )
            
            logger.info(f"Quote {quote.quote_number} emailed to {quote.client.email} by {request.user.username}")
            
            return JsonResponse({
                'success': True,
                'message': f'Quote sent successfully to {quote.client.email}'
            })
        else:
            return JsonResponse({
                'success': False,
                'error': 'Failed to send email. Please check email configuration.'
            })
            
    except Exception as e:
        logger.error(f"Error emailing quote {quote_id}: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': 'Failed to send email. Please try again.'
        })

def generate_quote_pdf_buffer(quote, request):
    """
    Generate PDF content in memory for email attachment.
    This is a helper function that creates the PDF without triggering a download.
    """
    
    quote_items = quote.items.all().select_related('product', 'supplier')
    
    context = {
        'quote': quote,
        'quote_items': quote_items,
        'company_info': get_company_info(),
        'generated_date': timezone.now(),
        'generated_by': request.user,
        'pdf_settings': get_pdf_settings(),
    }
    
    template = get_template('quotes/quote_pdf.html')
    html_content = template.render(context)
    
    # Generate PDF in memory
    font_config = FontConfiguration()
    html_doc = HTML(string=html_content, base_url=request.build_absolute_uri())
    pdf_buffer = BytesIO()
    html_doc.write_pdf(pdf_buffer, font_config=font_config)
    pdf_buffer.seek(0)
    
    return pdf_buffer

def send_quote_email(quote, pdf_buffer, sent_by_user):
    """
    Send professional email with quote PDF attachment.
    This function creates a branded, professional email that reinforces your company image.
    """
    
    from django.core.mail import EmailMessage
    from django.template.loader import render_to_string
    
    try:
        # Prepare email context with personalized content
        email_context = {
            'quote': quote,
            'client': quote.client,
            'company_info': get_company_info(),
            'sent_by': sent_by_user,
            'access_url': f"{settings.SITE_URL}/quotes/{quote.id}/preview/",
        }
        
        # Render professional email template
        subject = f"Quote {quote.quote_number} from {get_company_info()['name']}"
        
        # Create both HTML and text versions for better compatibility
        html_message = render_to_string('quotes/emails/quote_email.html', email_context)
        text_message = render_to_string('quotes/emails/quote_email.txt', email_context)
        
        # Create email message with professional styling
        email = EmailMessage(
            subject=subject,
            body=html_message,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[quote.client.email],
            cc=[sent_by_user.email] if sent_by_user.email else [],
            reply_to=[sent_by_user.email] if sent_by_user.email else [settings.DEFAULT_FROM_EMAIL],
        )
        
        # Set content type for HTML email
        email.content_subtype = 'html'
        
        # Attach PDF with professional filename
        filename = f"Quote_{quote.quote_number}_{quote.client.name.replace(' ', '_')}.pdf"
        email.attach(filename, pdf_buffer.read(), 'application/pdf')
        
        # Send the email
        email.send()
        
        return True
        
    except Exception as e:
        logger.error(f"Error sending quote email: {str(e)}")
        return False


@password_expiration_check
def quote_edit(request, quote_id):
    """
    Redirect to quote builder for editing quotes.
    This maintains consistency in the editing interface.
    """
    quote = get_object_or_404(Quote, id=quote_id)
    
    # Permission check
    if not request.user.profile.is_admin:
        if quote.assigned_to != request.user and quote.created_by != request.user:
            messages.error(request, "You don't have permission to edit this quote.")
            return redirect('quotes:quote_list')
    
    # Only allow editing of certain statuses
    if quote.status not in ['draft', 'sent']:
        messages.warning(request, f"Cannot edit quote with status: {quote.get_status_display()}")
        return redirect('quotes:quote_detail', quote_id=quote.id)
    
    return redirect('quotes:quote_builder', quote_id=quote.id)


@password_expiration_check
def quote_duplicate(request, quote_id):
    """
    Create a copy of an existing quote for quick creation of similar quotes.
    """
    original_quote = get_object_or_404(Quote, id=quote_id)
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Create new quote
                new_quote = Quote.objects.create(
                    quote_number=generate_quote_number(),
                    client=original_quote.client,
                    title=f"Copy of {original_quote.title}",
                    description=original_quote.description,
                    priority=original_quote.priority,
                    payment_terms=original_quote.payment_terms,
                    delivery_terms=original_quote.delivery_terms,
                    validity_date=timezone.now().date() + timezone.timedelta(days=30),
                    discount_percentage=original_quote.discount_percentage,
                    tax_rate=original_quote.tax_rate,
                    currency=original_quote.currency,
                    created_by=request.user,
                    assigned_to=request.user,
                    status='draft'
                )
                
                # Copy all items
                for item in original_quote.items.all():
                    QuoteItem.objects.create(
                        quote=new_quote,
                        product=item.product,
                        description=item.description,
                        detailed_specs=item.detailed_specs,
                        quantity=item.quantity,
                        unit_price=item.unit_price,
                        unit_cost=item.unit_cost,
                        markup_percentage=item.markup_percentage,
                        source_type=item.source_type,
                        supplier=item.supplier,
                        supplier_lead_time=item.supplier_lead_time,
                        estimated_delivery=item.estimated_delivery,
                        notes=item.notes,
                        sort_order=item.sort_order
                    )
                
                # Recalculate totals
                new_quote.calculate_totals()
                
                # Create CRM interaction
                CustomerInteraction.objects.create(
                    client=new_quote.client,
                    interaction_type='quote_draft',
                    subject=f'Quote {new_quote.quote_number} created (duplicated from {original_quote.quote_number})',
                    notes=f'Quote duplicated from {original_quote.quote_number}',
                    created_by=request.user
                )
                
                logger.info(f"Quote {original_quote.quote_number} duplicated as {new_quote.quote_number}")
                messages.success(request, f'Quote duplicated successfully as {new_quote.quote_number}')
                
                return JsonResponse({
                    'success': True,
                    'new_quote_id': new_quote.id,
                    'new_quote_number': new_quote.quote_number
                })
                
        except Exception as e:
            logger.error(f"Error duplicating quote {quote_id}: {str(e)}")
            return JsonResponse({'success': False, 'error': 'Failed to duplicate quote'})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})

@ajax_required

def send_quote(request, quote_id):
    """
    Send quote to client via email with professional presentation.
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'})
    
    quote = get_object_or_404(Quote, id=quote_id)
    
    # Permission check
    if not request.user.profile.is_admin:
        if quote.assigned_to != request.user and quote.created_by != request.user:
            return JsonResponse({'success': False, 'error': 'Permission denied'})
    
    # Validate quote can be sent
    if quote.status not in ['draft', 'sent']:
        return JsonResponse({
            'success': False, 
            'error': f'Cannot send quote with status: {quote.get_status_display()}'
        })
    
    if not quote.items.exists():
        return JsonResponse({
            'success': False,
            'error': 'Cannot send empty quote. Please add items first.'
        })
    
    # Get custom message if provided
    try:
        data = json.loads(request.body)
        custom_message = data.get('custom_message', '').strip()
    except (json.JSONDecodeError, AttributeError):
        custom_message = None
    
    # Send email
    result = send_quote_email(quote, request, custom_message)
    
    if result['success']:
        # Send notification to internal team
        send_quote_notification(quote, {
            'old_status': 'draft',
            'new_status': 'sent'
        })
        
        messages.success(request, f'Quote {quote.quote_number} sent successfully!')
    
    return JsonResponse(result)

@ajax_required

def update_quote_status(request, quote_id):
    """
    Update quote status with proper workflow validation.
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'})
    
    quote = get_object_or_404(Quote, id=quote_id)
    
    try:
        data = json.loads(request.body)
        new_status = data.get('status')
        
        if new_status not in dict(Quote.STATUS_CHOICES):
            return JsonResponse({'success': False, 'error': 'Invalid status'})
        
        old_status = quote.status
        
        # Validate status transition
        valid_transitions = {
            'draft': ['sent', 'cancelled'],
            'sent': ['viewed', 'under_review', 'accepted', 'rejected', 'expired'],
            'viewed': ['under_review', 'accepted', 'rejected', 'expired'],
            'under_review': ['accepted', 'rejected', 'expired'],
            'accepted': ['converted'],
            'rejected': [],  # Final state
            'expired': [],   # Final state
            'converted': [], # Final state
            'cancelled': []  # Final state
        }
        
        if new_status not in valid_transitions.get(old_status, []):
            return JsonResponse({
                'success': False,
                'error': f'Cannot change status from {old_status} to {new_status}'
            })
        
        # Update quote
        quote.status = new_status
        if new_status in ['accepted', 'rejected']:
            quote.response_date = timezone.now()
        quote.save()
        
        # Create CRM interaction
        CustomerInteraction.objects.create(
            client=quote.client,
            interaction_type='quote',
            subject=f'Quote {quote.quote_number} status changed to {new_status}',
            notes=f'Status changed from {old_status} to {new_status} by {request.user.get_full_name()}',
            created_by=request.user
        )
        
        # Send notifications
        send_quote_notification(quote, {
            'old_status': old_status,
            'new_status': new_status
        })
        
        logger.info(f"Quote {quote.quote_number} status changed from {old_status} to {new_status}")
        
        return JsonResponse({
            'success': True,
            'message': f'Quote status updated to {quote.get_status_display()}'
        })
        
    except Exception as e:
        logger.error(f"Error updating quote status: {str(e)}")
        return JsonResponse({'success': False, 'error': 'Failed to update status'})

@ajax_required

def quote_status_check(request, quote_id):
    """
    Check for quote status changes for real-time updates.
    """
    quote = get_object_or_404(Quote, id=quote_id)
    
    # Get last known status from request
    last_known_status = request.GET.get('last_status')
    
    status_changed = last_known_status and last_known_status != quote.status
    
    return JsonResponse({
        'current_status': quote.status,
        'status_display': quote.get_status_display(),
        'status_changed': status_changed,
        'last_updated': quote.updated_at.isoformat()
    })

@password_expiration_check
def approve_quote(request, quote_id):
    """
    Approve quotes that require management approval (high value, high discount).
    """
    quote = get_object_or_404(Quote, id=quote_id)
    
    if request.method == 'POST':
        try:
            quote.approved_by = request.user
            quote.save()
            
            # Create CRM interaction
            CustomerInteraction.objects.create(
                client=quote.client,
                interaction_type='quote',
                subject=f'Quote {quote.quote_number} approved by management',
                notes=f'Quote approved by {request.user.get_full_name()}',
                created_by=request.user
            )
            
            # Notify assigned user
            if quote.assigned_to and quote.assigned_to != request.user:
                create_notification(
                    user=quote.assigned_to,
                    title="Quote Approved",
                    message=f"Quote {quote.quote_number} has been approved and can now be sent.",
                    notification_type="success"
                )
            
            messages.success(request, f'Quote {quote.quote_number} approved successfully')
            return redirect('quotes:quote_detail', quote_id=quote.id)
            
        except Exception as e:
            logger.error(f"Error approving quote {quote_id}: {str(e)}")
            messages.error(request, 'Failed to approve quote')
    
    return redirect('quotes:quote_detail', quote_id=quote.id)

# =====================================
# ADDITIONAL HELPER VIEWS
# =====================================

@ajax_required

def generate_quote_number_ajax(request):
    """
    Generate a new quote number via AJAX.
    """
    try:
        quote_number = generate_quote_number()
        return JsonResponse({'success': True, 'quote_number': quote_number})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@ajax_required

def convert_currency(request):
    """
    Convert currency amounts for multi-currency quotes.
    """
    try:
        from_currency = request.GET.get('from', 'USD')
        to_currency = request.GET.get('to', 'ZWG')
        amount = Decimal(request.GET.get('amount', '0'))
        
        # Simple conversion - in production, use real exchange rates
        conversion_rates = {
            ('USD', 'ZWG'): Decimal('25.0'),  # Example rate
            ('ZWG', 'USD'): Decimal('0.04'),  # Example rate
        }
        
        rate = conversion_rates.get((from_currency, to_currency), Decimal('1.0'))
        converted_amount = amount * rate
        
        return JsonResponse({
            'success': True,
            'converted_amount': float(converted_amount),
            'rate': float(rate),
            'from_currency': from_currency,
            'to_currency': to_currency
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@ajax_required

def get_dashboard_stats(request):
    """
    Get real-time dashboard statistics for AJAX updates.
    """
    try:
        # Filter by user permissions
        user_filter = Q()
        if not request.user.profile.is_admin:
            user_filter = Q(assigned_to=request.user) | Q(created_by=request.user)
        
        # Calculate stats
        stats = {
            'total_quotes': Quote.objects.filter(user_filter).count(),
            'active_quotes': Quote.objects.filter(
                user_filter, 
                status__in=['draft', 'sent', 'viewed', 'under_review']
            ).count(),
            'quotes_needing_attention': Quote.objects.filter(
                user_filter,
                status='draft'
            ).count(),
            'expiring_soon': Quote.objects.filter(
                user_filter,
                status__in=['sent', 'viewed', 'under_review'],
                validity_date__lte=timezone.now().date() + timezone.timedelta(days=7)
            ).count(),
            'month_value': float(Quote.objects.filter(
                user_filter,
                created_at__gte=timezone.now().replace(day=1),
                status__in=['accepted', 'converted']
            ).aggregate(total=Sum('total_amount'))['total'] or 0)
        }
        
        return JsonResponse({'success': True, 'stats': stats})
        
    except Exception as e:
        logger.error(f"Error getting dashboard stats: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)})
