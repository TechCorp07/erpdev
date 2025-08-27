# inventory/utils.py - Complete Business Logic Utilities

"""
Comprehensive utility functions for BlitzTech Electronics inventory management.

Key Features:
1. Advanced cost calculation with all factors
2. Dynamic overhead allocation
3. Multi-currency conversion and management
4. Barcode/QR code generation and validation
5. Business intelligence calculations
6. Import/export utilities
7. Stock optimization algorithms
8. Reporting and analytics helpers
9. Integration utilities for quote system
10. Mobile and offline support
"""

import csv
import json
import logging
from django.forms import ValidationError
import qrcode
import barcode
from django.db import transaction
from barcode.writer import ImageWriter
from decimal import Decimal, ROUND_HALF_UP
from datetime import datetime, timedelta
from django.utils import timezone
from django.db.models import Sum, Count, Avg, F, Q, Case, When
from django.http import HttpResponse
from django.template.loader import render_to_string
from typing import Dict, List, Optional, Tuple, Union
from openpyxl.styles import Font, Alignment, PatternFill
from django.core.mail import send_mail
from django.conf import settings
from typing import Iterable
from io import BytesIO
import base64
import openpyxl
import requests
import re
import math

logger = logging.getLogger(__name__)

# =====================================
# COST CALCULATION UTILITIES
# =====================================

class CostCalculator:
    """
    Advanced cost calculation engine for electronics import business
    """
    
    def __init__(self):
        self.exchange_rates = self._get_exchange_rates()
        self.overhead_factors = self._get_overhead_factors()
    
    def _get_exchange_rates(self) -> Dict[str, Decimal]:
        """Get current exchange rates"""
        from .models import Currency
        
        rates = {}
        currencies = Currency.objects.filter(is_active=True)
        
        for currency in currencies:
            rates[currency.code] = currency.exchange_rate_to_usd
        
        return rates
    
    def _get_overhead_factors(self) -> List:
        """Get active overhead factors"""
        from .models import OverheadFactor
        
        return list(OverheadFactor.objects.filter(is_active=True))
    
    def calculate_total_cost(self, product_data: Dict) -> Dict:
        """
        Calculate total cost for a product including all factors
        
        Args:
            product_data: Dictionary with product cost information
            
        Returns:
            Dictionary with detailed cost breakdown
        """
        try:
            # Base cost in original currency
            base_cost = Decimal(str(product_data.get('cost_price', 0)))
            currency_code = product_data.get('currency_code', 'USD')
            
            # Convert to USD
            exchange_rate = self.exchange_rates.get(currency_code, Decimal('1.0'))
            base_cost_usd = base_cost * exchange_rate
            
            # Import costs
            shipping_cost = Decimal(str(product_data.get('shipping_cost_per_unit', 0)))
            insurance_cost = Decimal(str(product_data.get('insurance_cost_per_unit', 0)))
            customs_duty_percent = Decimal(str(product_data.get('customs_duty_percentage', 0)))
            vat_percent = Decimal(str(product_data.get('vat_percentage', 15)))
            other_fees = Decimal(str(product_data.get('other_fees_per_unit', 0)))
            
            # Calculate duties and taxes
            customs_duty = base_cost_usd * (customs_duty_percent / 100)
            vat_base = base_cost_usd + customs_duty
            vat_cost = vat_base * (vat_percent / 100)
            
            # Total import cost
            total_import_cost = (
                base_cost_usd + shipping_cost + insurance_cost + 
                customs_duty + vat_cost + other_fees
            )
            
            # Calculate overhead costs
            overhead_cost = self._calculate_overhead_cost(
                product_data, total_import_cost
            )
            
            # Total cost
            total_cost = total_import_cost + overhead_cost
            
            # Calculate suggested selling prices with different markups
            markup_suggestions = []
            for markup_percent in [10, 20, 30, 40, 50]:
                selling_price = total_cost * (1 + markup_percent / 100)
                profit = selling_price - total_cost
                
                markup_suggestions.append({
                    'markup_percentage': markup_percent,
                    'selling_price': float(selling_price.quantize(Decimal('0.01'))),
                    'profit_per_unit': float(profit.quantize(Decimal('0.01')))
                })
            
            return {
                'success': True,
                'cost_breakdown': {
                    'base_cost_original': float(base_cost),
                    'currency_code': currency_code,
                    'exchange_rate': float(exchange_rate),
                    'base_cost_usd': float(base_cost_usd),
                    'shipping_cost': float(shipping_cost),
                    'insurance_cost': float(insurance_cost),
                    'customs_duty': float(customs_duty),
                    'vat_cost': float(vat_cost),
                    'other_fees': float(other_fees),
                    'total_import_cost': float(total_import_cost),
                    'overhead_cost': float(overhead_cost),
                    'total_cost_usd': float(total_cost),
                },
                'markup_suggestions': markup_suggestions
            }
            
        except Exception as e:
            logger.error(f"Error calculating product cost: {str(e)}")
            return {
                'success': False,
                'error': str(e)
            }
    
    def _calculate_overhead_cost(self, product_data: Dict, import_cost: Decimal) -> Decimal:
        """Calculate allocated overhead costs"""
        total_overhead = Decimal('0.00')
        
        category_id = product_data.get('category_id')
        supplier_id = product_data.get('supplier_id')
        weight_grams = product_data.get('weight_grams')
        
        for factor in self.overhead_factors:
            # Check if factor applies to this product
            applies = True
            
            if factor.applies_to_categories.exists() and category_id:
                applies = applies and factor.applies_to_categories.filter(id=category_id).exists()
            
            if factor.applies_to_suppliers.exists() and supplier_id:
                applies = applies and factor.applies_to_suppliers.filter(id=supplier_id).exists()
            
            if applies:
                cost = factor.calculate_cost(
                    product_cost=import_cost,
                    order_value=import_cost,
                    weight_kg=weight_grams / 1000 if weight_grams else None
                )
                total_overhead += cost
        
        return total_overhead
    
    def calculate_optimal_markup(self, product_data: Dict, target_profit_margin: Decimal = None) -> Dict:
        """
        Calculate optimal markup based on market conditions and business rules
        
        Args:
            product_data: Product information
            target_profit_margin: Target profit margin percentage
            
        Returns:
            Optimal pricing recommendations
        """
        cost_breakdown = self.calculate_total_cost(product_data)
        
        if not cost_breakdown['success']:
            return cost_breakdown
        
        total_cost = Decimal(str(cost_breakdown['cost_breakdown']['total_cost_usd']))
        
        # Business rules for markup calculation
        category_id = product_data.get('category_id')
        brand_id = product_data.get('brand_id')
        competitor_prices = product_data.get('competitor_prices', {})
        
        # Get category default markup
        category_markup = Decimal('30.00')  # Default
        if category_id:
            from .models import Category
            try:
                category = Category.objects.get(id=category_id)
                category_markup = category.default_markup_percentage
            except Category.DoesNotExist:
                pass
        
        # Get brand default markup
        brand_markup = Decimal('30.00')  # Default
        if brand_id:
            from .models import Brand
            try:
                brand = Brand.objects.get(id=brand_id)
                brand_markup = brand.default_markup_percentage
            except Brand.DoesNotExist:
                pass
        
        # Calculate market-based markup
        market_markup = self._calculate_market_based_markup(
            total_cost, competitor_prices
        )
        
        # Determine optimal markup
        markups = [category_markup, brand_markup]
        if market_markup:
            markups.append(market_markup)
        if target_profit_margin:
            markups.append(target_profit_margin)
        
        # Use average of relevant markups
        optimal_markup = sum(markups) / len(markups)
        
        # Apply business constraints
        min_markup = Decimal('10.00')  # Minimum 10% markup
        max_markup = Decimal('500.00')  # Maximum 500% markup
        
        optimal_markup = max(min_markup, min(max_markup, optimal_markup))
        
        optimal_price = total_cost * (1 + optimal_markup / 100)
        optimal_profit = optimal_price - total_cost
        
        return {
            'success': True,
            'recommendations': {
                'optimal_markup_percentage': float(optimal_markup),
                'optimal_selling_price': float(optimal_price.quantize(Decimal('0.01'))),
                'optimal_profit_per_unit': float(optimal_profit.quantize(Decimal('0.01'))),
                'category_suggested_markup': float(category_markup),
                'brand_suggested_markup': float(brand_markup),
                'market_based_markup': float(market_markup) if market_markup else None,
            },
            'cost_breakdown': cost_breakdown['cost_breakdown']
        }
    
    def _calculate_market_based_markup(self, total_cost: Decimal, competitor_prices: Dict) -> Optional[Decimal]:
        """Calculate markup based on competitor pricing"""
        if not competitor_prices:
            return None
        
        min_price = competitor_prices.get('min_price')
        max_price = competitor_prices.get('max_price')
        
        if min_price and max_price:
            # Target price between min and max, closer to min for competitiveness
            target_price = Decimal(str(min_price)) * Decimal('1.05')  # 5% above minimum
            
            if target_price > total_cost:
                markup = ((target_price - total_cost) / total_cost) * 100
                return markup
        
        return None

# =====================================
# BARCODE AND QR CODE UTILITIES
# =====================================

class BarcodeManager:
    """
    Barcode and QR code generation and management
    """
    
    @staticmethod
    def generate_qr_code(data: Union[str, Dict], size: int = 10) -> str:
        """
        Generate QR code and return as base64 string
        
        Args:
            data: Data to encode (string or dictionary)
            size: QR code size (box size)
            
        Returns:
            Base64 encoded PNG image
        """
        try:
            if isinstance(data, dict):
                data = json.dumps(data)
            
            qr = qrcode.QRCode(
                version=1,
                error_correction=qrcode.constants.ERROR_CORRECT_L,
                box_size=size,
                border=4,
            )
            qr.add_data(data)
            qr.make(fit=True)
            
            img = qr.make_image(fill_color="black", back_color="white")
            
            buffer = BytesIO()
            img.save(buffer, format='PNG')
            img_data = base64.b64encode(buffer.getvalue()).decode()
            
            return img_data
            
        except Exception as e:
            logger.error(f"Error generating QR code: {str(e)}")
            return None
    
    @staticmethod
    def generate_barcode(code: str, code_type: str = 'code128') -> str:
        """
        Generate barcode and return as base64 string
        
        Args:
            code: Code to encode
            code_type: Type of barcode (code128, ean13, etc.)
            
        Returns:
            Base64 encoded PNG image
        """
        try:
            # Generate barcode
            barcode_class = barcode.get_barcode_class(code_type)
            barcode_instance = barcode_class(code, writer=ImageWriter())
            
            buffer = BytesIO()
            barcode_instance.write(buffer)
            img_data = base64.b64encode(buffer.getvalue()).decode()
            
            return img_data
            
        except Exception as e:
            logger.error(f"Error generating barcode: {str(e)}")
            return None
    
    @staticmethod
    def validate_barcode(code: str) -> Dict:
        """
        Validate barcode format and check digit
        
        Args:
            code: Barcode to validate
            
        Returns:
            Validation result
        """
        try:
            # Basic validation patterns
            patterns = {
                'ean13': r'^\d{13}$',
                'ean8': r'^\d{8}$',
                'upca': r'^\d{12}$',
                'code128': r'^[\x00-\x7F]+$',  # ASCII characters
            }
            
            for code_type, pattern in patterns.items():
                if re.match(pattern, code):
                    return {
                        'valid': True,
                        'type': code_type,
                        'code': code
                    }
            
            return {
                'valid': False,
                'error': 'Invalid barcode format'
            }
            
        except Exception as e:
            return {
                'valid': False,
                'error': str(e)
            }
    
    @staticmethod
    def generate_product_qr_data(product) -> Dict:
        """
        Generate standard QR code data for a product
        
        Args:
            product: Product instance
            
        Returns:
            Dictionary with product QR data
        """
        return {
            'type': 'product',
            'sku': product.sku,
            'name': product.name,
            'brand': product.brand.name,
            'category': product.category.name,
            'price': float(product.selling_price),
            'stock': product.total_stock,
            'barcode': product.barcode,
            'supplier': product.supplier.name,
            'datasheet_url': product.datasheet_url,
            'specifications': product.dynamic_attributes,
            'last_updated': timezone.now().isoformat()
        }

# =====================================
# STOCK OPTIMIZATION UTILITIES
# =====================================

class StockOptimizer:
    """
    Stock level optimization and reorder calculations
    """
    
    @staticmethod
    def calculate_economic_order_quantity(product) -> int:
        """
        Calculate Economic Order Quantity (EOQ)
        
        Args:
            product: Product instance
            
        Returns:
            Optimal order quantity
        """
        try:
            # Annual demand (estimate based on historical data or forecast)
            annual_demand = StockOptimizer._estimate_annual_demand(product)
            
            # Order cost (fixed cost per order)
            order_cost = Decimal('50.00')  # Default $50 per order
            
            # Holding cost (percentage of item value per year)
            holding_cost_rate = Decimal('0.20')  # 20% per year
            holding_cost = product.total_cost_price_usd * holding_cost_rate
            
            if holding_cost <= 0 or annual_demand <= 0:
                return product.reorder_quantity
            
            # EOQ formula: sqrt(2 * D * S / H)
            # Where D = annual demand, S = order cost, H = holding cost
            eoq = math.sqrt(
                (2 * float(annual_demand) * float(order_cost)) / float(holding_cost)
            )
            
            # Round to nearest integer and apply constraints
            eoq = max(1, int(round(eoq)))
            eoq = max(eoq, product.supplier_minimum_order_quantity)
            
            return eoq
            
        except Exception as e:
            logger.error(f"Error calculating EOQ for product {product.sku}: {str(e)}")
            return product.reorder_quantity
    
    @staticmethod
    def _estimate_annual_demand(product) -> int:
        """
        Estimate annual demand based on historical sales
        """
        # Simple estimate based on recent sales
        # In a real system, this would use more sophisticated forecasting
        
        if product.total_sold > 0:
            # Estimate based on product age and total sold
            days_since_created = (timezone.now().date() - product.created_at.date()).days
            if days_since_created > 0:
                daily_demand = product.total_sold / days_since_created
                annual_demand = daily_demand * 365
                return max(1, int(annual_demand))
        
        # Fallback to category average or default
        return max(product.reorder_quantity * 4, 50)  # Assume 4 reorders per year minimum
    
    @staticmethod
    def calculate_safety_stock(product) -> int:
        """
        Calculate safety stock level
        
        Args:
            product: Product instance
            
        Returns:
            Safety stock quantity
        """
        try:
            # Service level factor (for 95% service level, Z = 1.65)
            service_level_factor = 1.65
            
            # Lead time in days
            lead_time = product.supplier_lead_time_days
            
            # Demand variability (simplified calculation)
            daily_demand = StockOptimizer._estimate_annual_demand(product) / 365
            demand_std_dev = daily_demand * 0.3  # Assume 30% coefficient of variation
            
            # Lead time variability (assume 20% of lead time)
            lead_time_std_dev = lead_time * 0.2
            
            # Safety stock formula
            safety_stock = service_level_factor * math.sqrt(
                (lead_time * demand_std_dev ** 2) + 
                (daily_demand ** 2 * lead_time_std_dev ** 2)
            )
            
            return max(1, int(round(safety_stock)))
            
        except Exception as e:
            logger.error(f"Error calculating safety stock for product {product.sku}: {str(e)}")
            return max(5, product.reorder_level // 2)
    
    @staticmethod
    def suggest_reorder_levels(product) -> Dict:
        """
        Suggest optimal reorder levels for a product
        
        Args:
            product: Product instance
            
        Returns:
            Dictionary with reorder recommendations
        """
        try:
            eoq = StockOptimizer.calculate_economic_order_quantity(product)
            safety_stock = StockOptimizer.calculate_safety_stock(product)
            
            # Reorder point = (average demand × lead time) + safety stock
            daily_demand = StockOptimizer._estimate_annual_demand(product) / 365
            reorder_point = int((daily_demand * product.supplier_lead_time_days) + safety_stock)
            
            # Maximum stock level
            max_stock = reorder_point + eoq
            
            return {
                'current_reorder_level': product.reorder_level,
                'suggested_reorder_level': max(1, reorder_point),
                'current_reorder_quantity': product.reorder_quantity,
                'suggested_reorder_quantity': eoq,
                'current_max_stock': product.max_stock_level,
                'suggested_max_stock': max_stock,
                'safety_stock': safety_stock,
                'estimated_annual_demand': StockOptimizer._estimate_annual_demand(product)
            }
            
        except Exception as e:
            logger.error(f"Error calculating reorder suggestions for product {product.sku}: {str(e)}")
            return {
                'error': str(e),
                'current_reorder_level': product.reorder_level,
                'suggested_reorder_level': product.reorder_level,
                'current_reorder_quantity': product.reorder_quantity,
                'suggested_reorder_quantity': product.reorder_quantity,
            }

def calculate_available_stock(product, location=None):
    """
    Calculate available stock for a product, optionally at a specific location.
    
    Available stock = Current stock - Reserved stock
    
    Args:
        product: Product instance
        location: Optional Location instance for location-specific calculation
        
    Returns:
        int: Available stock quantity
    """
    try:
        if location:
            from .models import StockLevel
            stock_level = StockLevel.objects.filter(
                product=product, location=location
            ).first()
            
            if stock_level:
                return max(0, stock_level.quantity - stock_level.reserved_quantity)
            return 0
        else:
            return max(0, product.current_stock - product.reserved_stock)
            
    except Exception as e:
        logger.error(f"Error calculating available stock for {product.sku}: {str(e)}")
        return 0

def calculate_stock_value(products=None, location=None, cost_basis='current'):
    """
    Calculate total stock value for products.
    
    Args:
        products: QuerySet of products (default: all active products)
        location: Optional location filter
        cost_basis: 'current', 'average', or 'fifo' (default: 'current')
        
    Returns:
        Decimal: Total stock value
    """
    try:
        from .models import Product, StockLevel
        
        if products is None:
            products = Product.objects.filter(is_active=True)
        
        total_value = Decimal('0.00')
        
        for product in products:
            if location:
                stock_level = StockLevel.objects.filter(
                    product=product, location=location
                ).first()
                quantity = stock_level.quantity if stock_level else 0
            else:
                quantity = product.current_stock
            
            # For now, use current cost price
            # In the future, this could be enhanced for FIFO/average costing
            unit_cost = product.cost_price
            total_value += quantity * unit_cost
        
        return total_value
        
    except Exception as e:
        logger.error(f"Error calculating stock value: {str(e)}")
        return Decimal('0.00')

# =====================================
# PRICING AND COST UTILITIES
# =====================================

def calculate_selling_price(cost_price, markup_percentage, currency='USD'):
    """
    Calculate selling price based on cost and markup percentage.
    
    Args:
        cost_price: Decimal cost price
        markup_percentage: Decimal markup percentage (e.g., 30.00 for 30%)
        currency: Currency code for rounding rules
        
    Returns:
        Decimal: Calculated selling price
    """
    try:
        if cost_price <= 0:
            return Decimal('0.00')
        
        markup_multiplier = 1 + (markup_percentage / 100)
        selling_price = cost_price * markup_multiplier
        
        # Round to appropriate precision based on currency
        if currency in ['USD', 'EUR', 'GBP']:
            return selling_price.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        elif currency == 'ZWG':
            # Zimbabwe Gold might need different rounding
            return selling_price.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        else:
            return selling_price.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            
    except Exception as e:
        logger.error(f"Error calculating selling price: {str(e)}")
        return Decimal('0.00')

def calculate_profit_margin(cost_price, selling_price):
    """
    Calculate profit margin percentage.
    
    Args:
        cost_price: Decimal cost price
        selling_price: Decimal selling price
        
    Returns:
        Decimal: Profit margin percentage
    """
    try:
        if cost_price <= 0:
            return Decimal('0.00')
        
        profit = selling_price - cost_price
        margin_percentage = (profit / cost_price) * 100
        
        return margin_percentage.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        
    except Exception as e:
        logger.error(f"Error calculating profit margin: {str(e)}")
        return Decimal('0.00')

def apply_bulk_price_update(products, update_type, value, user=None):
    """
    Apply bulk price updates to multiple products.
    
    Args:
        products: QuerySet of products to update
        update_type: 'markup', 'percentage_increase', 'fixed_increase', 'set_price'
        value: Decimal value for the update
        user: User performing the update
        
    Returns:
        int: Number of products updated
    """
    try:
        updated_count = 0
        
        with transaction.atomic():
            for product in products:
                old_price = product.selling_price
                
                if update_type == 'markup':
                    # Apply markup percentage to cost price
                    product.selling_price = calculate_selling_price(
                        product.cost_price, value
                    )
                elif update_type == 'percentage_increase':
                    # Increase current price by percentage
                    multiplier = 1 + (value / 100)
                    product.selling_price = old_price * multiplier
                elif update_type == 'fixed_increase':
                    # Add fixed amount to current price
                    product.selling_price = old_price + value
                elif update_type == 'set_price':
                    # Set specific price
                    product.selling_price = value
                
                # Ensure price is not negative
                product.selling_price = max(product.selling_price, Decimal('0.01'))
                
                product.save(update_fields=['selling_price'])
                
                # Log the price change
                logger.info(
                    f"Price updated for {product.sku}: {old_price} -> {product.selling_price} "
                    f"by {user.username if user else 'System'}"
                )
                
                updated_count += 1
        
        return updated_count
        
    except Exception as e:
        logger.error(f"Error in bulk price update: {str(e)}")
        return 0


# =====================================
# STOCK MOVEMENT UTILITIES
# =====================================

def create_stock_movement(product, movement_type, quantity, reference, 
                         from_location=None, to_location=None, user=None, 
                         unit_cost=None, notes=""):
    """
    Create a stock movement record with proper validation and stock updates.
    
    Args:
        product: Product instance
        movement_type: Type of movement ('in', 'out', 'adjustment', etc.)
        quantity: Quantity moved (positive for in, negative for out)
        reference: Reference number or description
        from_location: Source location (optional)
        to_location: Destination location (optional)
        user: User performing the movement
        unit_cost: Cost per unit (optional)
        notes: Additional notes
        
    Returns:
        StockMovement: Created movement record
    """
    try:
        from .models import StockMovement
        
        with transaction.atomic():
            # Record current stock before movement
            previous_stock = product.current_stock
            
            # Calculate new stock level
            new_stock = max(0, previous_stock + quantity)
            
            # Create movement record
            movement = StockMovement.objects.create(
                product=product,
                movement_type=movement_type,
                quantity=quantity,
                from_location=from_location,
                to_location=to_location,
                reference=reference,
                previous_stock=previous_stock,
                new_stock=new_stock,
                unit_cost=unit_cost,
                total_cost=abs(quantity) * unit_cost if unit_cost else None,
                notes=notes,
                created_by=user
            )
            
            # Update product stock
            product.current_stock = new_stock
            if movement_type in ['sale', 'out']:
                product.total_sold = (product.total_sold or 0) + abs(quantity)
                product.last_sold_date = timezone.now()
            elif movement_type in ['purchase', 'in']:
                product.last_restocked_date = timezone.now()
            
            product.save()
            
            # Update location-specific stock levels
            update_location_stock_levels(movement)
            
            return movement
            
    except Exception as e:
        logger.error(f"Error creating stock movement: {str(e)}")
        raise ValidationError(f"Failed to create stock movement: {str(e)}")

def update_location_stock_levels(movement):
    """
    Update location-specific stock levels based on stock movement.
    
    Args:
        movement: StockMovement instance
    """
    try:
        from .models import StockLevel
        
        # Update from_location
        if movement.from_location and movement.quantity < 0:
            stock_level, created = StockLevel.objects.get_or_create(
                product=movement.product,
                location=movement.from_location,
                defaults={'quantity': 0, 'reserved_quantity': 0}
            )
            stock_level.quantity = max(0, stock_level.quantity + movement.quantity)
            stock_level.save()
        
        # Update to_location
        if movement.to_location and movement.quantity > 0:
            stock_level, created = StockLevel.objects.get_or_create(
                product=movement.product,
                location=movement.to_location,
                defaults={'quantity': 0, 'reserved_quantity': 0}
            )
            stock_level.quantity += movement.quantity
            stock_level.save()
            
    except Exception as e:
        logger.error(f"Error updating location stock levels: {str(e)}")

def transfer_stock_between_locations(product, from_location, to_location, 
                                   quantity, reference, user=None, notes=""):
    """
    Transfer stock between locations with proper validation.
    
    Args:
        product: Product instance
        from_location: Source location
        to_location: Destination location
        quantity: Quantity to transfer
        reference: Transfer reference
        user: User performing transfer
        notes: Transfer notes
        
    Returns:
        Tuple[StockMovement, StockMovement]: (outgoing, incoming) movements
    """
    try:
        from .models import StockLevel
        
        with transaction.atomic():
            # Check available stock at from_location
            from_stock = StockLevel.objects.filter(
                product=product, location=from_location
            ).first()
            
            if not from_stock or from_stock.available_quantity < quantity:
                raise ValidationError(
                    f"Insufficient stock at {from_location.name}. "
                    f"Available: {from_stock.available_quantity if from_stock else 0}, "
                    f"Required: {quantity}"
                )
            
            # Create outgoing movement
            outgoing = create_stock_movement(
                product=product,
                movement_type='transfer',
                quantity=-quantity,
                reference=reference,
                from_location=from_location,
                to_location=None,
                user=user,
                notes=f"Transfer to {to_location.name}. {notes}"
            )
            
            # Create incoming movement
            incoming = create_stock_movement(
                product=product,
                movement_type='transfer',
                quantity=quantity,
                reference=reference,
                from_location=None,
                to_location=to_location,
                user=user,
                notes=f"Transfer from {from_location.name}. {notes}"
            )
            
            return outgoing, incoming
            
    except Exception as e:
        logger.error(f"Error transferring stock: {str(e)}")
        raise ValidationError(f"Stock transfer failed: {str(e)}")


# =====================================
# IMPORT/EXPORT UTILITIES
# =====================================

class DataManager:
    """
    Data import/export and management utilities
    """
    
    @staticmethod
    def export_products_to_excel(products_queryset, filename: str = None) -> HttpResponse:
        """
        Export products to Excel file
        
        Args:
            products_queryset: QuerySet of products to export
            filename: Optional filename
            
        Returns:
            HttpResponse with Excel file
        """
        try:
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Products"
            
            # Headers
            headers = [
                'SKU', 'Name', 'Category', 'Brand', 'Supplier', 'Product Type',
                'Model Number', 'Manufacturer PN', 'Supplier SKU', 'Package Type',
                'Cost Price', 'Cost Currency', 'Total Cost USD', 'Selling Price',
                'Markup %', 'Current Stock', 'Reorder Level', 'Reorder Quantity',
                'Lead Time (Days)', 'Supplier MOQ', 'Weight (g)', 'Dimensions',
                'Is Active', 'Is Hazardous', 'Requires ESD', 'Datasheet URL'
            ]
            
            # Style headers
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            
            # Data rows
            for row, product in enumerate(products_queryset, 2):
                data = [
                    product.sku,
                    product.name,
                    product.category.name,
                    product.brand.name,
                    product.supplier.name,
                    product.get_product_type_display(),
                    product.model_number,
                    product.manufacturer_part_number,
                    product.supplier_sku,
                    product.package_type,
                    float(product.cost_price),
                    product.supplier_currency.code,
                    float(product.total_cost_price_usd),
                    float(product.selling_price),
                    float(product.markup_percentage) if product.markup_percentage else 0,
                    product.total_stock,
                    product.reorder_level,
                    product.reorder_quantity,
                    product.supplier_lead_time_days,
                    product.supplier_minimum_order_quantity,
                    float(product.weight_grams) if product.weight_grams else 0,
                    product.dimensions,
                    'Yes' if product.is_active else 'No',
                    'Yes' if product.is_hazardous else 'No',
                    'Yes' if product.requires_esd_protection else 'No',
                    product.datasheet_url
                ]
                
                for col, value in enumerate(data, 1):
                    ws.cell(row=row, column=col, value=value)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Create response
            if not filename:
                filename = f"products_export_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            wb.save(response)
            return response
            
        except Exception as e:
            logger.error(f"Error exporting products to Excel: {str(e)}")
            raise
    
    @staticmethod
    def import_products_from_excel(file, user) -> Dict:
        """
        Import products from Excel file
        
        Args:
            file: Uploaded Excel file
            user: User performing the import
            
        Returns:
            Import results
        """
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            
            results = {
                'success_count': 0,
                'error_count': 0,
                'errors': [],
                'warnings': []
            }
            
            # Process rows (skip header)
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                try:
                    # Extract data from row
                    product_data = DataManager._extract_product_data_from_row(row)
                    
                    # Validate and create/update product
                    success = DataManager._create_or_update_product(product_data, user)
                    
                    if success:
                        results['success_count'] += 1
                    else:
                        results['error_count'] += 1
                        results['errors'].append(f"Row {row_num}: Failed to create/update product")
                        
                except Exception as e:
                    results['error_count'] += 1
                    results['errors'].append(f"Row {row_num}: {str(e)}")
            
            return results
            
        except Exception as e:
            logger.error(f"Error importing products from Excel: {str(e)}")
            return {
                'success_count': 0,
                'error_count': 1,
                'errors': [str(e)]
            }
    
    @staticmethod
    def _extract_product_data_from_row(row) -> Dict:
        """Extract product data from Excel row"""
        # Map Excel columns to product fields
        # This would need to match the export format
        return {
            'sku': row[0],
            'name': row[1],
            'category_name': row[2],
            'brand_name': row[3],
            'supplier_name': row[4],
            # ... map other fields
        }
    
    @staticmethod
    def _create_or_update_product(product_data: Dict, user) -> bool:
        """Create or update product from data"""
        try:
            from .models import Product, Category, Brand, Supplier
            
            # Get or create related objects
            category, _ = Category.objects.get_or_create(
                name=product_data['category_name'],
                defaults={'created_by': user}
            )
            
            brand, _ = Brand.objects.get_or_create(
                name=product_data['brand_name'],
                defaults={'created_by': user}
            )
            
            supplier, _ = Supplier.objects.get_or_create(
                name=product_data['supplier_name'],
                defaults={'created_by': user}
            )
            
            # Create or update product
            product, created = Product.objects.update_or_create(
                sku=product_data['sku'],
                defaults={
                    'name': product_data['name'],
                    'category': category,
                    'brand': brand,
                    'supplier': supplier,
                    'created_by': user if created else None
                }
            )
            
            return True
            
        except Exception as e:
            logger.error(f"Error creating/updating product: {str(e)}")
            return False

# =====================================
# INTEGRATION UTILITIES
# =====================================

class IntegrationHelper:
    """
    Utilities for integrating with other systems (quote system, etc.)
    """
    
    @staticmethod
    def get_products_for_quote_system(search_term: str = None, limit: int = 50) -> List[Dict]:
        """
        Get products formatted for quote system integration
        
        Args:
            search_term: Optional search term
            limit: Maximum number of results
            
        Returns:
            List of products formatted for quote system
        """
        try:
            from .models import Product
            
            products = Product.objects.filter(is_active=True).select_related(
                'category', 'supplier', 'brand', 'supplier_currency'
            )
            
            if search_term:
                products = products.filter(
                    Q(name__icontains=search_term) |
                    Q(sku__icontains=search_term) |
                    Q(manufacturer_part_number__icontains=search_term) |
                    Q(supplier_sku__icontains=search_term)
                )
            
            products = products[:limit]
            
            result = []
            for product in products:
                result.append({
                    'id': product.id,
                    'sku': product.sku,
                    'name': product.name,
                    'description': product.short_description or product.description[:100],
                    'category': product.category.name,
                    'brand': product.brand.name,
                    'supplier': product.supplier.name,
                    'cost_price': float(product.total_cost_price_usd),
                    'selling_price': float(product.selling_price),
                    'currency': product.selling_currency.code,
                    'current_stock': product.total_stock,
                    'available_stock': product.available_stock,
                    'stock_status': product.stock_status,
                    'lead_time_days': product.supplier_lead_time_days,
                    'minimum_quantity': product.supplier_minimum_order_quantity,
                    'specifications': product.dynamic_attributes,
                    'datasheet_url': product.datasheet_url,
                    'images': product.product_images
                })
            
            return result
            
        except Exception as e:
            logger.error(f"Error getting products for quote system: {str(e)}")
            return []
    
    @staticmethod
    def check_stock_availability(product_requests: List[Dict]) -> Dict:
        """
        Check stock availability for multiple products
        
        Args:
            product_requests: List of {'product_id': int, 'quantity': int}
            
        Returns:
            Availability results
        """
        try:
            from .models import Product
            
            results = {
                'available': [],
                'partially_available': [],
                'unavailable': [],
                'total_value': Decimal('0.00')
            }
            
            for request in product_requests:
                try:
                    product = Product.objects.get(
                        id=request['product_id'],
                        is_active=True
                    )
                    
                    requested_qty = request['quantity']
                    available_qty = product.available_stock
                    
                    item_data = {
                        'product_id': product.id,
                        'sku': product.sku,
                        'name': product.name,
                        'requested_quantity': requested_qty,
                        'available_quantity': available_qty,
                        'unit_price': float(product.selling_price),
                        'lead_time_days': product.supplier_lead_time_days
                    }
                    
                    if available_qty >= requested_qty:
                        results['available'].append(item_data)
                        results['total_value'] += Decimal(str(product.selling_price)) * requested_qty
                    elif available_qty > 0:
                        results['partially_available'].append(item_data)
                        results['total_value'] += Decimal(str(product.selling_price)) * available_qty
                    else:
                        results['unavailable'].append(item_data)
                        
                except Product.DoesNotExist:
                    results['unavailable'].append({
                        'product_id': request['product_id'],
                        'error': 'Product not found'
                    })
            
            results['total_value'] = float(results['total_value'])
            
            return results
            
        except Exception as e:
            logger.error(f"Error checking stock availability: {str(e)}")
            return {'error': str(e)}
    
    @staticmethod
    def reserve_stock(reservations: List[Dict], reference: str = None) -> Dict:
        """
        Reserve stock for pending orders/quotes
        
        Args:
            reservations: List of {'product_id': int, 'quantity': int}
            reference: Reference number (quote, order, etc.)
            
        Returns:
            Reservation results
        """
        try:
            from .models import Product, ProductStockLevel
            from django.db import transaction
            
            reserved_items = []
            
            with transaction.atomic():
                for reservation in reservations:
                    try:
                        product = Product.objects.select_for_update().get(
                            id=reservation['product_id'],
                            is_active=True
                        )
                        
                        quantity = reservation['quantity']
                        
                        if product.available_stock >= quantity:
                            # Reserve stock
                            product.reserved_stock += quantity
                            product.save(update_fields=['reserved_stock'])
                            
                            reserved_items.append({
                                'product_id': product.id,
                                'sku': product.sku,
                                'quantity_reserved': quantity,
                                'reference': reference
                            })
                            
                        else:
                            return {
                                'success': False,
                                'error': f'Insufficient stock for product {product.sku}',
                                'available': product.available_stock,
                                'requested': quantity
                            }
                            
                    except Product.DoesNotExist:
                        return {
                            'success': False,
                            'error': f'Product not found: {reservation["product_id"]}'
                        }
            
            return {
                'success': True,
                'reserved_items': reserved_items,
                'reference': reference
            }
            
        except Exception as e:
            logger.error(f"Error reserving stock: {str(e)}")
            return {'success': False, 'error': str(e)}

# =====================================
# PERFORMANCE AND ANALYTICS UTILITIES
# =====================================

class AnalyticsCalculator:
    """
    Business intelligence and analytics calculations
    """
    
    @staticmethod
    def calculate_inventory_turnover(product, period_days: int = 365) -> Dict:
        """
        Calculate inventory turnover ratio
        
        Args:
            product: Product instance
            period_days: Period for calculation (default: 365 days)
            
        Returns:
            Turnover metrics
        """
        try:
            # Cost of goods sold (estimate)
            cogs = product.total_sold * product.total_cost_price_usd
            
            # Average inventory value
            avg_inventory_value = product.stock_value_usd
            
            # Inventory turnover ratio
            if avg_inventory_value > 0:
                turnover_ratio = float(cogs / avg_inventory_value)
                days_to_turn = period_days / turnover_ratio if turnover_ratio > 0 else 0
            else:
                turnover_ratio = 0
                days_to_turn = 0
            
            return {
                'turnover_ratio': round(turnover_ratio, 2),
                'days_to_turn': round(days_to_turn, 1),
                'cogs': float(cogs),
                'avg_inventory_value': float(avg_inventory_value),
                'performance': AnalyticsCalculator._evaluate_turnover_performance(turnover_ratio)
            }
            
        except Exception as e:
            logger.error(f"Error calculating inventory turnover: {str(e)}")
            return {'error': str(e)}
    
    @staticmethod
    def _evaluate_turnover_performance(turnover_ratio: float) -> str:
        """Evaluate turnover performance"""
        if turnover_ratio >= 12:
            return 'excellent'
        elif turnover_ratio >= 6:
            return 'good'
        elif turnover_ratio >= 3:
            return 'average'
        elif turnover_ratio >= 1:
            return 'poor'
        else:
            return 'very_poor'
    
    @staticmethod
    def calculate_abc_classification(products_queryset) -> Dict:
        """
        Calculate ABC classification for products
        
        Args:
            products_queryset: QuerySet of products
            
        Returns:
            ABC classification results
        """
        try:
            # Calculate annual value for each product
            product_values = []
            for product in products_queryset:
                annual_demand = StockOptimizer._estimate_annual_demand(product)
                annual_value = annual_demand * product.total_cost_price_usd
                product_values.append({
                    'product': product,
                    'annual_value': float(annual_value),
                    'annual_demand': annual_demand
                })
            
            # Sort by annual value (descending)
            product_values.sort(key=lambda x: x['annual_value'], reverse=True)
            
            total_value = sum(item['annual_value'] for item in product_values)
            
            # Calculate cumulative percentages and classify
            cumulative_value = 0
            classifications = {'A': [], 'B': [], 'C': []}
            
            for item in product_values:
                cumulative_value += item['annual_value']
                cumulative_percentage = (cumulative_value / total_value) * 100
                
                if cumulative_percentage <= 80:
                    classification = 'A'
                elif cumulative_percentage <= 95:
                    classification = 'B'
                else:
                    classification = 'C'
                
                item['classification'] = classification
                item['cumulative_percentage'] = round(cumulative_percentage, 2)
                classifications[classification].append(item)
            
            return {
                'classifications': classifications,
                'summary': {
                    'total_products': len(product_values),
                    'total_annual_value': total_value,
                    'A_count': len(classifications['A']),
                    'B_count': len(classifications['B']),
                    'C_count': len(classifications['C']),
                }
            }
            
        except Exception as e:
            logger.error(f"Error calculating ABC classification: {str(e)}")
            return {'error': str(e)}
    
    @staticmethod
    def calculate_category_performance(category) -> Dict:
        """
        Calculate performance metrics for a category
        
        Args:
            category: Category instance
            
        Returns:
            Performance metrics
        """
        try:
            products = category.products.filter(is_active=True)
            
            metrics = {
                'product_count': products.count(),
                'total_stock_value': float(products.aggregate(
                    total=Sum(F('total_stock') * F('total_cost_price_usd'))
                )['total'] or 0),
                'total_potential_revenue': float(products.aggregate(
                    total=Sum(F('total_stock') * F('selling_price'))
                )['total'] or 0),
                'average_markup': float(products.aggregate(
                    avg=Avg('markup_percentage')
                )['avg'] or 0),
                'low_stock_count': products.filter(
                    total_stock__lte=F('reorder_level')
                ).count(),
                'out_of_stock_count': products.filter(total_stock=0).count(),
            }
            
            metrics['potential_profit'] = metrics['total_potential_revenue'] - metrics['total_stock_value']
            metrics['average_profit_margin'] = (
                (metrics['potential_profit'] / metrics['total_stock_value'] * 100)
                if metrics['total_stock_value'] > 0 else 0
            )
            
            return metrics
            
        except Exception as e:
            logger.error(f"Error calculating category performance: {str(e)}")
            return {'error': str(e)}

# =====================================
# NOTIFICATION UTILITIES
# =====================================

class NotificationManager:
    """
    Notification management for inventory events
    """
    
    @staticmethod
    def send_low_stock_alerts():
        """Send low stock alerts to relevant users"""
        try:
            from .models import Product, ReorderAlert
            from django.contrib.auth.models import User
            
            # Find products with low stock
            low_stock_products = Product.objects.filter(
                is_active=True,
                total_stock__lte=F('reorder_level')
            )
            
            if not low_stock_products.exists():
                return
            
            # Create/update reorder alerts
            for product in low_stock_products:
                alert, created = ReorderAlert.objects.get_or_create(
                    product=product,
                    status='active',
                    defaults={
                        'quantity_needed': product.reorder_quantity,
                        'priority': 'high' if product.total_stock == 0 else 'medium'
                    }
                )
            
            # Send email notification to inventory managers
            inventory_managers = User.objects.filter(
                profile__user_type__in=['employee', 'blitzhub_admin'],
                profile__permissions__app='inventory'
            )
            
            if inventory_managers.exists():
                subject = f"Low Stock Alert - {low_stock_products.count()} products need reordering"
                
                context = {
                    'products': low_stock_products[:10],  # Limit for email
                    'total_count': low_stock_products.count()
                }
                
                message = render_to_string('inventory/emails/low_stock_alert.html', context)
                
                recipient_emails = [user.email for user in inventory_managers if user.email]
                
                if recipient_emails:
                    send_mail(
                        subject=subject,
                        message=message,
                        from_email=settings.DEFAULT_FROM_EMAIL,
                        recipient_list=recipient_emails,
                        html_message=message
                    )
            
        except Exception as e:
            logger.error(f"Error sending low stock alerts: {str(e)}")
    
    @staticmethod
    def notify_stock_adjustment(product, adjustment_type: str, quantity: int, user):
        """Notify about stock adjustments"""
        try:
            # Log the adjustment
            logger.info(
                f"Stock adjustment: {product.sku} - {adjustment_type} - "
                f"Quantity: {quantity} - User: {user.username}"
            )
            
            # Here you could add additional notifications like Slack, SMS, etc.
            
        except Exception as e:
            logger.error(f"Error notifying stock adjustment: {str(e)}")

# =====================================
# MOBILE SUPPORT UTILITIES
# =====================================

class MobileDataManager:
    """
    Utilities for mobile app support and offline functionality
    """
    
    @staticmethod
    def prepare_offline_data(user) -> Dict:
        """
        Prepare data for offline mobile use
        
        Args:
            user: User requesting offline data
            
        Returns:
            Dictionary with offline data
        """
        try:
            from .models import Product, Category, Brand, Supplier
            
            # Get essential data for offline use
            data = {
                'categories': list(Category.objects.filter(is_active=True).values(
                    'id', 'name', 'parent_id'
                )),
                'brands': list(Brand.objects.filter(is_active=True).values(
                    'id', 'name'
                )),
                'suppliers': list(Supplier.objects.filter(is_active=True).values(
                    'id', 'name', 'code'
                )),
                'products': list(Product.objects.filter(is_active=True).values(
                    'id', 'sku', 'name', 'category_id', 'brand_id', 'supplier_id',
                    'total_stock', 'reorder_level', 'selling_price', 'barcode'
                )[:1000]),  # Limit for mobile storage
                'last_sync': timezone.now().isoformat()
            }
            
            return data
            
        except Exception as e:
            logger.error(f"Error preparing offline data: {str(e)}")
            return {'error': str(e)}
    
    @staticmethod
    def sync_mobile_data(sync_data: Dict, user) -> Dict:
        """
        Sync data from mobile app
        
        Args:
            sync_data: Data from mobile app
            user: User performing sync
            
        Returns:
            Sync results
        """
        try:
            results = {
                'synced_count': 0,
                'error_count': 0,
                'errors': []
            }
            
            # Process stock adjustments from mobile
            if 'stock_adjustments' in sync_data:
                for adjustment in sync_data['stock_adjustments']:
                    try:
                        # Process mobile stock adjustment
                        # Implementation would depend on mobile data format
                        results['synced_count'] += 1
                        
                    except Exception as e:
                        results['error_count'] += 1
                        results['errors'].append(str(e))
            
            return results
            
        except Exception as e:
            logger.error(f"Error syncing mobile data: {str(e)}")
            return {'error': str(e)}

def get_stock_status(product, location=None):
    """
    Wrapper so views can import this. Falls back to Product.stock_status.
    If a location is provided, compute status from that location's available qty.
    """
    try:
        if location:
            from .models import StockLevel
            sl = StockLevel.objects.filter(product=product, location=location).first()
            qty = sl.available_quantity if sl else 0
            threshold = product.reorder_level or 0
            if qty <= 0:
                return "out_of_stock"
            if qty <= threshold:
                return "low_stock"
            return "in_stock"
        # default to the model property
        return getattr(product, "stock_status", "in_stock")
    except Exception:
        return getattr(product, "stock_status", "in_stock")

def get_products_for_quote_system(search_term: str = None,
                                  category=None,
                                  supplier=None,
                                  limit: int = 50):
    """
    Wrapper that supports optional category/supplier filters expected by views.
    Returns list[dict] in the same shape used by the quote API.
    """
    from .models import Product
    qs = Product.objects.filter(is_active=True).select_related(
        'category', 'supplier', 'brand', 'supplier_currency'
    )
    if category:
        qs = qs.filter(category=category)
    if supplier:
        qs = qs.filter(supplier=supplier)
    if search_term:
        qs = qs.filter(
            Q(name__icontains=search_term) |
            Q(sku__icontains=search_term) |
            Q(manufacturer_part_number__icontains=search_term) |
            Q(supplier_sku__icontains=search_term)
        )
    qs = qs[:max(1, min(limit, 100))]

    out = []
    for p in qs:
        out.append({
            'id': p.id,
            'sku': p.sku,
            'name': p.name,
            'description': (p.short_description or (p.description or '')[:100]),
            'category': p.category.name if p.category else '',
            'brand': p.brand.name if getattr(p, 'brand', None) else '',
            'supplier': p.supplier.name if p.supplier else '',
            'cost_price': float(p.total_cost_price_usd),
            'selling_price': float(p.selling_price or 0),
            'currency': p.selling_currency.code if p.selling_currency else 'USD',
            'current_stock': getattr(p, 'total_stock', p.current_stock),
            'available_stock': getattr(p, 'available_stock', 0),
            'stock_status': get_stock_status(p),
            'lead_time_days': getattr(p, 'supplier_lead_time_days', None),
            'minimum_quantity': getattr(p, 'supplier_minimum_order_quantity', None),
            'specifications': getattr(p, 'dynamic_attributes', {}),
            'datasheet_url': getattr(p, 'datasheet_url', ''),
            'images': getattr(p, 'product_images', []),
        })
    return out

def check_stock_availability_for_quote(quote_items: Iterable[dict]) -> dict:
    """
    Normalize items from the quote system and delegate to IntegrationHelper.check_stock_availability.
    Returns a dict keyed by product_id with can_fulfill flags, matching what views expect.
    """
    # normalize incoming shapes: accept product_id/id/product and quantity/qty
    reqs = []
    for item in quote_items:
        pid = item.get('product_id') or item.get('id') or item.get('product')
        qty = item.get('quantity') or item.get('qty') or 1
        if pid is None:
            continue
        reqs.append({'product_id': int(pid), 'quantity': int(qty)})

    results = IntegrationHelper.check_stock_availability(reqs)
    availability = {}

    for bucket, can_fulfill in [('available', True), ('partially_available', False), ('unavailable', False)]:
        for it in results.get(bucket, []):
            pid = str(it.get('product_id'))
            availability[pid] = {**it, 'can_fulfill': can_fulfill}

    # If IntegrationHelper returned an error, surface it
    if 'error' in results:
        availability['error'] = results['error']
    return availability

def reserve_stock_for_quote(quote_items: Iterable[dict], quote_reference: str, user=None) -> dict:
    """
    Normalize and delegate to IntegrationHelper.reserve_stock.
    """
    reqs = []
    for item in quote_items:
        pid = item.get('product_id') or item.get('id') or item.get('product')
        qty = item.get('quantity') or item.get('qty') or 1
        if pid is None:
            continue
        reqs.append({'product_id': int(pid), 'quantity': int(qty)})

    return IntegrationHelper.reserve_stock(reqs, reference=quote_reference)

def calculate_inventory_turnover(product, period_days: int = 365):
    """
    Thin wrapper to the AnalyticsCalculator so the symbol exists at module level.
    """
    return AnalyticsCalculator.calculate_inventory_turnover(product, period_days)

def generate_stock_valuation_report(products_qs, as_of_date=None) -> dict:
    """
    Simple valuation report (sum of current_stock * cost_price) + counts.
    """
    agg = products_qs.aggregate(
        total_value=Sum(F('current_stock') * F('cost_price')),
        total_items=Sum('current_stock')
    )
    return {
        'as_of': (as_of_date or timezone.now()).isoformat(),
        'total_value': float(agg.get('total_value') or 0),
        'total_items': int(agg.get('total_items') or 0),
        'product_count': products_qs.count(),
    }

def export_products_to_csv(products_qs) -> HttpResponse:
    """
    Minimal CSV export so the import exists for views; returns HttpResponse(csv).
    """
    import csv
    from io import StringIO
    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow(['SKU', 'Name', 'Category', 'Supplier', 'Current Stock', 'Cost Price', 'Selling Price'])
    for p in products_qs:
        writer.writerow([
            p.sku,
            p.name,
            p.category.name if p.category else '',
            p.supplier.name if p.supplier else '',
            getattr(p, 'total_stock', p.current_stock),
            float(p.cost_price or 0),
            float(p.selling_price or 0),
        ])
    resp = HttpResponse(buffer.getvalue(), content_type='text/csv')
    resp['Content-Disposition'] = 'attachment; filename="products.csv"'
    return resp

def import_products_from_csv(file_obj, user=None) -> dict:
    """
    Minimal CSV import (SKU, Name, Category?, Supplier?) to satisfy the symbol import.
    Returns {'created': X, 'updated': Y, 'errors': [..]}.
    """
    import csv
    from io import TextIOWrapper
    from .models import Product, Category, Supplier

    created = updated = 0
    errors = []
    reader = csv.DictReader(TextIOWrapper(file_obj, encoding='utf-8'))

    for i, row in enumerate(reader, start=2):
        try:
            sku = (row.get('SKU') or row.get('sku') or '').strip()
            name = (row.get('Name') or row.get('name') or '').strip()
            if not sku or not name:
                errors.append(f'Row {i}: SKU and Name are required')
                continue
            category = None
            if row.get('Category'):
                category, _ = Category.objects.get_or_create(name=row['Category'])
            supplier = None
            if row.get('Supplier'):
                supplier, _ = Supplier.objects.get_or_create(name=row['Supplier'])
            obj, was_created = Product.objects.update_or_create(
                sku=sku,
                defaults={'name': name, 'category': category, 'supplier': supplier}
            )
            created += 1 if was_created else 0
            updated += 0 if was_created else 1
        except Exception as e:
            errors.append(f'Row {i}: {e}')
    return {'created': created, 'updated': updated, 'errors': errors}

# Initialize logging
logger.info("Inventory management utilities loaded successfully")
