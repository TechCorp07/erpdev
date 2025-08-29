# inventory/views.py - Inventory Management Views

"""
Django Views for Inventory Management System

Key Features:
1. Dynamic attribute handling per component family
2. Advanced cost calculation with overhead factors
3. Multi-currency support with real-time conversion
4. Automated low stock ordering lists
5. Barcode/QR code support
6. Advanced search with electronics-specific filters
7. Business intelligence dashboards
8. Integration with quote system
9. Comprehensive reporting
"""

import base64
from io import BytesIO
import json
import csv
import logging
from datetime import datetime, timedelta
from decimal import Decimal
from django.template.loader import render_to_string
import qrcode
import requests
from weasyprint import HTML
from django.db import transaction

from django.utils.decorators import method_decorator
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views import View
from django.views.generic import (
    ListView, DetailView, CreateView, UpdateView, DeleteView, TemplateView
)
from django.views.decorators.http import require_http_methods, require_POST
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Q, Sum, Count, Avg, F, Max
from django.db.models.functions import TruncDate
from django.utils import timezone
from django.urls import reverse_lazy, reverse
from django.contrib.auth.mixins import LoginRequiredMixin
from openpyxl import Workbook

from .models import (
    Brand, Category, ComponentFamily, Currency, OverheadFactor, ProductAttributeDefinition, StorageBin,
    StorageLocation, Supplier, Location, Product, StockLevel,
    StockMovement, StockTake, StockTakeItem, PurchaseOrder,
    PurchaseOrderItem, ReorderAlert,
)
from .forms import (
    CategoryForm, CurrencyForm, ProductAttributeDefinitionForm, SupplierForm, LocationForm, ProductForm, ProductSearchForm,
    StockAdjustmentForm, StockTransferForm, PurchaseOrderForm, StockTakeForm,
    AdvancedProductSearchForm, OverheadFactorForm
)
from .decorators import (
    inventory_permission_required, stock_adjustment_permission, location_access_required,
    purchase_order_permission, stock_take_permission, cost_data_access, bulk_operation_permission
)
from .utils import (
    calculate_available_stock, get_stock_status, calculate_stock_value,
    create_stock_movement, BarcodeManager
)

logger = logging.getLogger(__name__)

# =====================================
# UNIVERSAL OVERVIEW VIEWS
# =====================================

class BaseInventoryListView(LoginRequiredMixin, ListView):
    """Base class for inventory list views"""
    paginate_by = 25
    
    @method_decorator(inventory_permission_required('view'))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

class BaseInventoryCreateView(LoginRequiredMixin, CreateView):
    """Base class for inventory create views"""
    
    @method_decorator(inventory_permission_required('edit'))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    
    def form_valid(self, form):
        messages.success(
            self.request, 
            f'{self.model._meta.verbose_name.title()} "{form.instance}" created successfully'
        )
        return super().form_valid(form)

class BaseInventoryUpdateView(LoginRequiredMixin, UpdateView):
    """Base class for inventory update views"""
    
    @method_decorator(inventory_permission_required('edit'))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'page_title': f'Edit {self.model._meta.verbose_name.title()}: {self.object}',
            'form_action': 'Update',
        })
        return context
    
    def form_valid(self, form):
        messages.success(
            self.request, 
            f'{self.model._meta.verbose_name.title()} "{form.instance}" updated successfully'
        )
        return super().form_valid(form)

class BaseInventoryDeleteView(LoginRequiredMixin, DeleteView):
    """Base class for inventory delete views"""
    
    @method_decorator(inventory_permission_required('admin'))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

class BaseInventoryDetailView(LoginRequiredMixin, DetailView):
    """Base class for inventory detail views"""
    
    @method_decorator(inventory_permission_required('view'))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

class InventoryViewMixin:
    """Common mixin for inventory views"""
    
    @method_decorator(inventory_permission_required('view'))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

class InventoryEditMixin:
    """Mixin for inventory edit operations"""
    
    @method_decorator(inventory_permission_required('edit'))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

class InventoryAdminMixin:
    """Mixin for inventory admin operations"""
    
    @method_decorator(inventory_permission_required('admin'))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

# =====================================
# DASHBOARD AND OVERVIEW VIEWS
# =====================================

class InventoryDashboardView(LoginRequiredMixin, TemplateView):
    """
    Main inventory dashboard with comprehensive overview.
    
    Provides key metrics, alerts, and quick access to common operations.
    This is the command center for inventory management operations.
    """
    template_name = 'inventory/dashboard.html'
    
    @method_decorator(inventory_permission_required('view'))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Basic statistics
        products = Product.objects.filter(is_active=True)
        context.update({
            'total_products': products.count(),
            'total_categories': Category.objects.filter(is_active=True).count(),
            'total_suppliers': Supplier.objects.filter(is_active=True).count(),
            'total_brands': Brand.objects.filter(is_active=True).count(),
            'total_locations': StorageLocation.objects.filter(is_active=True).count(),
        })
        
        # Financial metrics
        total_stock_value = products.aggregate(
            total_cost=Sum(F('total_stock') * F('total_cost_price_usd')),
            total_selling=Sum(F('total_stock') * F('selling_price'))
        )
        context.update({
            'total_stock_value_cost': total_stock_value['total_cost'] or Decimal('0.00'),
            'total_stock_value_selling': total_stock_value['total_selling'] or Decimal('0.00'),
            'potential_profit': (total_stock_value['total_selling'] or Decimal('0.00')) - 
                              (total_stock_value['total_cost'] or Decimal('0.00')),
        })
        
        # Stock status breakdown
        context['stock_status'] = {
            'in_stock': products.filter(total_stock__gt=F('reorder_level')).count(),
            'low_stock': products.filter(
                total_stock__lte=F('reorder_level'),
                total_stock__gt=0
            ).count(),
            'out_of_stock': products.filter(total_stock=0).count(),
            'needs_reorder': products.filter(total_stock__lte=F('reorder_level')).count()
        }
        
        # Component family analysis
        context['component_families_stats'] = ComponentFamily.objects.filter(
            products__is_active=True
        ).annotate(
            product_count=Count('products'),
            total_value=Sum(F('products__total_stock') * F('products__total_cost_price_usd')),
            avg_markup=Avg('products__markup_percentage')
        ).order_by('-total_value')[:8]
        
        # Supplier performance
        context['top_suppliers'] = Supplier.objects.filter(
            products__is_active=True,
            is_active=True
        ).annotate(
            product_count=Count('products'),
            avg_rating=Avg('rating'),
            total_inventory_value=Sum(F('products__total_stock') * F('products__total_cost_price_usd'))
        ).order_by('-product_count')[:6]
        
        # Recent alerts and activities
        context['recent_reorder_alerts'] = ReorderAlert.objects.filter(
            status__in=['active', 'acknowledged']
        ).select_related('product', 'product__supplier').order_by('-created_at')[:10]
        
        # Low stock products by category
        context['low_stock_by_category'] = Category.objects.filter(
            products__is_active=True,
            products__total_stock__lte=F('products__reorder_level')
        ).annotate(
            low_stock_count=Count('products')
        ).order_by('-low_stock_count')[:5]
        
        # Currency rates (for multi-currency awareness)
        context['currencies'] = Currency.objects.filter(is_active=True).order_by('code')
        
        # Storage utilization
        context['storage_locations'] = StorageLocation.objects.filter(
            is_active=True
        ).annotate(
            product_count=Count('productstocklevel__product', distinct=True),
            total_items=Sum('productstocklevel__quantity_on_hand')
        )[:3]
        
        return context

    def _calculate_total_stock_value(self):
        """Calculate total value of current stock"""
        try:
            return Product.objects.filter(is_active=True).aggregate(
                total_value=Sum(F('current_stock') * F('cost_price'))
            )['total_value'] or Decimal('0.00')
        except Exception:
            return Decimal('0.00')
    
    def _get_low_stock_count(self):
        """Get count of products with low stock"""
        return Product.objects.filter(
            is_active=True,
            current_stock__lte=F('reorder_level'),
            current_stock__gt=0
        ).count()
    
    def _get_out_of_stock_count(self):
        """Get count of products that are out of stock"""
        return Product.objects.filter(
            is_active=True,
            current_stock=0
        ).count()
    
    def _get_pending_po_count(self):
        """Get count of pending purchase orders"""
        return PurchaseOrder.objects.filter(
            status__in=['draft', 'sent', 'acknowledged']
        ).count()
    
    def _get_active_alerts_count(self):
        """Get count of active reorder alerts"""
        return ReorderAlert.objects.filter(
            status__in=['active', 'acknowledged']
        ).count()
    
    def _get_recent_movements(self):
        """Get recent stock movements"""
        return StockMovement.objects.select_related(
            'product', 'created_by'
        ).order_by('-created_at')[:10]
    
    def _get_recent_alerts(self):
        """Get recent reorder alerts"""
        return ReorderAlert.objects.select_related(
            'product', 'suggested_supplier'
        ).order_by('-created_at')[:5]
    
    def _get_pending_stock_takes(self):
        """Get pending stock takes"""
        return StockTake.objects.filter(
            status__in=['planned', 'in_progress']
        ).order_by('scheduled_date')[:5]
    
    def _get_category_stats(self):
        """Get statistics by category"""
        return Category.objects.filter(is_active=True).annotate(
            product_count=Count('products', filter=Q(products__is_active=True)),
            total_value=Sum(
                F('products__current_stock') * F('products__cost_price'),
                filter=Q(products__is_active=True)
            )
        ).order_by('-total_value')[:10]
    
    def _get_supplier_stats(self):
        """Get statistics by supplier"""
        return Supplier.objects.filter(is_active=True).annotate(
            product_count=Count('products', filter=Q(products__is_active=True)),
            total_value=Sum(
                F('products__current_stock') * F('products__cost_price'),
                filter=Q(products__is_active=True)
            )
        ).order_by('-total_value')[:10]
    
    def _get_location_stats(self):
        """Get statistics by location"""
        return Location.objects.filter(is_active=True).annotate(
            product_count=Count('stock_levels__product', distinct=True),
            total_stock=Sum('stock_levels__quantity')
        ).order_by('-total_stock')
    
    def _get_performance_metrics(self):
        """Get key performance indicators"""
        thirty_days_ago = timezone.now() - timedelta(days=30)
        
        return {
            'movements_last_30_days': StockMovement.objects.filter(
                created_at__gte=thirty_days_ago
            ).count(),
            'pos_created_last_30_days': PurchaseOrder.objects.filter(
                created_at__gte=thirty_days_ago
            ).count(),
            'alerts_resolved_last_30_days': ReorderAlert.objects.filter(
                resolved_at__gte=thirty_days_ago
            ).count(),
        }

class InventoryValuationReportView(LoginRequiredMixin, TemplateView):
    template_name = "inventory/reports/inventory_valuation.html"

    @method_decorator(inventory_permission_required('view'))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        import datetime
        from django.db.models import Sum, F, Count, Q
        from .models import Product, Category, Supplier, Location  # Import Location!
        context = super().get_context_data(**kwargs)

        # 1. Support filters from GET
        request = self.request
        as_of_date = request.GET.get('as_of_date') or datetime.date.today().isoformat()
        category_id = request.GET.get('category')
        location_id = request.GET.get('location')

        # 2. Base queryset: only active products
        products = Product.objects.filter(is_active=True)

        if category_id:
            products = products.filter(category_id=category_id)
        if location_id:
            # If you track stock by location via a StockLevel model, you would join here.
            # For now, just show all products, or implement as needed.
            pass

        # 3. Compute stats for products
        total_products = products.count()
        total_quantity = products.aggregate(total=Sum('current_stock'))['total'] or 0
        total_value = products.aggregate(
            value=Sum(F('current_stock') * F('cost_price'))
        )['value'] or 0
        avg_margin = products.aggregate(
            avg=Sum(
                F('selling_price') - F('cost_price')
            ) / Count('id')
        )['avg'] or 0

        # 4. Category breakdown for charts
        categories = Category.objects.filter(is_active=True).annotate(
            total_quantity=Sum('products__current_stock', filter=Q(products__is_active=True)),
            total_value=Sum(F('products__current_stock') * F('products__cost_price'), filter=Q(products__is_active=True)),
            product_count=Count('products', filter=Q(products__is_active=True)),
        ).order_by('-total_value')

        total_value_all = categories.aggregate(tot=Sum('total_value'))['tot'] or 1  # avoid divide by zero

        category_list = []
        for cat in categories:
            if not cat.total_value:
                continue
            percent = (cat.total_value / total_value_all) * 100 if total_value_all else 0
            category_list.append({
                'id': cat.id,
                'name': cat.name,
                'total_quantity': cat.total_quantity or 0,
                'total_value': cat.total_value or 0,
                'product_count': cat.product_count or 0,
                'percentage': percent,
            })

        # 5. Detailed product list for table
        product_list = []
        for p in products.select_related('category', 'supplier'):
            margin = (p.selling_price - p.cost_price) if p.cost_price else 0
            margin_pct = ((p.selling_price - p.cost_price) / p.cost_price * 100) if p.cost_price else 0
            product_list.append({
                'id': p.id,
                'sku': p.sku,
                'name': p.name,
                'category': p.category.name if p.category else '',
                'supplier': p.supplier.name if p.supplier else '',
                'quantity': p.current_stock,
                'cost_price': p.cost_price,
                'selling_price': p.selling_price,
                'total_value': (p.current_stock * p.cost_price) if p.current_stock else 0,
                'margin_amount': margin,
                'margin_percentage': margin_pct,
                'stock_status': p.stock_status,
            })

        # 6. Other context for filters
        all_categories = Category.objects.filter(is_active=True)
        all_locations = Location.objects.filter(is_active=True)

        # 7. Provide 'today' for the date picker default
        today = datetime.date.today().isoformat()

        context.update({
            "today": today,
            "categories": all_categories,
            "locations": all_locations,
            "report_data": {
                "as_of_date": as_of_date,
                "total_products": total_products,
                "total_quantity": total_quantity,
                "total_value": total_value,
                "avg_margin": avg_margin,
                "categories": category_list,
                "products": product_list,
                "location": all_locations.get(id=location_id).name if location_id else "All Locations",
            },
            "last_updated": datetime.datetime.now(),
        })
        return context

class InventoryOverviewView(LoginRequiredMixin, TemplateView):
    """Detailed inventory overview with filtering and analysis"""
    template_name = 'inventory/overview.html'
    
    @method_decorator(inventory_permission_required('view'))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get filter parameters
        category_id = self.request.GET.get('category')
        supplier_id = self.request.GET.get('supplier')
        location_id = self.request.GET.get('location')
        
        # Build base query
        products = Product.objects.filter(is_active=True).select_related(
            'category', 'supplier'
        )
        
        # Apply filters
        if category_id:
            products = products.filter(category_id=category_id)
        if supplier_id:
            products = products.filter(supplier_id=supplier_id)
        
        # Calculate metrics
        context.update({
            'products': products,
            'categories': Category.objects.filter(is_active=True),
            'suppliers': Supplier.objects.filter(is_active=True),
            'locations': Location.objects.filter(is_active=True),
            'selected_category': category_id,
            'selected_supplier': supplier_id,
            'selected_location': location_id,
            'total_value': calculate_stock_value(products),
            'total_quantity': products.aggregate(
                total=Sum('current_stock')
            )['total'] or 0
        })
        
        return context

@login_required
@inventory_permission_required('view')
def inventory_alerts_view(request):
    """View for managing inventory alerts and notifications"""
    template_name = 'inventory/alerts.html'
    
    # Get active alerts
    reorder_alerts = ReorderAlert.objects.filter(
        status__in=['active', 'acknowledged']
    ).select_related('product', 'suggested_supplier').order_by('priority', '-created_at')
    
    # Get low stock products
    low_stock_products = Product.objects.filter(
        is_active=True,
        current_stock__lte=F('reorder_level')
    ).select_related('category', 'supplier').order_by('current_stock')
    
    # Get overdue purchase orders
    overdue_pos = PurchaseOrder.objects.filter(
        status__in=['sent', 'acknowledged'],
        expected_delivery_date__lt=timezone.now().date()
    ).select_related('supplier')
    
    context = {
        'reorder_alerts': reorder_alerts,
        'low_stock_products': low_stock_products,
        'overdue_pos': overdue_pos,
        'alert_count': reorder_alerts.count(),
        'low_stock_count': low_stock_products.count(),
        'overdue_po_count': overdue_pos.count()
    }
    
    return render(request, template_name, context)

@login_required
@inventory_permission_required('view')
def quick_stats_api(request):
    """
    API endpoint for real-time dashboard statistics.
    
    Provides current inventory metrics for dashboard widgets and mobile apps.
    """
    try:
        stats = {
            'timestamp': timezone.now().isoformat(),
            'total_products': Product.objects.filter(is_active=True).count(),
            'low_stock_count': Product.objects.filter(
                is_active=True,
                current_stock__lte=F('reorder_level')
            ).count(),
            'out_of_stock_count': Product.objects.filter(
                is_active=True,
                current_stock=0
            ).count(),
            'total_stock_value': float(
                Product.objects.filter(is_active=True).aggregate(
                    total=Sum(F('current_stock') * F('cost_price'))
                )['total'] or 0
            ),
            'active_alerts': ReorderAlert.objects.filter(
                status__in=['active', 'acknowledged']
            ).count(),
            'pending_pos': PurchaseOrder.objects.filter(
                status__in=['draft', 'sent', 'acknowledged']
            ).count()
        }
        
        return JsonResponse({'success': True, 'stats': stats})
        
    except Exception as e:
        logger.error(f"Error in quick stats API: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
@inventory_permission_required('view')
def search_suggestions_api(request):
    """API for search suggestions"""
    query = request.GET.get('q', '')
    if len(query) < 2:
        return JsonResponse({'suggestions': []})
    
    products = Product.objects.filter(
        Q(name__icontains=query) | Q(sku__icontains=query),
        is_active=True
    )[:10]
    
    suggestions = [
        {
            'id': product.id,
            'text': f"{product.name} ({product.sku})",
            'type': 'product'
        }
        for product in products
    ]
    
    return JsonResponse({'suggestions': suggestions})

# =====================================
# PRODUCT MANAGEMENT VIEWS
# =====================================

class ProductListView(LoginRequiredMixin, ListView):
    """
    Product listing with search, filtering, and pagination.
    
    Provides comprehensive product management interface with bulk operations.
    """
    model = Product
    template_name = 'inventory/product/product_list.html'
    context_object_name = 'products'
    paginate_by = 25
    
    @method_decorator(inventory_permission_required('view'))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    
    def get_queryset(self):
        queryset = Product.objects.filter(is_active=True).select_related(
            'category', 'supplier', 'brand', 'component_family', 'supplier_currency'
        )
        
        # Apply filters
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) |
                Q(sku__icontains=search) |
                Q(manufacturer_part_number__icontains=search) |
                Q(supplier_sku__icontains=search) |
                Q(description__icontains=search)
            )
        
        category = self.request.GET.get('category')
        if category:
            queryset = queryset.filter(category_id=category)
        
        component_family = self.request.GET.get('component_family')
        if component_family:
            queryset = queryset.filter(component_family_id=component_family)
        
        brand = self.request.GET.get('brand')
        if brand:
            queryset = queryset.filter(brand_id=brand)
        
        supplier = self.request.GET.get('supplier')
        if supplier:
            queryset = queryset.filter(supplier_id=supplier)
        
        stock_status = self.request.GET.get('stock_status')
        if stock_status:
            if stock_status == 'in_stock':
                queryset = queryset.filter(total_stock__gt=F('reorder_level'))
            elif stock_status == 'low_stock':
                queryset = queryset.filter(
                    total_stock__lte=F('reorder_level'),
                    total_stock__gt=0
                )
            elif stock_status == 'out_of_stock':
                queryset = queryset.filter(total_stock=0)
            elif stock_status == 'needs_reorder':
                queryset = queryset.filter(total_stock__lte=F('reorder_level'))
        
        # Sorting
        sort_by = self.request.GET.get('sort', 'name')
        if sort_by in ['name', '-name', 'sku', '-sku', 'total_stock', '-total_stock', 
                       'selling_price', '-selling_price', 'markup_percentage', '-markup_percentage']:
            queryset = queryset.order_by(sort_by)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Add filter options
        context.update({
            'categories': Category.objects.filter(is_active=True).order_by('name'),
            'component_families': ComponentFamily.objects.filter(is_active=True).order_by('name'),
            'brands': Brand.objects.filter(is_active=True).order_by('name'),
            'suppliers': Supplier.objects.filter(is_active=True).order_by('name'),
            'current_filters': {
                'search': self.request.GET.get('search', ''),
                'category': self.request.GET.get('category', ''),
                'component_family': self.request.GET.get('component_family', ''),
                'brand': self.request.GET.get('brand', ''),
                'supplier': self.request.GET.get('supplier', ''),
                'stock_status': self.request.GET.get('stock_status', ''),
                'sort': self.request.GET.get('sort', 'name'),
            }
        })
        
        return context

class ProductDetailView(LoginRequiredMixin, DetailView):
    """
    Detailed product view with stock information and analytics.
    
    Shows comprehensive product information including stock levels,
    movement history, supplier details, and performance metrics.
    """
    model = Product
    template_name = 'inventory/product/product_detail.html'
    context_object_name = 'product'
    
    @method_decorator(inventory_permission_required('view'))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        product = self.object
        
        # Get dynamic attributes for this product's component family
        if product.component_family:
            context['attribute_definitions'] = product.component_family.all_attributes
        
        # Cost breakdown analysis
        cost_breakdown = {
            'base_cost_original': product.cost_price,
            'base_cost_usd': product.cost_price_usd,
            'shipping_cost': product.shipping_cost_per_unit,
            'insurance_cost': product.insurance_cost_per_unit,
            'customs_duty': product.cost_price_usd * (product.customs_duty_percentage / 100),
            'vat_cost': (product.cost_price_usd + product.cost_price_usd * (product.customs_duty_percentage / 100)) * (product.vat_percentage / 100),
            'other_fees': product.other_fees_per_unit,
            'overhead_cost': product.overhead_cost_per_unit,
            'total_cost': product.total_cost_price_usd,
        }
        context['cost_breakdown'] = cost_breakdown
        
        # Stock levels by location
        context['stock_levels'] = product.stock_levels.select_related(
            'location', 'storage_bin'
        ).filter(is_active=True)
        
        # Recent stock movements
        context['recent_movements'] = StockMovement.objects.filter(
            product=product
        ).select_related('location', 'created_by').order_by('-created_at')[:10]
        
        # Competitive analysis
        if product.competitor_min_price or product.competitor_max_price:
            context['competitive_analysis'] = {
                'our_price': product.selling_price,
                'competitor_min': product.competitor_min_price,
                'competitor_max': product.competitor_max_price,
                'position': product.price_position,
                'price_advantage': None
            }
            
            if product.competitor_min_price:
                diff = float(product.selling_price) - float(product.competitor_min_price)
                context['competitive_analysis']['price_advantage'] = diff
        
        # Generate QR code
        context['qr_code_data'] = self.generate_qr_code(product)
        
        return context
    
    def generate_qr_code(self, product):
        """Generate QR code for product"""
        qr_data = {
            'sku': product.sku,
            'name': product.name,
            'price': float(product.selling_price),
            'stock': product.total_stock,
            'url': self.request.build_absolute_uri(
                reverse('inventory:product_detail', args=[product.id])
            )
        }
        
        qr = qrcode.QRCode(version=1, box_size=10, border=5)
        qr.add_data(json.dumps(qr_data))
        qr.make(fit=True)
        
        img = qr.make_image(fill_color="black", back_color="white")
        buffer = BytesIO()
        img.save(buffer, format='PNG')
        qr_code_data = base64.b64encode(buffer.getvalue()).decode()
        
        return qr_code_data

class ProductCreateView(LoginRequiredMixin, CreateView):
    """Product creation view with intelligent defaults"""
    model = Product
    form_class = ProductForm
    template_name = 'inventory/product/product_form.html'
    
    @method_decorator(inventory_permission_required('edit'))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs
    
    def form_valid(self, form):
        form.instance.created_by = self.request.user
        
        messages.success(
            self.request,
            f'Product "{form.instance.name}" created successfully.'
        )
        
        # Create notification for inventory managers
        from core.utils import create_bulk_notifications
        from django.contrib.auth.models import User
        
        managers = User.objects.filter(
            profile__user_type__in=['sales_manager', 'blitzhub_admin', 'it_admin'],
            profile__is_active=True
        ).exclude(id=self.request.user.id)
        
        create_bulk_notifications(
            users=managers,
            title="New Product Added",
            message=f"Product '{form.instance.name}' ({form.instance.sku}) has been added to inventory",
            notification_type="info",
            action_url=form.instance.get_absolute_url(),
            action_text="View Product"
        )
        
        return super().form_valid(form)
    
    def get_success_url(self):
        return reverse('inventory:product_detail', kwargs={'pk': self.object.pk})

class ProductUpdateView(LoginRequiredMixin, UpdateView):
    """Product update view with change tracking"""
    model = Product
    form_class = ProductForm
    template_name = 'inventory/product/product_form.html'
    
    @method_decorator(inventory_permission_required('edit'))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs
    
    def form_valid(self, form):
        old_instance = Product.objects.get(pk=self.object.pk)
        response = super().form_valid(form)
        
        # Track significant changes
        changes = []
        if old_instance.selling_price != form.instance.selling_price:
            changes.append(f"Price: ${old_instance.selling_price} → ${form.instance.selling_price}")
        
        if old_instance.reorder_level != form.instance.reorder_level:
            changes.append(f"Reorder level: {old_instance.reorder_level} → {form.instance.reorder_level}")
        
        if changes:
            logger.info(f"Product {form.instance.sku} updated by {self.request.user.username}: {', '.join(changes)}")
        
        messages.success(
            self.request,
            f'Product "{form.instance.name}" updated successfully.'
        )
        
        return response
    
    def get_success_url(self):
        return reverse('inventory:product_detail', kwargs={'pk': self.object.pk})

class ProductDeleteView(LoginRequiredMixin, DeleteView):
    """Product deletion view with safety checks"""
    model = Product
    template_name = 'inventory/product/product_confirm_delete.html'
    success_url = reverse_lazy('inventory:product_list')
    
    @method_decorator(inventory_permission_required('admin'))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    
    def delete(self, request, *args, **kwargs):
        product = self.get_object()
        
        # Check for dependencies
        if product.stock_movements.exists():
            messages.error(
                request,
                f'Cannot delete product "{product.name}" because it has stock movement history.'
            )
            return redirect('inventory:product_detail', pk=product.pk)
        
        if product.current_stock > 0:
            messages.error(
                request,
                f'Cannot delete product "{product.name}" because it has current stock.'
            )
            return redirect('inventory:product_detail', pk=product.pk)
        
        # Log deletion
        logger.warning(f"Product {product.sku} deleted by {request.user.username}")
        
        messages.success(
            request,
            f'Product "{product.name}" deleted successfully.'
        )
        
        return super().delete(request, *args, **kwargs)

class ProductSearchView(LoginRequiredMixin, ListView):
    template_name = "inventory/product/product_search.html"
    model = Product
    context_object_name = "products"
    paginate_by = 30

    @method_decorator(inventory_permission_required('view'))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def get_queryset(self):
        queryset = Product.objects.select_related('category', 'supplier').all()
        search_form = ProductSearchForm(self.request.GET or None)
        
        if search_form.is_valid():
            data = search_form.cleaned_data
            search = data.get('search')
            if search:
                queryset = queryset.filter(
                    Q(name__icontains=search) |
                    Q(sku__icontains=search) |
                    Q(description__icontains=search) |
                    Q(barcode__icontains=search)
                )
            category = data.get('category')
            if category:
                queryset = queryset.filter(category=category)
            supplier = data.get('supplier')
            if supplier:
                queryset = queryset.filter(supplier=supplier)
            stock_status = data.get('stock_status')
            if stock_status == 'in_stock':
                queryset = queryset.filter(current_stock__gt=F('reorder_level'))
            elif stock_status == 'low_stock':
                queryset = queryset.filter(current_stock__lte=F('reorder_level'), current_stock__gt=0)
            elif stock_status == 'out_of_stock':
                queryset = queryset.filter(current_stock=0)
            is_active = data.get('is_active')
            if is_active == 'true':
                queryset = queryset.filter(is_active=True)
            elif is_active == 'false':
                queryset = queryset.filter(is_active=False)
            min_price = data.get('min_price')
            if min_price is not None:
                queryset = queryset.filter(selling_price__gte=min_price)
            max_price = data.get('max_price')
            if max_price is not None:
                queryset = queryset.filter(selling_price__lte=max_price)
        else:
            # If not valid, show all products, or restrict further as you like.
            queryset = queryset.filter(is_active=True)
        
        # You can add more custom filtering logic here as needed.

        return queryset.order_by('name')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["search_form"] = ProductSearchForm(self.request.GET)
        context["categories"] = Category.objects.filter(is_active=True)
        context["suppliers"] = Supplier.objects.filter(is_active=True)
        context["can_edit"] = self.request.user.is_staff or (
            hasattr(self.request.user, 'profile') and self.request.user.profile.user_type in [
                'sales_manager', 'blitzhub_admin', 'it_admin'
            ]
        )
        return context

class ProductBulkImportView(LoginRequiredMixin, TemplateView):
    """
    Advanced bulk import with validation and error handling.
    """
    template_name = 'inventory/data/product_bulk_import.html'
    
    @method_decorator(bulk_operation_permission)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        context.update({
            'page_title': 'Bulk Import Products',
            'categories': Category.objects.filter(is_active=True),
            'suppliers': Supplier.objects.filter(is_active=True),
            'brands': Brand.objects.filter(is_active=True),
            'supported_formats': ['CSV', 'Excel (.xlsx)'],
        })
        return context
    
    def post(self, request, *args, **kwargs):
        try:
            uploaded_file = request.FILES.get('import_file')
            
            if not uploaded_file:
                messages.error(request, 'No file uploaded')
                return self.get(request, *args, **kwargs)
            
            # Determine file type and process
            file_extension = uploaded_file.name.split('.')[-1].lower()
            
            if file_extension == 'csv':
                return self._process_csv_import(request, uploaded_file)
            elif file_extension in ['xlsx', 'xls']:
                return self._process_excel_import(request, uploaded_file)
            else:
                messages.error(request, 'Unsupported file format. Please use CSV or Excel files.')
                return self.get(request, *args, **kwargs)
            
        except Exception as e:
            messages.error(request, f'Import failed: {str(e)}')
            return self.get(request, *args, **kwargs)
    
    def _process_csv_import(self, request, uploaded_file):
        """Process CSV file import"""
        import csv
        from io import TextIOWrapper
        
        results = {'created': 0, 'updated': 0, 'errors': []}
        
        with transaction.atomic():
            # Read CSV
            file_data = TextIOWrapper(uploaded_file.file, encoding='utf-8')
            csv_reader = csv.DictReader(file_data)
            
            for row_num, row in enumerate(csv_reader, start=2):  # Start from row 2 (header is row 1)
                try:
                    # Validate required fields
                    if not row.get('sku') or not row.get('name'):
                        results['errors'].append(f'Row {row_num}: SKU and Name are required')
                        continue
                    
                    # Get or create related objects
                    category = None
                    if row.get('category'):
                        category, _ = Category.objects.get_or_create(
                            name=row['category'],
                            defaults={'is_active': True}
                        )
                    
                    supplier = None
                    if row.get('supplier'):
                        supplier, _ = Supplier.objects.get_or_create(
                            name=row['supplier'],
                            defaults={'is_active': True}
                        )
                    
                    brand = None
                    if row.get('brand'):
                        brand, _ = Brand.objects.get_or_create(
                            name=row['brand'],
                            defaults={'is_active': True}
                        )
                    
                    # Create or update product
                    product, created = Product.objects.update_or_create(
                        sku=row['sku'],
                        defaults={
                            'name': row['name'],
                            'description': row.get('description', ''),
                            'category': category,
                            'supplier': supplier,
                            'brand': brand,
                            'cost_price': Decimal(row.get('cost_price', '0')),
                            'selling_price': Decimal(row.get('selling_price', '0')),
                            'current_stock': int(row.get('current_stock', 0)),
                            'reorder_level': int(row.get('reorder_level', 0)),
                            'economic_order_quantity': int(row.get('economic_order_quantity', 0)) or None,
                            'barcode': row.get('barcode', ''),
                            'is_active': row.get('is_active', 'True').lower() in ['true', '1', 'yes'],
                        }
                    )
                    
                    if created:
                        results['created'] += 1
                    else:
                        results['updated'] += 1
                        
                except Exception as e:
                    results['errors'].append(f'Row {row_num}: {str(e)}')
                    continue
        
        # Show results
        if results['created'] or results['updated']:
            messages.success(
                request,
                f"Import completed: {results['created']} created, {results['updated']} updated"
            )
        
        if results['errors']:
            error_msg = f"{len(results['errors'])} errors occurred:\n" + '\n'.join(results['errors'][:10])
            if len(results['errors']) > 10:
                error_msg += f"\n... and {len(results['errors']) - 10} more errors"
            messages.error(request, error_msg)
        
        return redirect('inventory:product_list')
    
    def _process_excel_import(self, request, uploaded_file):
        """Process Excel file import"""
        from openpyxl import load_workbook
        
        results = {'created': 0, 'updated': 0, 'errors': []}
        
        try:
            workbook = load_workbook(uploaded_file)
            worksheet = workbook.active
            
            # Get headers from first row
            headers = []
            for cell in worksheet[1]:
                headers.append(cell.value.lower() if cell.value else '')
            
            with transaction.atomic():
                for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        # Convert row to dict
                        row_dict = {}
                        for i, value in enumerate(row):
                            if i < len(headers) and headers[i]:
                                row_dict[headers[i]] = value
                        
                        # Skip empty rows
                        if not any(row_dict.values()):
                            continue
                        
                        # Validate required fields
                        if not row_dict.get('sku') or not row_dict.get('name'):
                            results['errors'].append(f'Row {row_num}: SKU and Name are required')
                            continue
                        
                        # Process similar to CSV import...
                        # (Similar logic as _process_csv_import but adapted for Excel data)
                        
                    except Exception as e:
                        results['errors'].append(f'Row {row_num}: {str(e)}')
                        continue
            
            # Show results (similar to CSV)
            if results['created'] or results['updated']:
                messages.success(
                    request,
                    f"Import completed: {results['created']} created, {results['updated']} updated"
                )
            
            if results['errors']:
                error_msg = f"{len(results['errors'])} errors occurred:\n" + '\n'.join(results['errors'][:10])
                messages.error(request, error_msg)
            
        except Exception as e:
            messages.error(request, f'Excel processing failed: {str(e)}')
        
        return redirect('inventory:product_list')

class ProductBulkCreateView(LoginRequiredMixin, TemplateView):
    """
    Bulk create multiple products efficiently.
    """
    template_name = 'inventory/product/product_bulk_create.html'
    
    @method_decorator(bulk_operation_permission)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        context.update({
            'page_title': 'Bulk Create Products',
            'categories': Category.objects.filter(is_active=True),
            'suppliers': Supplier.objects.filter(is_active=True),
            'brands': Brand.objects.filter(is_active=True),
        })
        return context
    
    def post(self, request, *args, **kwargs):
        try:
            products_data = json.loads(request.POST.get('products', '[]'))
            
            if not products_data:
                messages.error(request, 'No product data provided')
                return redirect('inventory:product_bulk_create')
            
            created_count = 0
            with transaction.atomic():
                for product_data in products_data:
                    try:
                        # Get foreign key objects
                        category = Category.objects.get(id=product_data.get('category_id')) if product_data.get('category_id') else None
                        supplier = Supplier.objects.get(id=product_data.get('supplier_id')) if product_data.get('supplier_id') else None
                        brand = Brand.objects.get(id=product_data.get('brand_id')) if product_data.get('brand_id') else None
                        
                        Product.objects.create(
                            name=product_data['name'],
                            sku=product_data['sku'],
                            description=product_data.get('description', ''),
                            category=category,
                            supplier=supplier,
                            brand=brand,
                            cost_price=Decimal(str(product_data.get('cost_price', '0'))),
                            selling_price=Decimal(str(product_data.get('selling_price', '0'))),
                            current_stock=int(product_data.get('current_stock', 0)),
                            reorder_level=int(product_data.get('reorder_level', 0)),
                            barcode=product_data.get('barcode', ''),
                            created_by=request.user,
                            is_active=True,
                        )
                        created_count += 1
                        
                    except Exception as e:
                        logger.error(f"Failed to create product {product_data.get('name', 'Unknown')}: {str(e)}")
                        continue
            
            messages.success(request, f'Successfully created {created_count} products')
            return redirect('inventory:product_list')
            
        except Exception as e:
            messages.error(request, f'Bulk create failed: {str(e)}')
            return redirect('inventory:product_bulk_create')

class ProductAdvancedSearchView(LoginRequiredMixin, ListView):
    """
    Advanced product search & filtering using AdvancedProductSearchForm.
    """
    model = Product
    template_name = "inventory/product/product_advanced_search.html"  # create this if you don't have it
    context_object_name = "products"
    paginate_by = 30

    @method_decorator(inventory_permission_required('view'))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def get_queryset(self):
        qs = Product.objects.all().select_related(
            'category', 'supplier', 'brand', 'component_family'
        )

        form = AdvancedProductSearchForm(self.request.GET or None)
        if not form.is_valid():
            return qs

        cd = form.cleaned_data

        # Free text
        search = cd.get('search')
        if search:
            qs = qs.filter(
                Q(name__icontains=search) |
                Q(sku__icontains=search) |
                Q(manufacturer_part_number__icontains=search) |
                Q(supplier_sku__icontains=search) |
                Q(description__icontains=search) |
                Q(barcode__icontains=search)
            )

        # Entity filters
        if cd.get('category'):
            qs = qs.filter(category=cd['category'])
        if cd.get('component_family'):
            qs = qs.filter(component_family=cd['component_family'])
        if cd.get('brand'):
            qs = qs.filter(brand=cd['brand'])
        if cd.get('supplier'):
            qs = qs.filter(supplier=cd['supplier'])
        if cd.get('supplier_country'):
            qs = qs.filter(supplier__country=cd['supplier_country'])

        # Choice filters (only apply if model has those fields)
        if cd.get('product_type') and hasattr(Product, 'product_type'):
            qs = qs.filter(product_type=cd['product_type'])
        if cd.get('quality_grade') and hasattr(Product, 'quality_grade'):
            qs = qs.filter(quality_grade=cd['quality_grade'])

        # Stock status
        stock_status = cd.get('stock_status')
        if stock_status == 'in_stock':
            qs = qs.filter(total_stock__gt=F('reorder_level'))
        elif stock_status == 'low_stock':
            qs = qs.filter(total_stock__lte=F('reorder_level'), total_stock__gt=0)
        elif stock_status == 'out_of_stock':
            qs = qs.filter(total_stock=0)
        elif stock_status == 'needs_reorder':
            qs = qs.filter(total_stock__lte=F('reorder_level'))

        # Price range
        if cd.get('price_range_min') is not None:
            qs = qs.filter(selling_price__gte=cd['price_range_min'])
        if cd.get('price_range_max') is not None:
            qs = qs.filter(selling_price__lte=cd['price_range_max'])

        # Markup range (guard for field existence)
        if hasattr(Product, 'markup_percentage'):
            if cd.get('markup_range_min') is not None:
                qs = qs.filter(markup_percentage__gte=cd['markup_range_min'])
            if cd.get('markup_range_max') is not None:
                qs = qs.filter(markup_percentage__lte=cd['markup_range_max'])

        return qs.order_by('name')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['form'] = AdvancedProductSearchForm(self.request.GET or None)
        return ctx

class ProductAttributeListView(LoginRequiredMixin, ListView):
    """List and search dynamic product attribute definitions."""
    model = ProductAttributeDefinition
    template_name = 'inventory/configuration/product_attribute_list.html'
    context_object_name = 'attributes'

    @method_decorator(inventory_permission_required('view'))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def get_queryset(self):
        qs = ProductAttributeDefinition.objects.all().order_by('display_order', 'name')
        q = self.request.GET.get('q', '').strip()
        if q:
            qs = qs.filter(Q(name__icontains=q) | Q(help_text__icontains=q))
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx.update({
            'page_title': 'Product Attributes',
            'total_active': ProductAttributeDefinition.objects.filter(is_active=True).count(),
        })
        return ctx

class ProductAttributeCreateView(LoginRequiredMixin, CreateView):
    """Create a new product attribute definition."""
    model = ProductAttributeDefinition
    form_class = ProductAttributeDefinitionForm
    template_name = 'inventory/configuration/product_attribute_form.html'
    success_url = reverse_lazy('inventory:product_attribute_list')

    @method_decorator(inventory_permission_required('edit'))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def form_valid(self, form):
        if hasattr(form.instance, 'created_by') and self.request.user.is_authenticated:
            form.instance.created_by = self.request.user
        messages.success(self.request, f'Attribute "{form.instance.name}" created successfully.')
        return super().form_valid(form)

class ProductAttributeUpdateView(LoginRequiredMixin, UpdateView):
    """Update an existing product attribute definition."""
    model = ProductAttributeDefinition
    form_class = ProductAttributeDefinitionForm
    template_name = 'inventory/configuration/product_attribute_form.html'
    success_url = reverse_lazy('inventory:product_attribute_list')

    @method_decorator(inventory_permission_required('edit'))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def form_valid(self, form):
        messages.success(self.request, f'Attribute "{form.instance.name}" updated successfully.')
        return super().form_valid(form)

class ComponentFamilyListView(BaseInventoryListView):
    model = ComponentFamily
    template_name = "inventory/configuration/component_family_list.html"
    context_object_name = "families"

class ComponentFamilyCreateView(BaseInventoryCreateView):
    model = ComponentFamily
    fields = [
        "name", "slug", "description", "default_attributes",
        "typical_markup_percentage", "default_bin_prefix",
        "is_active", "display_order",
    ]
    template_name = "inventory/configuration/component_family_form.html"
    success_url = reverse_lazy("inventory:component_family_list")

class ComponentFamilyUpdateView(BaseInventoryUpdateView):
    model = ComponentFamily
    fields = [
        "name", "slug", "description", "default_attributes",
        "typical_markup_percentage", "default_bin_prefix",
        "is_active", "display_order",
    ]
    template_name = "inventory/configuration/component_family_form.html"
    success_url = reverse_lazy("inventory:component_family_list")

class ComponentFamilyDeleteView(BaseInventoryDeleteView):
    model = ComponentFamily
    template_name = "inventory/confirm_delete.html"
    success_url = reverse_lazy("inventory:component_family_list")

@login_required
@inventory_permission_required('view')
def product_import_template_view(request):
    """
    Download import template with proper headers and sample data.
    """
    try:
        format_type = request.GET.get('format', 'csv').lower()
        
        if format_type == 'excel':
            return product_import_template_excel(request)
        else:
            return product_import_template_csv(request)
            
    except Exception as e:
        messages.error(request, f'Failed to generate template: {str(e)}')
        return redirect('inventory:product_bulk_import')

def product_import_template_csv(request):
    """Generate CSV import template"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="product_import_template.csv"'
    
    writer = csv.writer(response)
    
    # Headers
    writer.writerow([
        'sku', 'name', 'description', 'category', 'supplier', 'brand',
        'cost_price', 'selling_price', 'current_stock', 'reorder_level',
        'economic_order_quantity', 'barcode', 'is_active'
    ])
    
    # Sample data
    writer.writerow([
        'BT-2024-001', 'Sample Electronic Component', 'Sample description',
        'Electronics', 'ABC Supplier', 'Generic Brand',
        '10.00', '15.00', '100', '10', '50', '1234567890', 'True'
    ])
    
    writer.writerow([
        'BT-2024-002', 'Another Component', 'Another description',
        'Components', 'XYZ Supplier', 'Premium Brand',
        '25.50', '40.00', '50', '5', '25', '0987654321', 'True'
    ])
    
    return response

@login_required
@inventory_permission_required('view')
def product_export_view(request):
    """
    Export products with advanced filtering options.
    """
    try:
        # Get filters
        category = request.GET.get('category')
        supplier = request.GET.get('supplier')
        brand = request.GET.get('brand')
        stock_status = request.GET.get('stock_status')
        format_type = request.GET.get('format', 'csv')
        
        # Build queryset
        products = Product.objects.select_related('category', 'supplier', 'brand')
        
        if category:
            products = products.filter(category_id=category)
        if supplier:
            products = products.filter(supplier_id=supplier)
        if brand:
            products = products.filter(brand_id=brand)
        
        if stock_status == 'low_stock':
            products = products.filter(current_stock__lte=F('reorder_level'))
        elif stock_status == 'out_of_stock':
            products = products.filter(current_stock=0)
        elif stock_status == 'in_stock':
            products = products.filter(current_stock__gt=0)
        
        products = products.filter(is_active=True).order_by('name')
        
        if format_type == 'excel':
            return _export_products_excel(products)
        else:
            return _export_products_csv(products)
            
    except Exception as e:
        messages.error(request, f'Export failed: {str(e)}')
        return redirect('inventory:product_list')

def _export_products_csv(products):
    """Export products as CSV"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="products_export_{timezone.now().strftime("%Y%m%d")}.csv"'
    
    writer = csv.writer(response)
    
    # Headers
    writer.writerow([
        'SKU', 'Name', 'Description', 'Category', 'Supplier', 'Brand',
        'Cost Price', 'Selling Price', 'Current Stock', 'Reorder Level',
        'Economic Order Qty', 'Stock Value', 'Barcode', 'Last Restocked',
        'Is Active', 'Created Date'
    ])
    
    # Data
    for product in products:
        stock_value = product.current_stock * product.cost_price
        
        writer.writerow([
            product.sku,
            product.name,
            product.description,
            product.category.name if product.category else '',
            product.supplier.name if product.supplier else '',
            product.brand.name if product.brand else '',
            product.cost_price,
            product.selling_price,
            product.current_stock,
            product.reorder_level,
            product.economic_order_quantity or '',
            stock_value,
            product.barcode,
            product.last_restocked_date.strftime('%Y-%m-%d') if product.last_restocked_date else '',
            'Yes' if product.is_active else 'No',
            product.created_at.strftime('%Y-%m-%d'),
        ])
    
    return response

def _export_products_excel(products):
    """Export products as Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Products Export"
    
    # Headers with styling
    headers = [
        'SKU', 'Name', 'Description', 'Category', 'Supplier', 'Brand',
        'Cost Price', 'Selling Price', 'Current Stock', 'Reorder Level',
        'Economic Order Qty', 'Stock Value', 'Barcode', 'Last Restocked',
        'Is Active', 'Created Date'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Data
    for row_num, product in enumerate(products, 2):
        stock_value = product.current_stock * product.cost_price
        
        worksheet.cell(row=row_num, column=1, value=product.sku)
        worksheet.cell(row=row_num, column=2, value=product.name)
        worksheet.cell(row=row_num, column=3, value=product.description)
        worksheet.cell(row=row_num, column=4, value=product.category.name if product.category else '')
        worksheet.cell(row=row_num, column=5, value=product.supplier.name if product.supplier else '')
        worksheet.cell(row=row_num, column=6, value=product.brand.name if product.brand else '')
        worksheet.cell(row=row_num, column=7, value=float(product.cost_price))
        worksheet.cell(row=row_num, column=8, value=float(product.selling_price))
        worksheet.cell(row=row_num, column=9, value=product.current_stock)
        worksheet.cell(row=row_num, column=10, value=product.reorder_level)
        worksheet.cell(row=row_num, column=11, value=product.economic_order_quantity or '')
        worksheet.cell(row=row_num, column=12, value=float(stock_value))
        worksheet.cell(row=row_num, column=13, value=product.barcode)
        worksheet.cell(row=row_num, column=14, value=product.last_restocked_date.strftime('%Y-%m-%d') if product.last_restocked_date else '')
        worksheet.cell(row=row_num, column=15, value='Yes' if product.is_active else 'No')
        worksheet.cell(row=row_num, column=16, value=product.created_at.strftime('%Y-%m-%d'))
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="products_export_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    workbook.save(response)
    return response

def _export_catalog_csv(products):
    """Export product catalog as CSV"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="product_catalog_{timezone.now().strftime("%Y%m%d")}.csv"'
    
    writer = csv.writer(response)
    
    # Headers
    writer.writerow([
        'SKU', 'Product Name', 'Description', 'Category', 'Brand',
        'Price', 'In Stock', 'Barcode'
    ])
    
    # Data
    for product in products:
        writer.writerow([
            product.sku,
            product.name,
            product.description,
            product.category.name if product.category else '',
            product.brand.name if product.brand else '',
            product.selling_price,
            'Yes' if product.current_stock > 0 else 'No',
            product.barcode,
        ])
    
    return response

def _export_catalog_excel(products):
    """Export product catalog as Excel (.xlsx)"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Product Catalog"

    headers = [
        'SKU', 'Product Name', 'Description', 'Category', 'Brand',
        'Price', 'In Stock', 'Barcode'
    ]
    # Header row styling
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    # Data rows
    for r, p in enumerate(products, start=2):
        ws.cell(row=r, column=1, value=p.sku)
        ws.cell(row=r, column=2, value=p.name)
        ws.cell(row=r, column=3, value=p.description)
        ws.cell(row=r, column=4, value=p.category.name if p.category else '')
        ws.cell(row=r, column=5, value=p.brand.name if getattr(p, 'brand', None) else '')
        ws.cell(row=r, column=6, value=float(p.selling_price or 0))
        ws.cell(row=r, column=7, value='Yes' if (p.current_stock or 0) > 0 else 'No')
        ws.cell(row=r, column=8, value=p.barcode)

    # Best-effort column widths
    for column in ws.columns:
        try:
            max_len = max(len(str(c.value)) if c.value is not None else 0 for c in column)
        except ValueError:
            max_len = 10
        column[0].column_letter  # ensure first cell exists
        ws.column_dimensions[column[0].column_letter].width = min(max_len + 2, 50)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="product_catalog_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response

def _export_catalog_pdf(products):
    """Export product catalog as a simple PDF using WeasyPrint"""
    # You can switch to a real template later; this inline HTML keeps it dependency-light.
    html = f"""
    <html>
    <head>
    <meta charset="utf-8">
    <style>
        body {{ font-family: Arial, sans-serif; font-size: 12px; }}
        h1 {{ text-align: center; }}
        table {{ width: 100%; border-collapse: collapse; }}
        th, td {{ border: 1px solid #ddd; padding: 8px; }}
        th {{ background: #f2f2f2; text-align: left; }}
        tr:nth-child(even) {{ background: #fafafa; }}
    </style>
    </head>
    <body>
    <h1>Product Catalog</h1>
    <table>
        <thead>
        <tr>
            <th>SKU</th><th>Name</th><th>Category</th><th>Brand</th>
            <th>Price</th><th>In Stock</th><th>Barcode</th>
        </tr>
        </thead>
        <tbody>
        {''.join(
            f"<tr>"
            f"<td>{p.sku}</td>"
            f"<td>{p.name}</td>"
            f"<td>{p.category.name if p.category else ''}</td>"
            f"<td>{p.brand.name if p.brand else ''}</td>"
            f"<td>{float(p.selling_price or 0):.2f}</td>"
            f"<td>{'Yes' if p.current_stock > 0 else 'No'}</td>"
            f"<td>{p.barcode or ''}</td>"
            f"</tr>"
            for p in products
        )}
        </tbody>
    </table>
    </body>
    </html>
    """
    pdf_bytes = HTML(string=html).write_pdf()
    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="product_catalog_{timezone.now().strftime("%Y%m%d")}.pdf"'
    return response

@login_required
@inventory_permission_required('view')
def product_catalog_export_view(request):
    """
    Export product catalog for external use (customers, sales teams).
    """
    try:
        # Get only active products with selling prices
        products = Product.objects.filter(
            is_active=True,
            selling_price__gt=0
        ).select_related('category', 'brand').order_by('category__name', 'name')
        
        format_type = request.GET.get('format', 'pdf')
        
        if format_type == 'pdf':
            return _export_catalog_pdf(products)
        elif format_type == 'excel':
            return _export_catalog_excel(products)
        else:
            return _export_catalog_csv(products)
            
    except Exception as e:
        messages.error(request, f'Catalog export failed: {str(e)}')
        return redirect('inventory:product_list')

@login_required
@inventory_permission_required('edit')
def adjust_stock_view(self, request, pk):
    """
    Stock adjustment view for individual products.
    
    Allows authorized users to adjust stock levels with proper audit trail.
    """
    product = get_object_or_404(Product, pk=pk)
    
    if request.method == 'POST':
        form = StockAdjustmentForm(request.POST, user=request.user)
        
        if form.is_valid():
            try:
                product = form.cleaned_data['product']
                quantity = int(form.cleaned_data['quantity'])
                adjustment_type = form.cleaned_data['adjustment_type']  # 'set' | 'add' | 'subtract'
                location = form.cleaned_data.get('location')
                reason = form.cleaned_data.get('reason', '')
                notes = form.cleaned_data.get('notes', '')
                
                if location:
                    stock_level, _ = StockLevel.objects.get_or_create(
                        product=product,
                        location=location,
                        defaults={'quantity': 0}
                    )
                    previous_stock = stock_level.quantity
                else:
                    previous_stock = product.current_stock
                
                # Calculate actual adjustment quantity
                if adjustment_type == 'set':
                    new_stock = quantity
                    actual_adjustment = new_stock - previous_stock
                elif adjustment_type == 'add':
                    new_stock = previous_stock + quantity
                    actual_adjustment = quantity
                else:  # 'subtract'
                    new_stock = max(0, previous_stock - quantity)
                    actual_adjustment = -(previous_stock - new_stock)
                
                if location:
                    stock_level.quantity = new_stock
                    stock_level.save()
                else:
                    product.current_stock = new_stock
                    product.save()
                    
                # Check for negative stock
                if product.current_stock + actual_adjustment < 0:
                    messages.error(
                        request,
                        f'Cannot adjust stock to negative value. Current stock: {product.current_stock}'
                    )
                    return render(request, 'inventory/product/adjust_stock.html', {
                        'product': product,
                        'form': form
                    })
                
                # Create stock movement
                StockMovement.objects.create(
                    product=product,
                    movement_type='adjustment',
                    quantity=actual_adjustment,
                    from_location=location if actual_adjustment < 0 else None,
                    to_location=location if actual_adjustment > 0 else None,
                    previous_stock=previous_stock,
                    new_stock=new_stock,
                    reference=f"ADJ-{timezone.now().strftime('%Y%m%d%H%M%S')}",
                    notes=f"Reason: {reason}. {notes}",
                    created_by=self.request.user
                )

                messages.success(self.request, f"Stock adjusted for {product.name}: {previous_stock} → {new_stock}")
                
                logger.info(
                    f"Stock adjustment: {actual_adjustment} units of {product.sku} "
                    f"by {request.user.username}. Reason: {reason}"
                )
                
                return redirect('inventory:product_detail', pk=product.pk)
                
            except Exception as e:
                logger.error(f"Error adjusting stock: {str(e)}")
                messages.error(request, f'Error adjusting stock: {str(e)}')
        
    else:
        form = StockAdjustmentForm(initial={'product': product}, user=request.user)
    
    context = {
        'product': product,
        'form': form,
        'stock_levels': StockLevel.objects.filter(product=product).select_related('location')
    }
    
    return render(request, 'inventory/product/adjust_stock.html', context)

@login_required
@inventory_permission_required('edit')
def product_import_view(request):
    """
    Product bulk import page (shows the wizard and handles CSV upload).
    If you want to support AJAX/Excel/validation steps, extend this!
    """
    if request.method == "POST" and request.FILES.get('import_file'):
        csv_file = request.FILES['import_file']
        try:
            decoded = csv_file.read().decode('utf-8').splitlines()
            reader = csv.DictReader(decoded)
            created, updated = 0, 0
            for row in reader:
                sku = row.get('sku') or row.get('SKU')
                name = row.get('name') or row.get('Name')
                if not sku or not name:
                    continue
                product, created_flag = Product.objects.update_or_create(
                    sku=sku,
                    defaults={
                        'name': name,
                        'cost_price': row.get('cost_price', 0),
                        'selling_price': row.get('selling_price', 0),
                        # Extend for more fields as needed!
                        'created_by': request.user,
                        'is_active': True,
                    }
                )
                if created_flag:
                    created += 1
                else:
                    updated += 1
            messages.success(request, f'Imported {created} new products, updated {updated}.')
            return redirect('inventory:product_list')
        except Exception as e:
            messages.error(request, f"Import failed: {str(e)}")

    return render(request, "inventory/product/product_import.html", {})

@login_required
@inventory_permission_required('view')
def product_import_template_excel(request):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Product Import Template"

    headers = [
        'sku', 'name', 'category', 'supplier', 'description', 'cost_price',
        'selling_price', 'current_stock', 'reorder_level', 'barcode', 'brand'
    ]
    sheet.append(headers)
    sheet.append([
        'BT-2024-001', 'Sample Product', 'Electronics', 'ABC Supplier', 'Sample desc',
        '10.00', '15.00', '100', '10', '0123456789012', 'Generic'
    ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=product_import_template.xlsx'
    workbook.save(response)
    return response

@login_required
@inventory_permission_required('view')
def export_data_view(request):
    # Export all products as CSV for now
    products = Product.objects.filter(is_active=True).select_related('category', 'supplier')
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="all_products.csv"'

    writer = csv.writer(response)
    writer.writerow([
        'SKU', 'Name', 'Category', 'Supplier', 'Cost Price', 'Selling Price', 'Stock', 'Barcode', 'Is Active'
    ])
    for p in products:
        writer.writerow([
            p.sku, p.name, p.category.name if p.category else '', p.supplier.name if p.supplier else '',
            p.cost_price, p.selling_price, p.current_stock, p.barcode, p.is_active
        ])
    return response

@login_required
@inventory_permission_required('edit')
def product_duplicate_view(request, pk):
    """
    Create a duplicate of an existing product for quick setup.
    """
    try:
        original_product = get_object_or_404(Product, pk=pk)
        
        if request.method == 'POST':
            # Create duplicate
            duplicate = Product.objects.create(
                name=f"{original_product.name} (Copy)",
                sku=f"{original_product.sku}-COPY-{timezone.now().strftime('%Y%m%d%H%M%S')}",
                description=original_product.description,
                category=original_product.category,
                supplier=original_product.supplier,
                brand=original_product.brand,
                component_family=original_product.component_family,
                cost_price=original_product.cost_price,
                selling_price=original_product.selling_price,
                reorder_level=original_product.reorder_level,
                economic_order_quantity=original_product.economic_order_quantity,
                created_by=request.user,
                current_stock=0,  # Don't copy stock
                is_active=False,  # Start inactive for review
            )
            
            messages.success(request, f'Product duplicated successfully. Please review and update the details.')
            return redirect('inventory:product_edit', pk=duplicate.pk)
        
        return render(request, 'inventory/product/product_duplicate_confirm.html', {
            'page_title': 'Duplicate Product',
            'original_product': original_product,
        })
        
    except Exception as e:
        messages.error(request, f'Failed to duplicate product: {str(e)}')
        return redirect('inventory:product_detail', pk=pk)

@login_required
@inventory_permission_required('view')
def product_cost_analysis_view(request, pk):
    """
    Detailed cost analysis for a specific product.
    Shows breakdown of costs, margins, and profitability.
    """
    try:
        product = get_object_or_404(Product, pk=pk)
        
        # Calculate cost breakdown
        base_cost = product.cost_price
        overhead_factors = OverheadFactor.objects.filter(is_active=True)
        
        total_overhead_rate = Decimal('0')
        overhead_breakdown = []
        for factor in overhead_factors:
            overhead_amount = base_cost * (factor.percentage / 100)
            overhead_breakdown.append({
                'name': factor.name,
                'percentage': factor.percentage,
                'amount': overhead_amount,
            })
            total_overhead_rate += factor.percentage
        
        total_overhead_cost = base_cost * (total_overhead_rate / 100)
        total_cost = base_cost + total_overhead_cost
        
        # Calculate margins
        gross_margin = product.selling_price - base_cost
        net_margin = product.selling_price - total_cost
        gross_margin_percent = (gross_margin / product.selling_price * 100) if product.selling_price > 0 else 0
        net_margin_percent = (net_margin / product.selling_price * 100) if product.selling_price > 0 else 0
        
        # Sales data (last 6 months)
        six_months_ago = timezone.now() - timedelta(days=180)
        sales_data = product.stock_movements.filter(
            movement_type='sale',
            created_at__gte=six_months_ago
        ).aggregate(
            total_units=Sum('quantity'),
            total_revenue=Sum(F('quantity') * product.selling_price)
        )
        
        return render(request, 'inventory/product/product_cost_analysis.html', {
            'page_title': f'Cost Analysis: {product.name}',
            'product': product,
            'base_cost': base_cost,
            'overhead_breakdown': overhead_breakdown,
            'total_overhead_cost': total_overhead_cost,
            'total_cost': total_cost,
            'gross_margin': gross_margin,
            'net_margin': net_margin,
            'gross_margin_percent': round(gross_margin_percent, 2),
            'net_margin_percent': round(net_margin_percent, 2),
            'sales_data': sales_data,
        })
        
    except Exception as e:
        messages.error(request, f'Failed to generate cost analysis: {str(e)}')
        return redirect('inventory:product_detail', pk=pk)

@login_required
@bulk_operation_permission
def product_bulk_update_view(request):
    """
    Bulk update multiple products at once.
    """
    if request.method == 'POST':
        try:
            updates_data = json.loads(request.POST.get('updates', '[]'))
            
            if not updates_data:
                messages.error(request, 'No update data provided')
                return redirect('inventory:product_list')
            
            updated_count = 0
            with transaction.atomic():
                for update in updates_data:
                    try:
                        product = Product.objects.get(id=update['product_id'])
                        
                        # Update fields if provided
                        if 'cost_price' in update:
                            product.cost_price = Decimal(str(update['cost_price']))
                        if 'selling_price' in update:
                            product.selling_price = Decimal(str(update['selling_price']))
                        if 'reorder_level' in update:
                            product.reorder_level = int(update['reorder_level'])
                        if 'category_id' in update and update['category_id']:
                            product.category = Category.objects.get(id=update['category_id'])
                        if 'supplier_id' in update and update['supplier_id']:
                            product.supplier = Supplier.objects.get(id=update['supplier_id'])
                        
                        product.save()
                        updated_count += 1
                        
                    except Exception as e:
                        logger.error(f"Failed to update product {update.get('product_id')}: {str(e)}")
                        continue
            
            messages.success(request, f'Successfully updated {updated_count} products')
            
        except Exception as e:
            messages.error(request, f'Bulk update failed: {str(e)}')
    
    return redirect('inventory:product_list')

@login_required
@inventory_permission_required('view')
def product_details_api(request, product_id):
    """
    API endpoint: Get detailed product information including stock levels.
    Used for quote system integration and mobile apps.
    """
    try:
        product = get_object_or_404(Product, id=product_id, is_active=True)
        
        # Get stock levels by location
        stock_levels = []
        for stock_level in product.stock_levels.select_related('location'):
            stock_levels.append({
                'location_id': stock_level.location.id,
                'location_name': stock_level.location.name,
                'quantity': stock_level.quantity,
                'available_quantity': stock_level.available_quantity,
                'reserved_quantity': stock_level.reserved_quantity,
                'last_counted': stock_level.last_counted.isoformat() if stock_level.last_counted else None,
            })
        
        # Calculate total cost including overheads
        overhead_factors = OverheadFactor.objects.filter(is_active=True)
        total_overhead_rate = sum(factor.percentage for factor in overhead_factors)
        total_cost = product.cost_price * (1 + total_overhead_rate / 100)
        
        # Get recent movements
        recent_movements = []
        for movement in product.stock_movements.select_related('from_location', 'to_location')[:10]:
            recent_movements.append({
                'date': movement.created_at.isoformat(),
                'type': movement.movement_type,
                'quantity': movement.quantity,
                'reference': movement.reference,
                'from_location': movement.from_location.name if movement.from_location else None,
                'to_location': movement.to_location.name if movement.to_location else None,
            })
        
        data = {
            'id': product.id,
            'sku': product.sku,
            'name': product.name,
            'description': product.description,
            'category': {
                'id': product.category.id,
                'name': product.category.name
            } if product.category else None,
            'supplier': {
                'id': product.supplier.id,
                'name': product.supplier.name,
                'currency': product.supplier.currency,
                'lead_time_days': product.supplier.average_lead_time_days,
            } if product.supplier else None,
            'brand': {
                'id': product.brand.id,
                'name': product.brand.name
            } if product.brand else None,
            'pricing': {
                'cost_price': float(product.cost_price),
                'selling_price': float(product.selling_price),
                'total_cost_with_overhead': float(total_cost),
                'margin': float(product.selling_price - total_cost),
                'margin_percent': float((product.selling_price - total_cost) / product.selling_price * 100) if product.selling_price > 0 else 0,
            },
            'stock': {
                'current_stock': product.current_stock,
                'reorder_level': product.reorder_level,
                'economic_order_quantity': product.economic_order_quantity,
                'last_restocked': product.last_restocked_date.isoformat() if product.last_restocked_date else None,
                'stock_status': get_stock_status(product),
                'stock_levels_by_location': stock_levels,
            },
            'barcode': product.barcode,
            'is_active': product.is_active,
            'created_at': product.created_at.isoformat(),
            'recent_movements': recent_movements,
        }
        
        return JsonResponse({'success': True, 'product': data})
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

# =====================================
# STOCK MANAGEMENT VIEWS
# =====================================

class StockOverviewView(LoginRequiredMixin, TemplateView):
    """
    Complete stock overview across all locations and products.
    Shows current stock levels, values, and alerts.
    """
    template_name = 'inventory/stock/stock_overview.html'
    
    @method_decorator(inventory_permission_required('view'))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get all stock levels with related data
        stock_levels = StockLevel.objects.select_related(
            'product', 'location'
        ).filter(quantity__gt=0)
        
        # Calculate summary metrics
        total_products = Product.objects.filter(is_active=True).count()
        total_stock_value = sum(sl.stock_value for sl in stock_levels)
        low_stock_count = Product.objects.filter(
            current_stock__lte=F('reorder_level'),
            is_active=True
        ).count()
        out_of_stock_count = Product.objects.filter(
            current_stock=0,
            is_active=True
        ).count()
        
        # Get stock by location
        locations = Location.objects.filter(is_active=True)
        stock_by_location = []
        for location in locations:
            location_stock = stock_levels.filter(location=location)
            location_value = sum(sl.stock_value for sl in location_stock)
            stock_by_location.append({
                'location': location,
                'product_count': location_stock.count(),
                'total_value': location_value,
                'stock_levels': location_stock[:10]  # Show top 10
            })
        
        # Get recent movements
        recent_movements = StockMovement.objects.select_related(
            'product', 'from_location', 'to_location'
        ).order_by('-created_at')[:20]
        
        context.update({
            'page_title': 'Stock Overview',
            'total_products': total_products,
            'total_stock_value': total_stock_value,
            'low_stock_count': low_stock_count,
            'out_of_stock_count': out_of_stock_count,
            'stock_by_location': stock_by_location,
            'recent_movements': recent_movements,
            'locations': locations,
        })
        return context

class StockByLocationView(LoginRequiredMixin, DetailView):
    """
    Detailed view of stock at a specific location.
    """
    model = Location
    template_name = 'inventory/stock/stock_by_location.html'
    context_object_name = 'location'
    
    @method_decorator(inventory_permission_required('view'))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        location = self.get_object()
        
        # Get stock levels for this location
        stock_levels = StockLevel.objects.filter(
            location=location
        ).select_related('product', 'product__category', 'product__supplier')
        
        # Calculate metrics
        total_value = sum(sl.stock_value for sl in stock_levels)
        total_products = stock_levels.count()
        
        context.update({
            'page_title': f'Stock at {location.name}',
            'stock_levels': stock_levels,
            'total_value': total_value,
            'total_products': total_products,
        })
        return context

class StockAdjustmentView(LoginRequiredMixin, FormView):
    """
    Adjust stock levels for products with proper audit trail.
    Supports single and bulk adjustments.
    """
    template_name = 'inventory/stock/stock_adjustment.html'
    form_class = StockAdjustmentForm
    
    @method_decorator(stock_adjustment_permission)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get products that might need adjustment
        products = Product.objects.filter(is_active=True).select_related(
            'category', 'supplier'
        ).order_by('name')
        
        locations = Location.objects.filter(is_active=True)
        
        context.update({
            'page_title': 'Stock Adjustment',
            'products': products,
            'locations': locations,
        })
        return context
    
    def form_valid(self, form):
        try:
            with transaction.atomic():
                product = form.cleaned_data['product']
                location = form.cleaned_data.get('location')
                adjustment_type = form.cleaned_data['adjustment_type']
                quantity = form.cleaned_data['quantity']
                reason = form.cleaned_data['reason']
                notes = form.cleaned_data.get('notes', '')
                
                # Get or create stock level
                if location:
                    stock_level, created = StockLevel.objects.get_or_create(
                        product=product,
                        location=location,
                        defaults={'quantity': 0}
                    )
                    previous_stock = stock_level.quantity
                else:
                    previous_stock = product.current_stock
                
                # Calculate new stock based on adjustment type
                if adjustment_type == 'set':
                    new_stock = quantity
                    actual_adjustment = new_stock - previous_stock
                elif adjustment_type == 'increase':
                    new_stock = previous_stock + quantity
                    actual_adjustment = quantity
                else:  # decrease
                    new_stock = max(0, previous_stock - quantity)
                    actual_adjustment = -(previous_stock - new_stock)
                
                # Update stock levels
                if location:
                    stock_level.quantity = new_stock
                    stock_level.last_counted = timezone.now()
                    stock_level.save()
                    
                    # Update product total stock
                    product.current_stock = product.stock_levels.aggregate(
                        total=Sum('quantity')
                    )['total'] or 0
                else:
                    product.current_stock = new_stock
                
                product.last_stock_check = timezone.now()
                product.save()
                
                # Create stock movement record
                StockMovement.objects.create(
                    product=product,
                    movement_type='adjustment',
                    quantity=actual_adjustment,
                    from_location=location if actual_adjustment < 0 else None,
                    to_location=location if actual_adjustment > 0 else None,
                    previous_stock=previous_stock,
                    new_stock=new_stock,
                    reference=f"ADJ-{timezone.now().strftime('%Y%m%d%H%M%S')}",
                    notes=f"Reason: {reason}. {notes}",
                    created_by=self.request.user
                )
                
                messages.success(
                    self.request, 
                    f'Stock adjusted for {product.name}: {previous_stock} → {new_stock}'
                )
                
                return redirect('inventory:stock_overview')
                
        except Exception as e:
            messages.error(self.request, f'Stock adjustment failed: {str(e)}')
            return self.form_invalid(form)

class BulkStockAdjustmentView(LoginRequiredMixin, TemplateView):
    """
    Bulk stock adjustment interface for efficient operations.
    """
    template_name = 'inventory/stock/bulk_stock_adjustment.html'
    
    @method_decorator(bulk_operation_permission)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        products = Product.objects.filter(is_active=True).select_related(
            'category', 'supplier'
        ).order_by('name')
        
        locations = Location.objects.filter(is_active=True)
        
        context.update({
            'page_title': 'Bulk Stock Adjustment',
            'products': products,
            'locations': locations,
        })
        return context
    
    def post(self, request, *args, **kwargs):
        try:
            adjustments_data = json.loads(request.POST.get('adjustments', '[]'))
            
            if not adjustments_data:
                messages.error(request, 'No adjustments data provided')
                return redirect('inventory:bulk_stock_adjustment')
            
            success_count = 0
            with transaction.atomic():
                for adj in adjustments_data:
                    try:
                        product = Product.objects.get(id=adj['product_id'])
                        location_id = adj.get('location_id')
                        location = Location.objects.get(id=location_id) if location_id else None
                        
                        # Get current stock
                        if location:
                            stock_level, created = StockLevel.objects.get_or_create(
                                product=product,
                                location=location,
                                defaults={'quantity': 0}
                            )
                            previous_stock = stock_level.quantity
                        else:
                            previous_stock = product.current_stock
                        
                        new_stock = adj['new_quantity']
                        adjustment = new_stock - previous_stock
                        
                        # Update stock
                        if location:
                            stock_level.quantity = new_stock
                            stock_level.last_counted = timezone.now()
                            stock_level.save()
                            
                            # Update product total
                            product.current_stock = product.stock_levels.aggregate(
                                total=Sum('quantity')
                            )['total'] or 0
                        else:
                            product.current_stock = new_stock
                        
                        product.last_stock_check = timezone.now()
                        product.save()
                        
                        # Create movement record
                        StockMovement.objects.create(
                            product=product,
                            movement_type='adjustment',
                            quantity=adjustment,
                            from_location=location if adjustment < 0 else None,
                            to_location=location if adjustment > 0 else None,
                            previous_stock=previous_stock,
                            new_stock=new_stock,
                            reference=f"BULK-ADJ-{timezone.now().strftime('%Y%m%d%H%M%S')}",
                            notes=f"Bulk adjustment: {adj.get('reason', 'No reason provided')}",
                            created_by=request.user
                        )
                        
                        success_count += 1
                        
                    except Exception as e:
                        logger.error(f"Bulk adjustment failed for item {adj}: {str(e)}")
                        continue
            
            messages.success(
                request, 
                f'Successfully adjusted {success_count} stock items'
            )
            return redirect('inventory:stock_overview')
            
        except Exception as e:
            messages.error(request, f'Bulk adjustment failed: {str(e)}')
            return redirect('inventory:bulk_stock_adjustment')

class StockTransferView(LoginRequiredMixin, FormView):
    """
    Transfer stock between locations with proper tracking.
    """
    template_name = 'inventory/stock/stock_transfer.html'
    form_class = StockTransferForm
    
    @method_decorator(stock_adjustment_permission)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        locations = Location.objects.filter(is_active=True)
        products = Product.objects.filter(
            is_active=True,
            current_stock__gt=0
        ).select_related('category', 'supplier')
        
        context.update({
            'page_title': 'Stock Transfer',
            'locations': locations,
            'products': products,
        })
        return context
    
    def form_valid(self, form):
        try:
            with transaction.atomic():
                product = form.cleaned_data['product']
                from_location = form.cleaned_data['from_location']
                to_location = form.cleaned_data['to_location']
                quantity = form.cleaned_data['quantity']
                notes = form.cleaned_data.get('notes', '')
                
                # Validate transfer
                if from_location == to_location:
                    messages.error(self.request, 'Cannot transfer to the same location')
                    return self.form_invalid(form)
                
                # Get source stock level
                try:
                    from_stock = StockLevel.objects.get(
                        product=product,
                        location=from_location
                    )
                except StockLevel.DoesNotExist:
                    messages.error(
                        self.request, 
                        f'No stock found for {product.name} at {from_location.name}'
                    )
                    return self.form_invalid(form)
                
                if from_stock.available_quantity < quantity:
                    messages.error(
                        self.request,
                        f'Insufficient stock. Available: {from_stock.available_quantity}, Requested: {quantity}'
                    )
                    return self.form_invalid(form)
                
                # Update source location
                from_stock.quantity -= quantity
                from_stock.last_movement = timezone.now()
                from_stock.save()
                
                # Update destination location
                to_stock, created = StockLevel.objects.get_or_create(
                    product=product,
                    location=to_location,
                    defaults={'quantity': 0}
                )
                to_stock.quantity += quantity
                to_stock.last_movement = timezone.now()
                to_stock.save()
                
                # Create transfer reference
                transfer_ref = f"TXF-{timezone.now().strftime('%Y%m%d%H%M%S')}"
                
                # Create stock movements (outbound)
                StockMovement.objects.create(
                    product=product,
                    movement_type='transfer',
                    quantity=-quantity,
                    from_location=from_location,
                    to_location=None,
                    previous_stock=from_stock.quantity + quantity,
                    new_stock=from_stock.quantity,
                    reference=transfer_ref,
                    notes=f"Transfer out to {to_location.name}. {notes}",
                    created_by=self.request.user
                )
                
                # Create stock movements (inbound)
                StockMovement.objects.create(
                    product=product,
                    movement_type='transfer',
                    quantity=quantity,
                    from_location=None,
                    to_location=to_location,
                    previous_stock=to_stock.quantity - quantity,
                    new_stock=to_stock.quantity,
                    reference=transfer_ref,
                    notes=f"Transfer in from {from_location.name}. {notes}",
                    created_by=self.request.user
                )
                
                messages.success(
                    self.request,
                    f'Successfully transferred {quantity} units of {product.name} from {from_location.name} to {to_location.name}'
                )
                
                return redirect('inventory:stock_overview')
                
        except Exception as e:
            messages.error(self.request, f'Stock transfer failed: {str(e)}')
            return self.form_invalid(form)

class StockMovementListView(LoginRequiredMixin, ListView):
    """
    Complete audit trail of all stock movements.
    """
    model = StockMovement
    template_name = 'inventory/stock/stock_movements.html'
    context_object_name = 'movements'
    paginate_by = 50
    
    @method_decorator(inventory_permission_required('view'))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    
    def get_queryset(self):
        queryset = StockMovement.objects.select_related(
            'product', 'from_location', 'to_location', 'created_by'
        ).order_by('-created_at')
        
        # Apply filters
        product_id = self.request.GET.get('product')
        if product_id:
            queryset = queryset.filter(product_id=product_id)
        
        movement_type = self.request.GET.get('movement_type')
        if movement_type:
            queryset = queryset.filter(movement_type=movement_type)
        
        location_id = self.request.GET.get('location')
        if location_id:
            queryset = queryset.filter(
                Q(from_location_id=location_id) | Q(to_location_id=location_id)
            )
        
        date_from = self.request.GET.get('date_from')
        if date_from:
            queryset = queryset.filter(created_at__date__gte=date_from)
        
        date_to = self.request.GET.get('date_to')
        if date_to:
            queryset = queryset.filter(created_at__date__lte=date_to)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        context.update({
            'page_title': 'Stock Movement History',
            'products': Product.objects.filter(is_active=True).order_by('name'),
            'locations': Location.objects.filter(is_active=True).order_by('name'),
            'movement_types': StockMovement.MOVEMENT_TYPES,
            'filters': {
                'product': self.request.GET.get('product'),
                'movement_type': self.request.GET.get('movement_type'),
                'location': self.request.GET.get('location'),
                'date_from': self.request.GET.get('date_from'),
                'date_to': self.request.GET.get('date_to'),
            }
        })
        return context

class StockAdjustDashboardView(LoginRequiredMixin, TemplateView):
    template_name = "inventory/stock/stock_adjust_dashboard.html"

    @method_decorator(inventory_permission_required('edit'))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        products = Product.objects.filter(is_active=True)
        categories = Category.objects.filter(is_active=True)
        locations = Location.objects.filter(is_active=True)

        # You could also support filtering here (e.g., by GET params)

        context.update({
            "products": products,
            "categories": categories,
            "locations": locations,
            "stock_adjustment_form": StockAdjustmentForm(),
        })
        return context

class LowStockReportView(LoginRequiredMixin, TemplateView):
    template_name = "inventory/stock/low_stock.html"

    @method_decorator(inventory_permission_required('view'))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        low_stock_products = Product.objects.filter(
            is_active=True,
            current_stock__lte=F('reorder_level'),
            current_stock__gt=0
        ).select_related('category', 'supplier').order_by('current_stock')
        context['low_stock_products'] = low_stock_products
        return context

class LowStockOrderingView(LoginRequiredMixin, TemplateView):
    """
    Generate and manage purchase orders from low stock alerts.
    This is critical for business continuity - prevents stockouts.
    """
    template_name = 'inventory/reorder/low_stock_ordering.html'
    
    @method_decorator(purchase_order_permission)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get products that need reordering
        low_stock_products = Product.objects.filter(
            current_stock__lte=F('reorder_level'),
            is_active=True
        ).select_related('category', 'supplier').annotate(
            stock_shortage=F('reorder_level') - F('current_stock'),
            suggested_order_qty=F('economic_order_quantity')
        )
        
        # Group by supplier for efficient ordering
        suppliers_data = {}
        for product in low_stock_products:
            if product.supplier:
                supplier = product.supplier
                if supplier not in suppliers_data:
                    suppliers_data[supplier] = {
                        'products': [],
                        'total_cost': Decimal('0'),
                        'currency': supplier.currency,
                        'lead_time': supplier.average_lead_time_days,
                    }
                
                order_quantity = max(
                    product.suggested_order_qty or product.economic_order_quantity,
                    product.stock_shortage
                )
                estimated_cost = order_quantity * product.cost_price
                
                suppliers_data[supplier]['products'].append({
                    'product': product,
                    'current_stock': product.current_stock,
                    'reorder_level': product.reorder_level,
                    'shortage': product.stock_shortage,
                    'suggested_qty': order_quantity,
                    'estimated_cost': estimated_cost,
                })
                suppliers_data[supplier]['total_cost'] += estimated_cost
        
        # Get existing reorder alerts
        active_alerts = ReorderAlert.objects.filter(
            status__in=['active', 'acknowledged']
        ).select_related('product', 'suggested_supplier').order_by('-priority', '-created_at')
        
        # Calculate summary metrics
        total_products_needing_reorder = low_stock_products.count()
        total_estimated_cost = sum(data['total_cost'] for data in suppliers_data.values())
        critical_shortages = low_stock_products.filter(current_stock=0).count()
        
        context.update({
            'page_title': 'Low Stock Ordering',
            'suppliers_data': suppliers_data,
            'active_alerts': active_alerts,
            'total_products_needing_reorder': total_products_needing_reorder,
            'total_estimated_cost': total_estimated_cost,
            'critical_shortages': critical_shortages,
        })
        return context
    
    def post(self, request, *args, **kwargs):
        """Generate purchase orders from selected products"""
        try:
            selected_items = json.loads(request.POST.get('selected_items', '[]'))
            
            if not selected_items:
                messages.error(request, 'No items selected for ordering')
                return redirect('inventory:low_stock_ordering')
            
            # Group items by supplier
            supplier_orders = {}
            for item in selected_items:
                product_id = item['product_id']
                supplier_id = item['supplier_id']
                quantity = int(item['quantity'])
                
                if supplier_id not in supplier_orders:
                    supplier_orders[supplier_id] = []
                
                supplier_orders[supplier_id].append({
                    'product_id': product_id,
                    'quantity': quantity,
                })
            
            # Create purchase orders
            created_orders = []
            with transaction.atomic():
                for supplier_id, items in supplier_orders.items():
                    supplier = Supplier.objects.get(id=supplier_id)
                    
                    # Generate PO number
                    po_number = f"PO-{timezone.now().strftime('%Y%m%d')}-{supplier.supplier_code}"
                    counter = 1
                    while PurchaseOrder.objects.filter(po_number=po_number).exists():
                        po_number = f"PO-{timezone.now().strftime('%Y%m%d')}-{supplier.supplier_code}-{counter:02d}"
                        counter += 1
                    
                    # Calculate expected delivery
                    expected_delivery = timezone.now().date() + timedelta(
                        days=supplier.average_lead_time_days or 7
                    )
                    
                    # Create PO
                    purchase_order = PurchaseOrder.objects.create(
                        po_number=po_number,
                        supplier=supplier,
                        status='draft',
                        order_date=timezone.now().date(),
                        expected_delivery_date=expected_delivery,
                        currency=supplier.currency,
                        payment_terms=supplier.payment_terms,
                        notes=f"Auto-generated from low stock ordering on {timezone.now().strftime('%Y-%m-%d')}",
                        created_by=request.user,
                    )
                    
                    # Add items to PO
                    total_amount = Decimal('0')
                    for item in items:
                        product = Product.objects.get(id=item['product_id'])
                        quantity = item['quantity']
                        unit_price = product.cost_price
                        total_price = quantity * unit_price
                        
                        PurchaseOrderItem.objects.create(
                            purchase_order=purchase_order,
                            product=product,
                            quantity_ordered=quantity,
                            unit_price=unit_price,
                            total_price=total_price,
                            expected_delivery_date=expected_delivery,
                        )
                        
                        total_amount += total_price
                    
                    # Update PO totals
                    purchase_order.subtotal = total_amount
                    purchase_order.total_amount = total_amount  # Add tax calculation if needed
                    purchase_order.save()
                    
                    created_orders.append(purchase_order)
                    
                    # Update reorder alerts
                    for item in items:
                        ReorderAlert.objects.filter(
                            product_id=item['product_id'],
                            status__in=['active', 'acknowledged']
                        ).update(
                            status='ordered',
                            purchase_order=purchase_order,
                            resolved_at=timezone.now()
                        )
            
            messages.success(
                request,
                f'Successfully created {len(created_orders)} purchase orders'
            )
            return redirect('inventory:purchase_order_list')
            
        except Exception as e:
            messages.error(request, f'Failed to create purchase orders: {str(e)}')
            return redirect('inventory:low_stock_ordering')

class StockLevelListView(LoginRequiredMixin, ListView):
    """List stock levels across locations"""
    model = StockLevel
    template_name = 'inventory/stock/stock_level_list.html'
    context_object_name = 'stock_levels'
    paginate_by = 50
    
    @method_decorator(inventory_permission_required('view'))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    
    def get_queryset(self):
        return StockLevel.objects.select_related(
            'product', 'location'
        ).filter(product__is_active=True).order_by('product__name')

@login_required
@stock_adjustment_permission
def stock_adjustment_view(request):
    """Adjust stock levels"""
    if request.method == 'POST':
        form = StockAdjustmentForm(request.POST)
        if form.is_valid():
            # Process stock adjustment
            adjustment = form.save(commit=False)
            adjustment.adjusted_by = request.user
            adjustment.save()
            
            # Create stock movement record
            create_stock_movement(
                product=adjustment.product,
                movement_type='adjustment',
                quantity=adjustment.adjustment_quantity,
                user=request.user,
                notes=adjustment.notes
            )
            
            messages.success(request, 'Stock adjustment completed successfully')
            return redirect('inventory:stock_level_list')
    else:
        form = StockAdjustmentForm()
    
    context = {
        'page_title': 'Stock Adjustment',
        'form': form,
    }
    return render(request, 'inventory/stock/stock_adjustment.html', context)

@login_required
@inventory_permission_required('view_stocktake')
def export_stock_take_pdf(request, pk):
    stock_take = StockTake.objects.get(pk=pk)
    html_string = render_to_string('inventory/stock_take_pdf.html', {'stock_take': stock_take})
    pdf_file = HTML(string=html_string).write_pdf()

    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="stock_take_{stock_take.reference}.pdf"'
    return response

@login_required
@inventory_permission_required('view_stocktake')
def export_stock_take_excel(request, pk):
    stock_take = StockTake.objects.get(pk=pk)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = f"Stock Take {stock_take.reference}"

    sheet.append(["SKU", "Product", "Expected", "Counted", "Variance", "Location", "Status"])
    for item in stock_take.items.select_related('product', 'location'):
        sheet.append([
            item.product.sku,
            item.product.name,
            item.expected_quantity,
            item.counted_quantity,
            item.variance,
            item.location.name if item.location else "N/A",
            item.status,
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"stock_take_{stock_take.reference}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook.save(response)
    return response

@login_required
def get_live_stock(request, product_id):
    try:
        product = Product.objects.get(id=product_id)
        return JsonResponse({'sku': product.sku, 'name': product.name, 'current_stock': product.current_stock})
    except Product.DoesNotExist:
        return JsonResponse({'error': 'Product not found'}, status=404)

@login_required
@inventory_permission_required('edit')
def stock_transfer_view(request):
    """Stock transfer between locations"""
    if request.method == 'POST':
        form = StockTransferForm(request.POST)
        
        if form.is_valid():
            try:
                from .utils import transfer_stock_between_locations
                
                product = form.cleaned_data['product']
                from_location = form.cleaned_data['from_location']
                to_location = form.cleaned_data['to_location']
                quantity = form.cleaned_data['quantity']
                reference = form.cleaned_data['reference']
                notes = form.cleaned_data['notes']
                
                # Perform transfer
                outgoing, incoming = transfer_stock_between_locations(
                    product=product,
                    from_location=from_location,
                    to_location=to_location,
                    quantity=quantity,
                    reference=reference,
                    user=request.user,
                    notes=notes
                )
                
                messages.success(
                    request,
                    f'Successfully transferred {quantity} units of {product.name} '
                    f'from {from_location.name} to {to_location.name}'
                )
                
                logger.info(
                    f"Stock transfer: {quantity} units of {product.sku} "
                    f"from {from_location.name} to {to_location.name} "
                    f"by {request.user.username}"
                )
                
                return redirect('inventory:stock_overview')
                
            except Exception as e:
                logger.error(f"Error in stock transfer: {str(e)}")
                messages.error(request, f'Transfer failed: {str(e)}')
    else:
        form = StockTransferForm()
    
    context = {
        'form': form,
        'page_title': 'Stock Transfer'
    }
    
    return render(request, 'inventory/stock/stock_transfer.html', context)

@login_required
@inventory_permission_required('view')
def low_stock_view(request):
    """
    List all products that are low in stock (current_stock <= reorder_level and > 0).
    """
    low_stock_products = Product.objects.filter(
        is_active=True,
        current_stock__lte=F('reorder_level'),
        current_stock__gt=0
    ).select_related('category', 'supplier').order_by('current_stock')

    context = {
        "page_title": "Low Stock Products",
        "low_stock_products": low_stock_products,
    }
    return render(request, "inventory/stock/low_stock.html", context)

@login_required
@inventory_permission_required('edit')
def stock_take_create(request):
    """
    Schedule a new stock take.
    """
    if request.method == "POST":
        form = StockTakeForm(request.POST)
        if form.is_valid():
            stock_take = form.save(commit=False)
            stock_take.created_by = request.user
            stock_take.status = 'planned'
            stock_take.save()
            messages.success(request, "Stock take scheduled successfully.")
            return redirect('inventory:stock_take_detail', pk=stock_take.pk)
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = StockTakeForm()

    context = {
        "form": form,
        "stock_take": None,  # So your template shows the creation form
    }
    return render(request, "inventory/stock/stock_take.html", context)

@login_required
@inventory_permission_required('view')
def reorder_alert_list(request):
    """
    List all active and recent reorder alerts.
    """
    alerts = ReorderAlert.objects.select_related('product', 'suggested_supplier').order_by(
        '-priority', '-created_at'
    )
    context = {
        "page_title": "Reorder Alerts",
        "alerts": alerts,
    }
    return render(request, "inventory/alerts/reorder_alert_list.html", context)

@login_required
@inventory_permission_required('view')
def low_stock_report(request):
    """Generate low stock report"""
    products = Product.objects.filter(
        is_active=True,
        current_stock__lte=F('reorder_level')
    ).select_related('category', 'supplier').order_by('current_stock')
    
    context = {
        'page_title': 'Low Stock Report',
        'products': products,
        'total_products': products.count(),
    }
    
    return render(request, 'inventory/reports/low_stock_report.html', context)


# =====================================
# PURCHASE ORDER VIEWS
# =====================================

class PurchaseOrderListView(LoginRequiredMixin, ListView):
    """List all purchase orders"""
    model = PurchaseOrder
    template_name = 'inventory/purchase_orders/po_list.html'
    context_object_name = 'purchase_orders'
    paginate_by = 25
    
    @method_decorator(purchase_order_permission)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    
    def get_queryset(self):
        return PurchaseOrder.objects.select_related('supplier').order_by('-created_at')

class PurchaseOrderDetailView(LoginRequiredMixin, DetailView):
    """Detailed view of purchase order"""
    model = PurchaseOrder
    template_name = 'inventory/purchase_orders/po_detail.html'
    context_object_name = 'purchase_order'
    
    @method_decorator(purchase_order_permission)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        po = self.get_object()
        context.update({
            'page_title': f'Purchase Order: {po.po_number}',
            'po_items': po.items.select_related('product').all(),
            'total_value': sum(item.total_price for item in po.items.all()),
        })
        return context

class PurchaseOrderCreateView(LoginRequiredMixin, CreateView):
    """Create new purchase order"""
    model = PurchaseOrder
    form_class = PurchaseOrderForm
    template_name = 'inventory/purchase_orders/po_form.html'
    
    @method_decorator(purchase_order_permission)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    
    def form_valid(self, form):
        form.instance.created_by = self.request.user
        messages.success(self.request, f'Purchase order {form.instance.po_number} created successfully')
        return super().form_valid(form)

class PurchaseOrderUpdateView(LoginRequiredMixin, UpdateView):
    """Update purchase order"""
    model = PurchaseOrder
    form_class = PurchaseOrderForm
    template_name = 'inventory/purchase_orders/po_form.html'
    
    @method_decorator(purchase_order_permission)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

@login_required
@purchase_order_permission
def purchase_order_receive_view(request, pk):
    """Receive items from purchase order"""
    po = get_object_or_404(PurchaseOrder, pk=pk)
    
    if request.method == 'POST':
        # Process receiving logic here
        po.status = 'received'
        po.received_date = timezone.now().date()
        po.save()
        
        messages.success(request, f'Purchase order {po.po_number} marked as received')
        return redirect('inventory:po_detail', pk=po.pk)
    
    context = {
        'page_title': f'Receive PO: {po.po_number}',
        'purchase_order': po,
        'po_items': po.items.select_related('product').all(),
    }
    return render(request, 'inventory/purchase_orders/po_receive.html', context)

@login_required
@purchase_order_permission
def purchase_order_cancel_view(request, pk):
    """Cancel purchase order"""
    po = get_object_or_404(PurchaseOrder, pk=pk)
    
    if request.method == 'POST':
        po.status = 'cancelled'
        po.save()
        
        messages.success(request, f'Purchase order {po.po_number} cancelled')
        return redirect('inventory:po_list')
    
    context = {
        'page_title': f'Cancel PO: {po.po_number}',
        'purchase_order': po,
    }
    return render(request, 'inventory/purchase_orders/po_cancel.html', context)

# =====================================
# PRICING VIEWS
# =====================================

class CostCalculatorView(LoginRequiredMixin, TemplateView):
    """Calculate product costs with overhead factors"""
    template_name = 'inventory/pricing/cost_calculator.html'
    
    @method_decorator(cost_data_access)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get overhead factors for calculation
        overhead_factors = OverheadFactor.objects.filter(is_active=True)
        
        context.update({
            'page_title': 'Cost Calculator',
            'overhead_factors': overhead_factors,
        })
        return context

class BulkPriceUpdateView(LoginRequiredMixin, TemplateView):
    """Bulk update product prices"""
    template_name = 'inventory/pricing/bulk_price_update.html'
    
    @method_decorator(inventory_permission_required('edit'))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    
    def post(self, request, *args, **kwargs):
        # Process bulk price update
        price_adjustment = request.POST.get('price_adjustment', '0')
        adjustment_type = request.POST.get('adjustment_type', 'percentage')
        
        try:
            adjustment = Decimal(price_adjustment)
            updated_count = 0
            
            products = Product.objects.filter(is_active=True)
            for product in products:
                if adjustment_type == 'percentage':
                    product.selling_price *= (1 + adjustment / 100)
                else:
                    product.selling_price += adjustment
                product.save()
                updated_count += 1
            
            messages.success(request, f'Updated prices for {updated_count} products')
        except Exception as e:
            messages.error(request, f'Price update failed: {str(e)}')
        
        return redirect('inventory:product_list')

class MarginAnalysisView(LoginRequiredMixin, TemplateView):
    """Analyze profit margins across products"""
    template_name = 'inventory/pricing/margin_analysis.html'
    
    @method_decorator(cost_data_access)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Calculate margin statistics
        products = Product.objects.filter(is_active=True).select_related('category')
        
        margin_stats = []
        for product in products:
            if product.selling_price > 0:
                margin = product.selling_price - product.cost_price
                margin_percentage = (margin / product.selling_price) * 100
                margin_stats.append({
                    'product': product,
                    'margin': margin,
                    'margin_percentage': margin_percentage,
                })
        
        # Sort by margin percentage
        margin_stats.sort(key=lambda x: x['margin_percentage'], reverse=True)
        
        context.update({
            'page_title': 'Margin Analysis',
            'margin_stats': margin_stats,
        })
        return context

class CompetitivePricingView(LoginRequiredMixin, TemplateView):
    """Compare pricing with competitors"""
    template_name = 'inventory/pricing/competitive_pricing.html'
    
    @method_decorator(cost_data_access)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'page_title': 'Competitive Pricing Analysis',
        })
        return context

class MarkupRuleListView(LoginRequiredMixin, TemplateView):
    """List markup rules"""
    template_name = 'inventory/pricing/markup_rules.html'
    
    @method_decorator(inventory_permission_required('view'))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

class MarkupRuleCreateView(LoginRequiredMixin, TemplateView):
    """Create new markup rule"""
    template_name = 'inventory/pricing/markup_rule_form.html'
    
    @method_decorator(inventory_permission_required('edit'))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

class OverheadAnalysisView(LoginRequiredMixin, TemplateView):
    """Analyze overhead costs"""
    template_name = 'inventory/pricing/overhead_analysis.html'
    
    @method_decorator(cost_data_access)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        overhead_factors = OverheadFactor.objects.filter(is_active=True)
        
        context.update({
            'page_title': 'Overhead Cost Analysis',
            'overhead_factors': overhead_factors,
        })
        return context

class OverheadFactorListView(LoginRequiredMixin, ListView):
    """List and manage overhead factors used in cost calculations."""
    model = OverheadFactor
    template_name = 'inventory/configuration/overhead_factor_list.html'
    context_object_name = 'overhead_factors'

    @method_decorator(inventory_permission_required('view'))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def get_queryset(self):
        return OverheadFactor.objects.order_by('display_order', 'name')

class OverheadFactorCreateView(LoginRequiredMixin, CreateView):
    """Create a new overhead factor."""
    model = OverheadFactor
    form_class = OverheadFactorForm
    template_name = 'inventory/configuration/overhead_factor_form.html'
    success_url = reverse_lazy('inventory:overhead_factor_list')

    @method_decorator(inventory_permission_required('edit'))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def form_valid(self, form):
        messages.success(self.request, f'Overhead factor "{form.instance.name}" created.')
        return super().form_valid(form)

class OverheadFactorUpdateView(LoginRequiredMixin, UpdateView):
    """Edit an existing overhead factor."""
    model = OverheadFactor
    form_class = OverheadFactorForm
    template_name = 'inventory/configuration/overhead_factor_form.html'
    success_url = reverse_lazy('inventory:overhead_factor_list')

    @method_decorator(inventory_permission_required('edit'))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def form_valid(self, form):
        messages.success(self.request, f'Overhead factor "{form.instance.name}" updated.')
        return super().form_valid(form)

# =====================================
# LOCATION MANAGEMENT VIEWS
# =====================================

class LocationListView(BaseInventoryListView):
    """
    Manage warehouse locations and storage areas.
    Essential for multi-location inventory tracking.
    """
    model = Location
    template_name = 'inventory/configuration/location_list.html'
    context_object_name = 'locations'
    
class LocationDetailView(BaseInventoryDetailView):
    """
    Detailed view of a specific location with stock levels.
    """
    model = Location
    template_name = 'inventory/configuration/location_detail.html'
    context_object_name = 'location'
    
class LocationCreateView(BaseInventoryCreateView):
    """
    Create new warehouse locations.
    """
    model = Location
    form_class = LocationForm
    template_name = 'inventory/configuration/location_form.html'
    success_url = reverse_lazy('inventory:location_list')
    
class LocationUpdateView(BaseInventoryUpdateView):
    """
    Update existing warehouse locations.
    """
    model = Location
    form_class = LocationForm
    template_name = 'inventory/configuration/location_form.html'
    success_url = reverse_lazy('inventory:location_list')
    
# =====================================
# BRAND MANAGEMENT VIEWS
# =====================================

class BrandListView(BaseInventoryListView):
    model = Brand
    template_name = 'inventory/configuration/brand_list.html'
    context_object_name = 'brands'

class BrandCreateView(BaseInventoryCreateView):
    model = Brand
    fields = ['name', 'description', 'is_active']
    template_name = 'inventory/configuration/brand_form.html'
    success_url = reverse_lazy('inventory:brand_list')

class BrandUpdateView(BaseInventoryUpdateView):
    model = Brand
    fields = ['name', 'description', 'is_active']
    template_name = 'inventory/configuration/brand_form.html'
    success_url = reverse_lazy('inventory:brand_list')

class BrandDeleteView(BaseInventoryDeleteView):
    model = Brand
    template_name = 'inventory/configuration/brand_confirm_delete.html'
    success_url = reverse_lazy('inventory:brand_list')

class BrandDetailView(BaseInventoryDetailView):
    model = Brand
    template_name = 'inventory/configuration/brand_detail.html'
    context_object_name = 'brand'

class BrandProductsView(ProductListView):
    template_name = 'inventory/brand/brand_products.html'

    def get_queryset(self):
        qs = super().get_queryset()
        return qs.filter(brand_id=self.kwargs['pk'])

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        brand = get_object_or_404(Brand, pk=self.kwargs['pk'])
        ctx.update({
            'brand': brand,
            'page_title': f'Products by {brand.name}',
        })
        return ctx

# =====================================
# STORAGE LOCATION MANAGEMENT VIEWS
# =====================================

class StorageLocationListView(BaseInventoryListView):
    """
    Manage detailed storage locations within warehouses.
    """
    model = StorageLocation
    template_name = 'inventory/configuration/storage_location_list.html'
    context_object_name = 'storage_locations'
    
class StorageLocationCreateView(BaseInventoryCreateView):
    """
    Create new storage locations.
    """
    model = StorageLocation
    fields = [
        'name', 'code', 'location_type',
        'address', 'city', 'country',
        'contact_person', 'phone', 'email',
        'max_capacity_cubic_meters',
        'is_active', 'is_default',
        'allows_sales', 'allows_receiving',
    ]
    template_name = 'inventory/configuration/storage_location_form.html'
    success_url = reverse_lazy('inventory:storage_location_list')
    
class StorageLocationUpdateView(BaseInventoryUpdateView):
    """
    Edit existing storage locations.
    """
    model = StorageLocation
    fields = [
        'name', 'code', 'location_type',
        'address', 'city', 'country',
        'contact_person', 'phone', 'email',
        'max_capacity_cubic_meters',
        'is_active', 'is_default',
        'allows_sales', 'allows_receiving',
    ]
    template_name = 'inventory/configuration/storage_location_form.html'
    success_url = reverse_lazy('inventory:storage_location_list')

class StorageBinListView(LoginRequiredMixin, TemplateView):
    """
    Manage individual storage bins for precise inventory placement.
    """
    template_name = 'inventory/configuration/storage_bin_list.html'
    
    @method_decorator(inventory_permission_required('view'))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # For now, show storage locations as bins can be managed through them
        storage_locations = StorageLocation.objects.filter(is_active=True)
        
        context.update({
            'page_title': 'Storage Bin Management',
            'storage_locations': storage_locations,
            'note': 'Storage bins are managed through Storage Locations. Each storage location can contain multiple bins.'
        })
        return context

class StorageBinCreateView(LoginRequiredMixin, CreateView):
    """
    Create storage bins within a location.
    """
    model = StorageBin
    fields = [
        'location', 'bin_code', 'name',
        'component_families', 'row', 'column', 'shelf',
        'max_capacity_items', 'requires_special_handling',
        'is_active', 'notes'
    ]
    template_name = 'inventory/configuration/storage_bin_form.html'
    success_url = reverse_lazy('inventory:storage_location_list')

    @method_decorator(inventory_permission_required('edit'))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def get_initial(self):
        initial = super().get_initial()
        # Optional: allow ?location=<id> to preselect a location
        loc_id = self.request.GET.get('location')
        if loc_id:
            initial['location'] = loc_id
        return initial

    def form_valid(self, form):
        messages.success(
            self.request,
            f'Bin "{form.instance.bin_code}" created successfully'
        )
        return super().form_valid(form)

class StorageBinUpdateView(LoginRequiredMixin, UpdateView):
    """
    Edit existing storage bins.
    """
    model = StorageBin
    fields = [
        'location', 'bin_code', 'name',
        'component_families', 'row', 'column', 'shelf',
        'max_capacity_items', 'requires_special_handling',
        'is_active', 'notes'
    ]
    template_name = 'inventory/configuration/storage_bin_form.html'
    success_url = reverse_lazy('inventory:storage_location_list')

    @method_decorator(inventory_permission_required('edit'))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def form_valid(self, form):
        messages.success(
            self.request,
            f'Bin "{form.instance.bin_code}" updated successfully'
        )
        return super().form_valid(form)

# =====================================
# CONFIGURATION VIEWS
# =====================================

class CategoryListView(BaseInventoryListView):
    model = Category
    template_name = 'inventory/configuration/category_list.html'
    context_object_name = 'categories'

class CategoryCreateView(BaseInventoryCreateView):
    model = Category
    form_class = CategoryForm
    template_name = 'inventory/configuration/category_form.html'
    success_url = reverse_lazy('inventory:category_list')

class CategoryUpdateView(BaseInventoryUpdateView):
    model = Category
    form_class = CategoryForm
    template_name = 'inventory/configuration/category_form.html'
    success_url = reverse_lazy('inventory:category_list')

class CategoryDeleteView(BaseInventoryDeleteView):
    model = Category
    template_name = 'inventory/configuration/category_confirm_delete.html'
    success_url = reverse_lazy('inventory:category_list')

class CategoryDetailView(BaseInventoryDetailView):
    model = Category
    template_name = 'inventory/configuration/category_detail.html'
    context_object_name = 'category'

class CategoryProductsView(LoginRequiredMixin, ListView):
    """List products belonging to a category"""
    model = Product
    template_name = 'inventory/configuration/category_products.html'
    context_object_name = 'products'
    paginate_by = 30

    @method_decorator(inventory_permission_required('view'))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def get_queryset(self):
        # URL uses <int:pk> for the category
        return (
            Product.objects.filter(category_id=self.kwargs['pk'], is_active=True)
            .select_related('category', 'brand', 'supplier')
            .order_by('name')
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['category'] = get_object_or_404(Category, pk=self.kwargs['pk'])
        ctx['page_title'] = f"Products in {ctx['category'].name}"
        return ctx

class SupplierListView(BaseInventoryListView):
    """List all suppliers"""
    model = Supplier
    template_name = 'inventory/configuration/supplier_list.html'
    context_object_name = 'suppliers'

class SupplierCreateView(BaseInventoryCreateView):
    """Create new supplier"""
    model = Supplier
    form_class = SupplierForm
    template_name = 'inventory/configuration/supplier_form.html'
    success_url = reverse_lazy('inventory:supplier_list')

class SupplierUpdateView(BaseInventoryUpdateView):
    """Update supplier"""
    model = Supplier
    form_class = SupplierForm
    template_name = 'inventory/configuration/supplier_form.html'
    success_url = reverse_lazy('inventory:supplier_list')

class SupplierDeleteView(BaseInventoryDeleteView):
    """Delete supplier"""
    model = Supplier
    template_name = 'inventory/configuration/supplier_confirm_delete.html'
    success_url = reverse_lazy('inventory:supplier_list')

class SupplierDetailView(BaseInventoryDetailView):
    """Detailed view for a supplier, with quick stats."""
    model = Supplier
    template_name = 'inventory/suppliers/supplier_detail.html'
    context_object_name = 'supplier'
    
class SupplierProductsView(LoginRequiredMixin, ListView):
    """All products sourced from a given supplier."""
    model = Product
    template_name = 'inventory/suppliers/supplier_products.html'
    context_object_name = 'products'
    paginate_by = 30

    @method_decorator(inventory_permission_required('view'))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def get_queryset(self):
        return (Product.objects
                .filter(supplier_id=self.kwargs['pk'], is_active=True)
                .select_related('category', 'brand')
                .order_by('name'))

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        supplier = get_object_or_404(Supplier, pk=self.kwargs['pk'])
        ctx.update({
            'supplier': supplier,
            'page_title': f'Products by {supplier.name}',
        })
        return ctx

class SupplierPerformanceView(LoginRequiredMixin, TemplateView):
    """Show KPIs/metrics for a supplier."""
    template_name = 'inventory/suppliers/performance.html'

    @method_decorator(inventory_permission_required('view'))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        supplier = get_object_or_404(Supplier, pk=self.kwargs['pk'])
        recent = supplier.get_recent_performance(days=30)
        ctx.update({
            'supplier': supplier,
            'page_title': f'Performance: {supplier.name}',
            'average_lead_time_days': supplier.average_lead_time_days,
            'on_time_delivery_rate': supplier.on_time_delivery_rate,
            'quality_score': supplier.quality_score,
            'reliability_rating': supplier.reliability_rating,
            'rating': supplier.rating,
            'recent_performance': recent,
        })
        return ctx

# =====================================
# API ENDPOINTS
# =====================================

def calculate_overhead_for_product(self, import_cost, category_id, supplier_id, weight_grams):
    """Calculate overhead costs for a product"""
    overhead_factors = OverheadFactor.objects.filter(is_active=True)
    total_overhead = Decimal('0.00')
    
    for factor in overhead_factors:
        # Check if factor applies
        applies = True
        
        if factor.applies_to_categories.exists() and category_id:
            applies = applies and factor.applies_to_categories.filter(id=category_id).exists()
        
        if factor.applies_to_suppliers.exists() and supplier_id:
            applies = applies and factor.applies_to_suppliers.filter(id=supplier_id).exists()
        
        if applies:
            cost = factor.calculate_cost(
                product_cost=import_cost,
                order_value=import_cost,
                weight_kg=weight_grams / 1000 if weight_grams else None
            )
            total_overhead += cost
    
    return total_overhead

@login_required
@inventory_permission_required('view')
def product_search_api(request):
    """
    API endpoint for product search with autocomplete support.
    
    Supports integration with quote system and other modules.
    """
    try:
        search_term = request.GET.get('q', '').strip()
        limit = min(int(request.GET.get('limit', 20)), 100)
        
        if len(search_term) < 2:
            return JsonResponse({'success': False, 'error': 'Search term too short'})
        
        products = Product.objects.filter(
            Q(name__icontains=search_term) |
            Q(sku__icontains=search_term) |
            Q(barcode__icontains=search_term),
            is_active=True
        ).select_related('category', 'supplier')[:limit]
        
        results = []
        for product in products:
            results.append({
                'id': product.id,
                'sku': product.sku,
                'name': product.name,
                'category': product.category.name,
                'supplier': product.supplier.name,
                'cost_price': float(product.cost_price),
                'selling_price': float(product.selling_price),
                'current_stock': product.current_stock,
                'available_stock': product.available_stock,
                'stock_status': get_stock_status(product),
                'barcode': product.barcode
            })
        
        return JsonResponse({
            'success': True,
            'results': results,
            'count': len(results)
        })
        
    except Exception as e:
        logger.error(f"Error in product search API: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)})

@require_POST
@login_required
@inventory_permission_required('edit')
def stock_adjust_api(request):
    """
    API endpoint for stock adjustments.
    
    Provides programmatic stock adjustment capability for integrations.
    """
    try:
        data = json.loads(request.body)
        
        product_id = data.get('product_id')
        adjustment_type = data.get('adjustment_type')  # 'set', 'add', 'subtract'
        quantity = int(data.get('quantity', 0))
        reason = data.get('reason', 'API adjustment')
        notes = data.get('notes', '')
        
        # Validate inputs
        if not all([product_id, adjustment_type, quantity >= 0]):
            return JsonResponse({
                'success': False,
                'error': 'Missing or invalid parameters'
            })
        
        product = get_object_or_404(Product, id=product_id, is_active=True)
        
        # Calculate actual adjustment
        if adjustment_type == 'set':
            actual_adjustment = quantity - product.current_stock
        elif adjustment_type == 'add':
            actual_adjustment = quantity
        elif adjustment_type == 'subtract':
            actual_adjustment = -quantity
        else:
            return JsonResponse({
                'success': False,
                'error': 'Invalid adjustment type'
            })
        
        # Check for negative stock
        if product.current_stock + actual_adjustment < 0:
            return JsonResponse({
                'success': False,
                'error': 'Adjustment would result in negative stock'
            })
        
        # Create stock movement
        movement = create_stock_movement(
            product=product,
            movement_type='adjustment',
            quantity=actual_adjustment,
            reference=f"API adjustment: {reason}",
            user=request.user,
            notes=notes
        )
        
        return JsonResponse({
            'success': True,
            'product_id': product.id,
            'previous_stock': movement.previous_stock,
            'new_stock': movement.new_stock,
            'adjustment': actual_adjustment,
            'movement_id': movement.id
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'})
    except Exception as e:
        logger.error(f"Error in stock adjust API: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
@inventory_permission_required('view')
def check_sku_availability(request):
    """Check if SKU is available for use"""
    sku = request.GET.get('sku', '').strip()
    product_id = request.GET.get('product_id')  # For updates
    
    if not sku:
        return JsonResponse({'available': False, 'error': 'SKU required'})
    
    existing = Product.objects.filter(sku=sku)
    if product_id:
        existing = existing.exclude(id=product_id)
    
    available = not existing.exists()
    
    return JsonResponse({
        'available': available,
        'message': 'SKU available' if available else 'SKU already in use'
    })

@login_required
@inventory_permission_required('view')
def check_barcode_availability(request):
    """Check if barcode is available for use"""
    barcode = request.GET.get('barcode', '').strip()
    product_id = request.GET.get('product_id')  # For updates
    
    if not barcode:
        return JsonResponse({'available': True})  # Empty barcode is allowed
    
    existing = Product.objects.filter(barcode=barcode)
    if product_id:
        existing = existing.exclude(id=product_id)
    
    available = not existing.exists()
    
    return JsonResponse({
        'available': available,
        'message': 'Barcode available' if available else 'Barcode already in use'
    })

@login_required
@inventory_permission_required('view')
def dashboard_metrics_api(request):
    """
    Return dashboard chart data and summary metrics as JSON for AJAX.
    """
    period = int(request.GET.get("period", 30))  # days, default: 30
    end_date = datetime.now()
    start_date = end_date - timedelta(days=period)

    # Prepare time labels
    labels = [(start_date + timedelta(days=i)).strftime('%Y-%m-%d') for i in range(period)]

    # Stock movement trends (simple version)
    movement_data = StockMovement.objects.filter(created_at__range=[start_date, end_date])
    stock_in = []
    stock_out = []
    label_map = {label: i for i, label in enumerate(labels)}
    stock_in_temp = [0 for _ in labels]
    stock_out_temp = [0 for _ in labels]

    for m in movement_data:
        label = m.created_at.strftime('%Y-%m-%d')
        idx = label_map.get(label)
        if idx is not None:
            if m.quantity > 0:
                stock_in_temp[idx] += m.quantity
            else:
                stock_out_temp[idx] += abs(m.quantity)

    # Category distribution (stock value by category)
    category_qs = Category.objects.all()
    category_labels = []
    category_values = []
    for cat in category_qs:
        value = Product.objects.filter(category=cat, is_active=True).aggregate(
            total=Sum(F('current_stock') * F('cost_price'))
        )['total'] or 0
        if value > 0:
            category_labels.append(cat.name)
            category_values.append(float(value))

    data = {
        "success": True,
        "movement_chart": {
            "labels": labels,
            "stock_in": stock_in_temp,
            "stock_out": stock_out_temp,
        },
        "category_chart": {
            "labels": category_labels,
            "values": category_values,
        },
        # Add more dashboard stats if you want!
    }
    return JsonResponse(data)

@login_required
@inventory_permission_required('view')
def recent_activities_api(request):
    """
    Return the most recent inventory activities as JSON.
    """
    activities = []

    # Stock movements (last 10)
    for m in StockMovement.objects.select_related('product', 'created_by').order_by('-created_at')[:10]:
        activities.append({
            "type": "Stock Movement",
            "date": m.created_at.strftime("%Y-%m-%d %H:%M"),
            "user": m.created_by.get_full_name() if m.created_by else "",
            "details": f"{m.get_movement_type_display()} of {m.quantity} - {m.product.name}",
        })
    
    # Product updates (last 5)
    for p in Product.objects.select_related('category').order_by('-updated_at')[:5]:
        activities.append({
            "type": "Product Update",
            "date": p.updated_at.strftime("%Y-%m-%d %H:%M") if p.updated_at else "",
            "user": p.modified_by.get_full_name() if hasattr(p, 'modified_by') and p.modified_by else "",
            "details": f"Product '{p.name}' updated.",
        })

    # Purchase Orders (last 5)
    for po in PurchaseOrder.objects.select_related('supplier').order_by('-created_at')[:5]:
        activities.append({
            "type": "Purchase Order",
            "date": po.created_at.strftime("%Y-%m-%d %H:%M"),
            "user": po.created_by.get_full_name() if po.created_by else "",
            "details": f"PO {po.po_number} for {po.supplier.name}",
        })

    # Sort all by date (desc)
    activities.sort(key=lambda x: x['date'], reverse=True)

    # Limit to latest 20
    data = {
        "success": True,
        "activities": activities[:20],
    }
    return JsonResponse(data)

@login_required
@inventory_permission_required('view')
def critical_alerts_api(request):
    """
    Return all critical (and out-of-stock) reorder alerts as JSON.
    """
    alerts = ReorderAlert.objects.filter(
        priority='critical',  # or whatever value you use for "critical" priority
        status__in=['active', 'acknowledged']
    ).select_related('product')

    data = {
        "success": True,
        "alerts": [
            {
                "id": alert.id,
                "product_id": alert.product.id,
                "product_name": alert.product.name,
                "current_stock": alert.current_stock,
                "reorder_level": alert.reorder_level,
                "suggested_order_quantity": alert.suggested_order_quantity,
                "estimated_cost": float(alert.estimated_cost) if alert.estimated_cost else None,
                "priority": alert.priority,
            }
            for alert in alerts
        ]
    }
    return JsonResponse(data)

@login_required
@inventory_permission_required('view')
def component_family_attributes_api(request, family_id):
    """Get dynamic attributes for a component family"""
    try:
        family = get_object_or_404(ComponentFamily, id=family_id, is_active=True)
        attributes = family.all_attributes
        
        attribute_data = []
        for attr in attributes:
            attribute_data.append({
                'id': attr.id,
                'name': attr.name,
                'field_type': attr.field_type,
                'is_required': attr.is_required,
                'default_value': attr.default_value,
                'help_text': attr.help_text,
                'choice_options': attr.choice_options,
                'min_value': float(attr.min_value) if attr.min_value else None,
                'max_value': float(attr.max_value) if attr.max_value else None,
                'validation_pattern': attr.validation_pattern,
            })
        
        return JsonResponse({
            'success': True,
            'family': family.name,
            'attributes': attribute_data
        })
        
    except ComponentFamily.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Component family not found'})

@login_required
@inventory_permission_required('view')
def calculate_product_cost_api(request, product_id):
    """
    API endpoint: Calculate real-time product cost with current overhead factors.
    Used for quote system and cost analysis.
    """
    try:
        product = get_object_or_404(Product, id=product_id)
        quantity = int(request.GET.get('quantity', 1))
        
        # Base cost calculation
        base_cost = product.cost_price * quantity
        
        # Apply overhead factors
        overhead_factors = OverheadFactor.objects.filter(is_active=True)
        overhead_breakdown = []
        total_overhead_amount = Decimal('0')
        
        for factor in overhead_factors:
            overhead_amount = base_cost * (factor.percentage / 100)
            total_overhead_amount += overhead_amount
            overhead_breakdown.append({
                'name': factor.name,
                'percentage': float(factor.percentage),
                'amount': float(overhead_amount),
            })
        
        total_cost = base_cost + total_overhead_amount
        
        # Supplier information
        supplier_info = None
        if product.supplier:
            supplier_info = {
                'name': product.supplier.name,
                'currency': product.supplier.currency,
                'lead_time_days': product.supplier.average_lead_time_days,
                'minimum_order': float(product.supplier.minimum_order_amount),
            }
        
        data = {
            'product_id': product.id,
            'product_name': product.name,
            'sku': product.sku,
            'quantity': quantity,
            'unit_cost': float(product.cost_price),
            'base_cost': float(base_cost),
            'overhead_breakdown': overhead_breakdown,
            'total_overhead': float(total_overhead_amount),
            'total_cost': float(total_cost),
            'unit_total_cost': float(total_cost / quantity) if quantity > 0 else 0,
            'supplier_info': supplier_info,
            'calculated_at': timezone.now().isoformat(),
        }
        
        return JsonResponse({'success': True, 'cost_calculation': data})
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
@inventory_permission_required('view')
def product_stock_levels_api(request, product_id):
    """
    API endpoint: Get current stock levels for a product across all locations.
    """
    try:
        product = get_object_or_404(Product, id=product_id)
        
        stock_levels = []
        total_available = 0
        total_reserved = 0
        
        for stock_level in product.stock_levels.select_related('location'):
            stock_levels.append({
                'location_id': stock_level.location.id,
                'location_name': stock_level.location.name,
                'location_code': stock_level.location.location_code,
                'quantity': stock_level.quantity,
                'available_quantity': stock_level.available_quantity,
                'reserved_quantity': stock_level.reserved_quantity,
                'last_counted': stock_level.last_counted.isoformat() if stock_level.last_counted else None,
                'last_movement': stock_level.last_movement.isoformat() if stock_level.last_movement else None,
            })
            total_available += stock_level.available_quantity
            total_reserved += stock_level.reserved_quantity
        
        data = {
            'product_id': product.id,
            'product_name': product.name,
            'sku': product.sku,
            'total_stock': product.current_stock,
            'total_available': total_available,
            'total_reserved': total_reserved,
            'reorder_level': product.reorder_level,
            'stock_status': get_stock_status(product),
            'locations': stock_levels,
            'checked_at': timezone.now().isoformat(),
        }
        
        return JsonResponse({'success': True, 'stock_levels': data})
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
@inventory_permission_required('edit')
@require_POST
@csrf_exempt
def reserve_stock_api(request):
    """
    API endpoint: Reserve stock for quotes/orders.
    """
    try:
        data = json.loads(request.body)
        product_id = data['product_id']
        location_id = data.get('location_id')
        quantity = int(data['quantity'])
        reference = data.get('reference', f'RESERVE-{timezone.now().strftime("%Y%m%d%H%M%S")}')
        
        product = get_object_or_404(Product, id=product_id)
        
        if location_id:
            location = get_object_or_404(Location, id=location_id)
            stock_level = get_object_or_404(StockLevel, product=product, location=location)
            
            if stock_level.available_quantity < quantity:
                return JsonResponse({
                    'success': False, 
                    'error': f'Insufficient stock. Available: {stock_level.available_quantity}, Requested: {quantity}'
                }, status=400)
            
            stock_level.reserved_quantity += quantity
            stock_level.save()
            
        else:
            # Reserve from total stock
            available_stock = calculate_available_stock(product)
            if available_stock < quantity:
                return JsonResponse({
                    'success': False,
                    'error': f'Insufficient stock. Available: {available_stock}, Requested: {quantity}'
                }, status=400)
            
            # Reserve from locations with stock (FIFO)
            remaining_to_reserve = quantity
            for stock_level in product.stock_levels.filter(quantity__gt=0):
                if remaining_to_reserve <= 0:
                    break
                
                available = stock_level.available_quantity
                if available > 0:
                    reserve_qty = min(available, remaining_to_reserve)
                    stock_level.reserved_quantity += reserve_qty
                    stock_level.save()
                    remaining_to_reserve -= reserve_qty
        
        return JsonResponse({
            'success': True,
            'message': f'Reserved {quantity} units of {product.name}',
            'reference': reference,
            'reserved_quantity': quantity,
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
@inventory_permission_required('edit')
@require_POST
@csrf_exempt
def release_reservation_api(request):
    """
    API endpoint: Release reserved stock.
    """
    try:
        data = json.loads(request.body)
        product_id = data['product_id']
        location_id = data.get('location_id')
        quantity = int(data['quantity'])
        
        product = get_object_or_404(Product, id=product_id)
        
        if location_id:
            location = get_object_or_404(Location, id=location_id)
            stock_level = get_object_or_404(StockLevel, product=product, location=location)
            
            release_qty = min(stock_level.reserved_quantity, quantity)
            stock_level.reserved_quantity -= release_qty
            stock_level.save()
            
        else:
            # Release from all locations
            remaining_to_release = quantity
            for stock_level in product.stock_levels.filter(reserved_quantity__gt=0):
                if remaining_to_release <= 0:
                    break
                
                release_qty = min(stock_level.reserved_quantity, remaining_to_release)
                stock_level.reserved_quantity -= release_qty
                stock_level.save()
                remaining_to_release -= release_qty
        
        return JsonResponse({
            'success': True,
            'message': f'Released {quantity} reserved units of {product.name}',
            'released_quantity': quantity,
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
@inventory_permission_required('view')
def product_availability_api(request):
    """
    API endpoint: Check product availability for multiple products.
    Used for quote system to validate stock availability.
    """
    try:
        if request.method == 'POST':
            data = json.loads(request.body)
            product_requests = data.get('products', [])
        else:
            # GET request with URL parameters
            product_requests = []
            product_ids = request.GET.getlist('product_id')
            quantities = request.GET.getlist('quantity')
            
            for i, product_id in enumerate(product_ids):
                qty = int(quantities[i]) if i < len(quantities) else 1
                product_requests.append({'product_id': product_id, 'quantity': qty})
        
        availability_results = []
        
        for req in product_requests:
            try:
                product = Product.objects.get(id=req['product_id'], is_active=True)
                requested_qty = int(req['quantity'])
                available_qty = calculate_available_stock(product)
                
                is_available = available_qty >= requested_qty
                shortage = max(0, requested_qty - available_qty)
                
                # Estimate restock date if out of stock
                restock_estimate = None
                if shortage > 0 and product.supplier:
                    lead_time = product.supplier.average_lead_time_days or 7
                    restock_estimate = (timezone.now() + timedelta(days=lead_time)).date().isoformat()
                
                availability_results.append({
                    'product_id': product.id,
                    'product_name': product.name,
                    'sku': product.sku,
                    'requested_quantity': requested_qty,
                    'available_quantity': available_qty,
                    'is_available': is_available,
                    'shortage': shortage,
                    'restock_estimate': restock_estimate,
                    'current_stock': product.current_stock,
                    'reorder_level': product.reorder_level,
                })
                
            except Product.DoesNotExist:
                availability_results.append({
                    'product_id': req['product_id'],
                    'error': 'Product not found or inactive',
                    'is_available': False,
                })
        
        return JsonResponse({
            'success': True,
            'availability': availability_results,
            'checked_at': timezone.now().isoformat(),
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
@inventory_permission_required('edit')
@require_POST
def generate_reorder_list_api(request):
    """Generate and return reorder list for download"""
    try:
        # Get filter parameters
        data = json.loads(request.body) if request.body else {}
        supplier_id = data.get('supplier_id')
        category_id = data.get('category_id')
        priority = data.get('priority')  # 'critical', 'high', 'medium'
        
        # Base query
        products = Product.objects.filter(
            is_active=True,
            total_stock__lte=F('reorder_level')
        ).select_related('supplier', 'category', 'brand', 'supplier_currency')
        
        # Apply filters
        if supplier_id:
            products = products.filter(supplier_id=supplier_id)
        if category_id:
            products = products.filter(category_id=category_id)
        
        # Generate reorder list
        reorder_data = []
        total_order_value = Decimal('0.00')
        
        for product in products:
            # Determine priority
            if product.total_stock == 0:
                item_priority = 'critical'
            elif product.total_stock <= (product.reorder_level * 0.5):
                item_priority = 'high'
            else:
                item_priority = 'medium'
            
            # Filter by priority if specified
            if priority and item_priority != priority:
                continue
            
            # Calculate order details
            recommended_qty = max(
                product.reorder_quantity,
                product.supplier_minimum_order_quantity,
                product.reorder_level - product.total_stock + 10
            )
            
            unit_cost = product.get_preferred_supplier_price(recommended_qty)
            order_value = unit_cost * recommended_qty
            total_order_value += order_value
            
            reorder_data.append({
                'supplier': product.supplier.name,
                'supplier_email': product.supplier.email,
                'sku': product.sku,
                'supplier_sku': product.supplier_sku,
                'name': product.name,
                'current_stock': product.total_stock,
                'reorder_level': product.reorder_level,
                'recommended_quantity': recommended_qty,
                'unit_cost': float(unit_cost),
                'currency': product.supplier_currency.code,
                'order_value': float(order_value),
                'lead_time_days': product.supplier_lead_time_days,
                'moq': product.supplier_minimum_order_quantity,
                'priority': item_priority
            })
        
        return JsonResponse({
            'success': True,
            'reorder_list': reorder_data,
            'summary': {
                'total_items': len(reorder_data),
                'total_order_value': float(total_order_value),
                'suppliers': len(set(item['supplier'] for item in reorder_data))
            }
        })
        
    except Exception as e:
        logger.error(f"Error generating reorder list: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)})

@csrf_exempt
@require_http_methods(["GET"])
def product_attributes_api(request, product_id):
    """API for product attributes"""
    try:
        product = Product.objects.get(id=product_id, is_active=True)
        data = {
            'product_id': product.id,
            'attributes': product.get_dynamic_attributes(),  # Assuming this method exists
        }
        return JsonResponse(data)
    except Product.DoesNotExist:
        return JsonResponse({'error': 'Product not found'}, status=404)

@csrf_exempt
@require_http_methods(["POST"])
def stock_adjustment_api(request):
    """API for stock adjustments"""
    try:
        data = json.loads(request.body)
        product_id = data.get('product_id')
        adjustment = data.get('adjustment')
        notes = data.get('notes', '')
        
        product = Product.objects.get(id=product_id, is_active=True)
        product.current_stock += adjustment
        product.save()
        
        # Create movement record
        create_stock_movement(
            product=product,
            movement_type='adjustment',
            quantity=adjustment,
            user=request.user,
            notes=notes
        )
        
        return JsonResponse({
            'success': True,
            'new_stock': product.current_stock
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

@csrf_exempt
@require_http_methods(["POST"])
def stock_transfer_api(request):
    """API for stock transfers"""
    try:
        data = json.loads(request.body)
        product_id = data.get('product_id')
        from_location_id = data.get('from_location_id')
        to_location_id = data.get('to_location_id')
        quantity = data.get('quantity')
        
        product = Product.objects.get(id=product_id, is_active=True)
        from_location = Location.objects.get(id=from_location_id)
        to_location = Location.objects.get(id=to_location_id)
        
        # Process transfer
        from_stock = StockLevel.objects.get(product=product, location=from_location)
        to_stock, created = StockLevel.objects.get_or_create(
            product=product, 
            location=to_location,
            defaults={'quantity': 0}
        )
        
        from_stock.quantity -= quantity
        to_stock.quantity += quantity
        
        from_stock.save()
        to_stock.save()
        
        # Create movement records
        create_stock_movement(
            product=product,
            movement_type='transfer_out',
            quantity=-quantity,
            user=request.user,
            from_location=from_location,
            to_location=to_location
        )
        
        create_stock_movement(
            product=product,
            movement_type='transfer_in',
            quantity=quantity,
            user=request.user,
            from_location=from_location,
            to_location=to_location
        )
        
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

@csrf_exempt
@require_http_methods(["GET"])
def reorder_suggestions_api(request):
    """API for reorder suggestions"""
    low_stock_products = Product.objects.filter(
        is_active=True,
        current_stock__lte=F('reorder_level')
    ).select_related('supplier')
    
    suggestions = []
    for product in low_stock_products:
        suggestions.append({
            'product_id': product.id,
            'product_name': product.name,
            'current_stock': product.current_stock,
            'reorder_level': product.reorder_level,
            'reorder_quantity': product.reorder_quantity,
            'supplier': product.supplier.name if product.supplier else '',
            'cost_price': str(product.cost_price),
            'total_cost': str(product.reorder_quantity * product.cost_price),
        })
    
    return JsonResponse({'suggestions': suggestions})

# =====================================
# INTEGRATION ENDPOINTS FOR QUOTE SYSTEM
# =====================================

@login_required
@inventory_permission_required('view')
def quote_products_api(request):
    """
    API endpoint for quote system to search and retrieve products.
    
    Provides product information optimized for quote creation.
    """
    try:
        from .utils import get_products_for_quote_system
        
        search_term = request.GET.get('search')
        category_id = request.GET.get('category')
        supplier_id = request.GET.get('supplier')
        limit = min(int(request.GET.get('limit', 50)), 100)
        
        # Get category and supplier objects if provided
        category = None
        supplier = None
        
        if category_id:
            try:
                category = Category.objects.get(id=category_id)
            except Category.DoesNotExist:
                pass
        
        if supplier_id:
            try:
                supplier = Supplier.objects.get(id=supplier_id)
            except Supplier.DoesNotExist:
                pass
        
        products = get_products_for_quote_system(
            search_term=search_term,
            category=category,
            supplier=supplier,
            limit=limit
        )
        
        return JsonResponse({
            'success': True,
            'products': products,
            'count': len(products)
        })
        
    except Exception as e:
        logger.error(f"Error in quote products API: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)})

@require_POST
@login_required
@inventory_permission_required('view')
def check_quote_availability_api(request):
    """
    Check stock availability for quote items.
    
    Used by quote system to validate stock before quote creation.
    """
    try:
        data = json.loads(request.body)
        quote_items = data.get('items', [])
        
        if not quote_items:
            return JsonResponse({'success': False, 'error': 'No items provided'})
        
        from .utils import check_stock_availability_for_quote
        
        availability = check_stock_availability_for_quote(quote_items)
        
        # Check if all items can be fulfilled
        all_available = all(
            item.get('can_fulfill', False) 
            for item in availability.values() 
            if 'error' not in item
        )
        
        return JsonResponse({
            'success': True,
            'all_available': all_available,
            'availability': availability
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'})
    except Exception as e:
        logger.error(f"Error checking quote availability: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)})

@require_POST
@login_required
@inventory_permission_required('edit')
def reserve_for_quote_api(request):
    """
    Reserve stock for quote items.
    
    Creates stock reservations that can be released if quote is not accepted.
    """
    try:
        data = json.loads(request.body)
        quote_items = data.get('items', [])
        quote_reference = data.get('quote_reference', '')
        
        if not quote_items or not quote_reference:
            return JsonResponse({
                'success': False,
                'error': 'Items and quote reference required'
            })
        
        from .utils import reserve_stock_for_quote
        
        results = reserve_stock_for_quote(
            quote_items=quote_items,
            quote_reference=quote_reference,
            user=request.user
        )
        
        return JsonResponse(results)
        
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'})
    except Exception as e:
        logger.error(f"Error reserving stock for quote: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)})

# =====================================
# BUSINESS INTELLIGENCE VIEWS
# =====================================

@login_required
@inventory_permission_required('view')
def inventory_analytics_view(request):
    """Comprehensive inventory analytics dashboard"""
    context = {}
    
    # Date range for analytics
    end_date = timezone.now()
    start_date = end_date - timedelta(days=30)
    
    # Stock value trends
    categories = Category.objects.filter(is_active=True).annotate(
        total_value=Sum(F('products__total_stock') * F('products__total_cost_price_usd')),
        product_count=Count('products', filter=Q(products__is_active=True))
    ).order_by('-total_value')
    
    context['category_analysis'] = categories[:10]
    
    # Supplier analysis
    suppliers = Supplier.objects.filter(is_active=True).annotate(
        inventory_value=Sum(F('products__total_stock') * F('products__total_cost_price_usd')),
        product_count=Count('products', filter=Q(products__is_active=True)),
        avg_markup=Avg('products__markup_percentage')
    ).order_by('-inventory_value')
    
    context['supplier_analysis'] = suppliers[:10]
    
    # Brand performance
    brands = Brand.objects.filter(is_active=True).annotate(
        inventory_value=Sum(F('products__total_stock') * F('products__total_cost_price_usd')),
        product_count=Count('products', filter=Q(products__is_active=True)),
        avg_markup=Avg('products__markup_percentage')
    ).order_by('-inventory_value')
    
    context['brand_analysis'] = brands[:10]
    
    # Stock status distribution
    total_products = Product.objects.filter(is_active=True).count()
    stock_distribution = {
        'in_stock': Product.objects.filter(is_active=True, total_stock__gt=F('reorder_level')).count(),
        'low_stock': Product.objects.filter(is_active=True, total_stock__lte=F('reorder_level'), total_stock__gt=0).count(),
        'out_of_stock': Product.objects.filter(is_active=True, total_stock=0).count(),
    }
    
    context['stock_distribution'] = stock_distribution
    context['total_products'] = total_products
    
    # Markup analysis
    markup_ranges = [
        ('0-20%', 0, 20),
        ('20-30%', 20, 30),
        ('30-40%', 30, 40),
        ('40-50%', 40, 50),
        ('50%+', 50, 1000),
    ]
    
    markup_analysis = []
    for label, min_markup, max_markup in markup_ranges:
        count = Product.objects.filter(
            is_active=True,
            markup_percentage__gte=min_markup,
            markup_percentage__lt=max_markup
        ).count()
        markup_analysis.append({'range': label, 'count': count})
    
    context['markup_analysis'] = markup_analysis
    
    return render(request, 'inventory/analytics.html', context)

# =====================================
# REPORTS VIEWS
# =====================================

class AlertsDashboardView(LoginRequiredMixin, TemplateView):
    template_name = "inventory/alerts/reorder_alert_list.html"

    @method_decorator(inventory_permission_required('view'))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        active_alerts = ReorderAlert.objects.filter(
            status__in=['active', 'acknowledged']
        ).select_related('product', 'suggested_supplier').order_by('priority', '-created_at')
        context['active_alerts'] = active_alerts
        return context

class ReorderAlertListView(LoginRequiredMixin, ListView):
    """
    Manage all reorder alerts - active, acknowledged, and resolved.
    """
    model = ReorderAlert
    template_name = 'inventory/reorder/reorder_alerts.html'
    context_object_name = 'alerts'
    paginate_by = 50
    
    @method_decorator(inventory_permission_required('view'))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    
    def get_queryset(self):
        queryset = ReorderAlert.objects.select_related(
            'product', 'suggested_supplier', 'acknowledged_by', 'purchase_order'
        ).order_by('-priority', '-created_at')
        
        # Apply filters
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)
        
        priority = self.request.GET.get('priority')
        if priority:
            queryset = queryset.filter(priority=priority)
        
        supplier = self.request.GET.get('supplier')
        if supplier:
            queryset = queryset.filter(suggested_supplier_id=supplier)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Calculate summary statistics
        alerts = ReorderAlert.objects.all()
        summary = {
            'total_alerts': alerts.count(),
            'active_alerts': alerts.filter(status='active').count(),
            'acknowledged_alerts': alerts.filter(status='acknowledged').count(),
            'critical_alerts': alerts.filter(priority='critical').count(),
            'high_priority_alerts': alerts.filter(priority='high').count(),
        }
        
        context.update({
            'page_title': 'Reorder Alerts',
            'summary': summary,
            'suppliers': Supplier.objects.filter(is_active=True).order_by('name'),
            'filters': {
                'status': self.request.GET.get('status'),
                'priority': self.request.GET.get('priority'),
                'supplier': self.request.GET.get('supplier'),
            }
        })
        return context

class CustomReportView(LoginRequiredMixin, TemplateView):
    """Custom report builder"""
    template_name = 'inventory/reports/custom_report.html'
    
    @method_decorator(inventory_permission_required('view'))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

@login_required
@inventory_permission_required('view')
def inventory_reports_view(request):
    """Inventory reports dashboard"""
    return render(request, 'inventory/reports.html', {
        'report_types': [
            {'name': 'Stock Valuation Report', 'url': 'stock_valuation'},
            {'name': 'Low Stock Report', 'url': 'low_stock'},
            {'name': 'Supplier Performance', 'url': 'supplier_performance'},
            {'name': 'Category Analysis', 'url': 'category_analysis'},
            {'name': 'Markup Analysis', 'url': 'markup_analysis'},
            {'name': 'ABC Analysis', 'url': 'abc_analysis'},
        ]
    })

@login_required
@inventory_permission_required('view')
def stock_valuation_report(request):
    """Generate stock valuation report"""
    # Implementation for stock valuation report
    products = Product.objects.filter(is_active=True).select_related(
        'category', 'supplier', 'brand'
    ).order_by('category__name', 'name')
    
    if request.GET.get('format') == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="stock_valuation_report.csv"'
        
        writer = csv.writer(response)
        writer.writerow([
            'Category', 'SKU', 'Product Name', 'Brand', 'Supplier',
            'Stock Quantity', 'Unit Cost (USD)', 'Total Value (USD)',
            'Unit Selling Price', 'Potential Revenue', 'Potential Profit'
        ])
        
        for product in products:
            writer.writerow([
                product.category.name,
                product.sku,
                product.name,
                product.brand.name,
                product.supplier.name,
                product.total_stock,
                product.total_cost_price_usd,
                product.stock_value_usd,
                product.selling_price,
                product.total_stock * product.selling_price,
                product.total_stock * product.profit_per_unit_usd
            ])
        
        return response
    
    # Return HTML version
    context = {'products': products}
    return render(request, 'inventory/reports/stock_valuation.html', context)

def supplier_details_api(request, pk):
    """
    API endpoint for AJAX: Returns supplier details as JSON.
    """
    try:
        supplier = Supplier.objects.get(pk=pk)
        data = {
            'id': supplier.id,
            'name': supplier.name,
            'supplier_code': supplier.supplier_code,
            'supplier_type': supplier.get_supplier_type_display(),
            'email': supplier.email,
            'phone': supplier.phone,
            'address': f"{supplier.address_line_1} {supplier.address_line_2} {supplier.city} {supplier.country}".strip(),
            'currency': supplier.currency,
            'minimum_order_amount': float(supplier.minimum_order_amount),
            'average_lead_time_days': supplier.average_lead_time_days,
            'reliability_rating': float(supplier.reliability_rating),
            'is_active': supplier.is_active,
            'is_preferred': supplier.is_preferred,
            # Add more fields if needed!
        }
        return JsonResponse({'success': True, 'supplier': data})
    except Supplier.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Supplier not found'}, status=404)

@login_required
@inventory_permission_required('view')
def quick_search_view(request):
    """Quick search across all inventory items"""
    search_term = request.GET.get('q', '').strip()
    
    if len(search_term) < 2:
        return JsonResponse({'results': []})
    
    # Search products
    products = Product.objects.filter(
        Q(name__icontains=search_term) |
        Q(sku__icontains=search_term),
        is_active=True
    )[:10]
    
    results = [
        {
            'type': 'product',
            'id': p.id,
            'title': p.name,
            'subtitle': p.sku,
            'url': reverse('inventory:product_detail', args=[p.id])
        }
        for p in products
    ]
    
    return JsonResponse({'results': results})

@login_required
@inventory_permission_required('view')
def stock_movements_export(request):
    """
    Export stock movements to CSV.
    """
    # Optional: Add filter params, e.g. by date, type, etc.
    movements = StockMovement.objects.select_related('product', 'from_location', 'to_location', 'created_by').order_by('-created_at')

    # Prepare CSV response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="stock_movements.csv"'

    writer = csv.writer(response)
    # Header
    writer.writerow([
        'Date', 'Product SKU', 'Product Name', 'Type', 'Quantity',
        'From Location', 'To Location', 'Reference', 'Previous Stock',
        'New Stock', 'Created By', 'Notes'
    ])

    # Rows
    for m in movements:
        writer.writerow([
            m.created_at.strftime('%Y-%m-%d %H:%M'),
            m.product.sku,
            m.product.name,
            m.get_movement_type_display(),
            m.quantity,
            m.from_location.name if m.from_location else '',
            m.to_location.name if m.to_location else '',
            m.reference,
            m.previous_stock,
            m.new_stock,
            m.created_by.get_full_name() if m.created_by else '',
            m.notes or '',
        ])
    return response

@login_required
@inventory_permission_required('view')
def reports_dashboard(request):
    """
    Enterprise-grade inventory reports dashboard view.
    Aggregates inventory and operational KPIs, and supports extensible reporting.
    """
    # Key operational stats
    total_products = Product.objects.filter(is_active=True).count()
    total_stock_value = Product.objects.filter(is_active=True).aggregate(
        total=Sum(F('current_stock') * F('cost_price'))
    )['total'] or 0

    low_stock_count = Product.objects.filter(
        is_active=True,
        current_stock__lte=F('reorder_level'),
        current_stock__gt=0
    ).count()

    out_of_stock_count = Product.objects.filter(
        is_active=True,
        current_stock=0
    ).count()

    # Per-category and per-supplier breakdowns
    top_categories = Category.objects.annotate(
        product_count=Count('products', filter=Q(products__is_active=True)),
        stock_value=Sum(F('products__current_stock') * F('products__cost_price'))
    ).order_by('-stock_value')[:6]

    top_suppliers = Supplier.objects.annotate(
        product_count=Count('products', filter=Q(products__is_active=True)),
        stock_value=Sum(F('products__current_stock') * F('products__cost_price'))
    ).order_by('-stock_value')[:6]

    # Recent stock movements
    recent_movements = StockMovement.objects.select_related('product', 'created_by').order_by('-created_at')[:10]

    # Alerts and overdue POs
    active_alerts = ReorderAlert.objects.filter(
        status__in=['active', 'acknowledged']
    ).select_related('product', 'suggested_supplier').order_by('-priority', '-created_at')[:10]

    overdue_pos = PurchaseOrder.objects.filter(
        status__in=['sent', 'acknowledged'],
        expected_delivery_date__lt=timezone.now().date()
    ).select_related('supplier').order_by('expected_delivery_date')[:5]

    # Stock take analytics
    pending_stock_takes = StockTake.objects.filter(
        status__in=['planned', 'in_progress']
    ).order_by('scheduled_date')[:3]

    # Historical analytics - can be expanded with real BI/analytics backend
    movements_last_30_days = StockMovement.objects.filter(
        created_at__gte=timezone.now() - timezone.timedelta(days=30)
    ).count()
    total_value_last_30_days = StockMovement.objects.filter(
        created_at__gte=timezone.now() - timezone.timedelta(days=30)
    ).aggregate(
        value=Sum(F('quantity') * F('product__cost_price'))
    )['value'] or 0

    context = {
        # KPIs for dashboard cards
        'total_products': total_products,
        'total_stock_value': total_stock_value,
        'low_stock_count': low_stock_count,
        'out_of_stock_count': out_of_stock_count,
        # Analytics
        'movements_last_30_days': movements_last_30_days,
        'total_value_last_30_days': total_value_last_30_days,
        'top_categories': top_categories,
        'top_suppliers': top_suppliers,
        'recent_movements': recent_movements,
        'active_alerts': active_alerts,
        'overdue_pos': overdue_pos,
        'pending_stock_takes': pending_stock_takes,
    }
    
    return render(request, "inventory/reports/reports_dashboard.html", context)

@login_required
@inventory_permission_required('view')
def export_report(request, report_type):
    """
    Export inventory reports (valuation, etc) as Excel (CSV) or PDF (placeholder).
    """
    # Only 'valuation' is supported here; you can extend for more report types!
    if report_type != "valuation":
        return HttpResponse("Unsupported report type", status=400)

    export_format = request.GET.get('export', 'excel')

    # Filter products as in your valuation view
    products = Product.objects.filter(is_active=True).select_related('category', 'supplier')

    # Handle export format
    if export_format == "excel":
        # Export as CSV (Excel-compatible)
        response = HttpResponse(content_type='text/csv')
        filename = f"inventory_valuation_{timezone.now().date()}.csv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        writer = csv.writer(response)
        writer.writerow([
            "SKU", "Product Name", "Category", "Supplier",
            "Quantity", "Cost Price", "Selling Price", "Total Value", "Margin (%)"
        ])
        for p in products:
            margin_pct = ((p.selling_price - p.cost_price) / p.cost_price * 100) if p.cost_price else 0
            writer.writerow([
                p.sku, p.name,
                p.category.name if p.category else "",
                p.supplier.name if p.supplier else "",
                p.current_stock,
                "%.2f" % p.cost_price,
                "%.2f" % p.selling_price,
                "%.2f" % (p.current_stock * p.cost_price),
                "%.1f" % margin_pct
            ])
        return response

    elif export_format == "pdf":
        # For real systems, use ReportLab, xhtml2pdf, or WeasyPrint here!
        # Here, return a simple HTML-to-PDF placeholder.
        context = {
            "products": products,
            "generated_on": timezone.now(),
        }
        # Render a minimal HTML template; real implementation would use a PDF generator
        return render(request, "inventory/reports/valuation_pdf.html", context)

    else:
        return HttpResponse("Unknown export format", status=400)

@login_required
@inventory_permission_required('edit')
def generate_reorder_recommendations_view(request):
    """
    Generate intelligent reorder recommendations based on:
    - Current stock levels
    - Sales velocity
    - Supplier lead times
    - Seasonal patterns
    """
    try:
        # Calculate sales velocity (last 30 days)
        thirty_days_ago = timezone.now() - timedelta(days=30)
        
        # Get products that need analysis
        products = Product.objects.filter(
            is_active=True,
            supplier__isnull=False
        ).select_related('supplier', 'category')
        
        recommendations = []
        for product in products:
            # Calculate sales velocity from stock movements
            sales_movements = product.stock_movements.filter(
                movement_type='sale',
                created_at__gte=thirty_days_ago
            ).aggregate(
                total_sold=Sum('quantity')
            )
            
            monthly_sales = abs(sales_movements['total_sold'] or 0)
            daily_sales = monthly_sales / 30 if monthly_sales > 0 else 0
            
            # Calculate days of stock remaining
            if daily_sales > 0 and product.current_stock > 0:
                days_remaining = product.current_stock / daily_sales
            else:
                days_remaining = 999  # Effectively infinite if no sales
            
            # Determine if reorder is needed
            supplier_lead_time = product.supplier.average_lead_time_days or 7
            safety_buffer = 5  # Extra safety days
            reorder_point = (supplier_lead_time + safety_buffer) * daily_sales
            
            should_reorder = (
                product.current_stock <= reorder_point or
                product.current_stock <= product.reorder_level
            )
            
            if should_reorder:
                # Calculate recommended order quantity
                if product.economic_order_quantity:
                    recommended_qty = product.economic_order_quantity
                else:
                    # Order for 60 days worth of sales (minimum)
                    recommended_qty = max(
                        int(daily_sales * 60),
                        product.reorder_level - product.current_stock + 10
                    )
                
                # Calculate priority
                if product.current_stock == 0:
                    priority = 'critical'
                elif days_remaining <= supplier_lead_time:
                    priority = 'high'
                elif days_remaining <= supplier_lead_time + 7:
                    priority = 'medium'
                else:
                    priority = 'low'
                
                recommendations.append({
                    'product': product,
                    'current_stock': product.current_stock,
                    'reorder_level': product.reorder_level,
                    'monthly_sales': monthly_sales,
                    'daily_sales': round(daily_sales, 2),
                    'days_remaining': round(days_remaining, 1),
                    'recommended_qty': recommended_qty,
                    'estimated_cost': recommended_qty * product.cost_price,
                    'priority': priority,
                    'supplier': product.supplier,
                })
        
        # Sort by priority and days remaining
        priority_order = {'critical': 0, 'high': 1, 'medium': 2, 'low': 3}
        recommendations.sort(key=lambda x: (priority_order[x['priority']], x['days_remaining']))
        
        if request.method == 'POST':
            # Create reorder alerts for recommendations
            alerts_created = 0
            for rec in recommendations:
                if rec['priority'] in ['critical', 'high']:  # Auto-create for urgent items
                    alert, created = ReorderAlert.objects.get_or_create(
                        product=rec['product'],
                        status='active',
                        defaults={
                            'priority': rec['priority'],
                            'current_stock': rec['current_stock'],
                            'reorder_level': rec['product'].reorder_level,
                            'suggested_order_quantity': rec['recommended_qty'],
                            'suggested_supplier': rec['supplier'],
                            'estimated_cost': rec['estimated_cost'],
                        }
                    )
                    if created:
                        alerts_created += 1
            
            messages.success(request, f'Created {alerts_created} new reorder alerts')
            return redirect('inventory:reorder_alert_list')
        
        return render(request, 'inventory/reorder/reorder_recommendations.html', {
            'page_title': 'Reorder Recommendations',
            'recommendations': recommendations,
            'total_recommendations': len(recommendations),
            'critical_count': sum(1 for r in recommendations if r['priority'] == 'critical'),
            'high_count': sum(1 for r in recommendations if r['priority'] == 'high'),
        })
        
    except Exception as e:
        messages.error(request, f'Failed to generate recommendations: {str(e)}')
        return redirect('inventory:dashboard')

@login_required
@inventory_permission_required('edit')
@require_POST
def acknowledge_reorder_alert_view(request, alert_id):
    """
    Acknowledge a reorder alert to indicate it's being handled.
    """
    try:
        alert = get_object_or_404(ReorderAlert, id=alert_id)
        
        if alert.status == 'active':
            alert.status = 'acknowledged'
            alert.acknowledged_by = request.user
            alert.acknowledged_at = timezone.now()
            alert.save()
            
            messages.success(request, f'Acknowledged reorder alert for {alert.product.name}')
        else:
            messages.info(request, 'Alert already acknowledged or resolved')
        
    except Exception as e:
        messages.error(request, f'Failed to acknowledge alert: {str(e)}')
    
    return redirect('inventory:reorder_alert_list')

@login_required
@inventory_permission_required('edit')
@require_POST
def complete_reorder_alert_view(request, alert_id):
    """
    Mark a reorder alert as resolved/complete.
    """
    try:
        alert = get_object_or_404(ReorderAlert, id=alert_id)
        
        if alert.status in ['active', 'acknowledged']:
            alert.status = 'resolved'
            alert.resolved_at = timezone.now()
            alert.save()
            
            messages.success(request, f'Completed reorder alert for {alert.product.name}')
        else:
            messages.info(request, 'Alert already resolved')
        
    except Exception as e:
        messages.error(request, f'Failed to complete alert: {str(e)}')
    
    return redirect('inventory:reorder_alert_list')

@login_required
@inventory_permission_required('view')
def download_reorder_csv(request):
    """
    Export reorder recommendations as CSV for procurement teams.
    """
    try:
        # Get active reorder alerts
        alerts = ReorderAlert.objects.filter(
            status__in=['active', 'acknowledged']
        ).select_related('product', 'suggested_supplier').order_by('-priority', 'product__name')
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="reorder_recommendations.csv"'
        
        writer = csv.writer(response)
        writer.writerow([
            'SKU', 'Product Name', 'Category', 'Current Stock', 'Reorder Level',
            'Suggested Quantity', 'Priority', 'Supplier', 'Estimated Cost',
            'Currency', 'Lead Time Days', 'Alert Created'
        ])
        
        for alert in alerts:
            product = alert.product
            supplier = alert.suggested_supplier
            
            writer.writerow([
                product.sku,
                product.name,
                product.category.name if product.category else '',
                alert.current_stock,
                alert.reorder_level,
                alert.suggested_order_quantity,
                alert.get_priority_display(),
                supplier.name if supplier else '',
                alert.estimated_cost or '',
                supplier.currency if supplier else '',
                supplier.average_lead_time_days if supplier else '',
                alert.created_at.strftime('%Y-%m-%d %H:%M'),
            ])
        
        return response
        
    except Exception as e:
        messages.error(request, f'Failed to export reorder data: {str(e)}')
        return redirect('inventory:reorder_alert_list')

@login_required
@inventory_permission_required('edit')
def bulk_create_reorder_alerts_view(request):
    """
    Bulk create reorder alerts based on current stock levels.
    """
    if request.method == 'POST':
        try:
            # Find all products that need reordering
            low_stock_products = Product.objects.filter(
                current_stock__lte=F('reorder_level'),
                is_active=True,
                supplier__isnull=False
            ).select_related('supplier')
            
            created_count = 0
            for product in low_stock_products:
                # Check if alert already exists
                existing_alert = ReorderAlert.objects.filter(
                    product=product,
                    status__in=['active', 'acknowledged']
                ).exists()
                
                if not existing_alert:
                    # Determine priority based on stock level
                    if product.current_stock == 0:
                        priority = 'critical'
                    elif product.current_stock <= (product.reorder_level * 0.5):
                        priority = 'high'
                    elif product.current_stock <= (product.reorder_level * 0.8):
                        priority = 'medium'
                    else:
                        priority = 'low'
                    
                    # Calculate suggested order quantity
                    suggested_qty = max(
                        product.economic_order_quantity or product.reorder_level,
                        product.reorder_level - product.current_stock + 10
                    )
                    
                    ReorderAlert.objects.create(
                        product=product,
                        priority=priority,
                        status='active',
                        current_stock=product.current_stock,
                        reorder_level=product.reorder_level,
                        suggested_order_quantity=suggested_qty,
                        suggested_supplier=product.supplier,
                        estimated_cost=suggested_qty * product.cost_price,
                    )
                    created_count += 1
            
            messages.success(request, f'Created {created_count} new reorder alerts')
            
        except Exception as e:
            messages.error(request, f'Failed to create bulk alerts: {str(e)}')
    
    return redirect('inventory:reorder_alert_list')

@login_required
@inventory_permission_required('edit')
def create_purchase_order_from_alerts_view(request):
    """
    Create purchase orders from selected reorder alerts.
    """
    if request.method == 'POST':
        try:
            alert_ids = request.POST.getlist('alert_ids')
            
            if not alert_ids:
                messages.error(request, 'No alerts selected')
                return redirect('inventory:reorder_alert_list')
            
            alerts = ReorderAlert.objects.filter(
                id__in=alert_ids,
                status__in=['active', 'acknowledged']
            ).select_related('product', 'suggested_supplier')
            
            # Group alerts by supplier
            supplier_groups = {}
            for alert in alerts:
                supplier = alert.suggested_supplier
                if supplier not in supplier_groups:
                    supplier_groups[supplier] = []
                supplier_groups[supplier].append(alert)
            
            created_pos = []
            with transaction.atomic():
                for supplier, supplier_alerts in supplier_groups.items():
                    # Generate PO number
                    po_number = f"PO-{timezone.now().strftime('%Y%m%d')}-{supplier.supplier_code}"
                    counter = 1
                    while PurchaseOrder.objects.filter(po_number=po_number).exists():
                        po_number = f"PO-{timezone.now().strftime('%Y%m%d')}-{supplier.supplier_code}-{counter:02d}"
                        counter += 1
                    
                    # Create PO
                    po = PurchaseOrder.objects.create(
                        po_number=po_number,
                        supplier=supplier,
                        status='draft',
                        order_date=timezone.now().date(),
                        expected_delivery_date=timezone.now().date() + timedelta(
                            days=supplier.average_lead_time_days or 7
                        ),
                        currency=supplier.currency,
                        payment_terms=supplier.payment_terms,
                        notes=f"Generated from reorder alerts on {timezone.now().strftime('%Y-%m-%d')}",
                        created_by=request.user,
                    )
                    
                    # Add items
                    total_amount = Decimal('0')
                    for alert in supplier_alerts:
                        PurchaseOrderItem.objects.create(
                            purchase_order=po,
                            product=alert.product,
                            quantity_ordered=alert.suggested_order_quantity,
                            unit_price=alert.product.cost_price,
                            total_price=alert.suggested_order_quantity * alert.product.cost_price,
                        )
                        total_amount += alert.estimated_cost
                        
                        # Update alert status
                        alert.status = 'ordered'
                        alert.purchase_order = po
                        alert.resolved_at = timezone.now()
                        alert.save()
                    
                    # Update PO totals
                    po.subtotal = total_amount
                    po.total_amount = total_amount
                    po.save()
                    
                    created_pos.append(po)
            
            messages.success(
                request,
                f'Successfully created {len(created_pos)} purchase orders from alerts'
            )
            return redirect('inventory:purchase_order_list')
            
        except Exception as e:
            messages.error(request, f'Failed to create purchase orders: {str(e)}')
    
    return redirect('inventory:reorder_alert_list')

@login_required
@inventory_permission_required('view')
def stock_aging_report(request):
    """
    Stock aging report to identify slow-moving and dead stock.
    Critical for inventory optimization and cash flow management.
    """
    try:
        # Date ranges for aging analysis
        today = timezone.now().date()
        date_ranges = [
            ('0-30', today - timedelta(days=30), today),
            ('31-60', today - timedelta(days=60), today - timedelta(days=31)),
            ('61-90', today - timedelta(days=90), today - timedelta(days=61)),
            ('91-180', today - timedelta(days=180), today - timedelta(days=91)),
            ('180+', None, today - timedelta(days=180)),
        ]
        
        # Get all active products with their last movement dates
        products = Product.objects.filter(
            is_active=True,
            current_stock__gt=0
        ).select_related('category', 'supplier').annotate(
            last_movement_date=Max('stock_movements__created_at')
        )
        
        aging_data = []
        total_value = Decimal('0')
        
        for product in products:
            last_movement = product.last_movement_date
            
            if last_movement:
                days_since_movement = (today - last_movement.date()).days
            else:
                # If no movements, use creation date
                days_since_movement = (today - product.created_at.date()).days
            
            # Determine aging category
            aging_category = '180+'
            for category, start_date, end_date in date_ranges:
                if start_date is None:  # 180+ category
                    if days_since_movement >= 180:
                        aging_category = category
                        break
                else:
                    if start_date <= (today - timedelta(days=days_since_movement)) <= end_date:
                        aging_category = category
                        break
            
            stock_value = product.current_stock * product.cost_price
            total_value += stock_value
            
            aging_data.append({
                'product': product,
                'days_since_movement': days_since_movement,
                'last_movement_date': last_movement,
                'aging_category': aging_category,
                'stock_value': stock_value,
                'current_stock': product.current_stock,
            })
        
        # Group by aging category
        aging_summary = {}
        for category, _, _ in date_ranges:
            category_items = [item for item in aging_data if item['aging_category'] == category]
            category_value = sum(item['stock_value'] for item in category_items)
            
            aging_summary[category] = {
                'count': len(category_items),
                'value': category_value,
                'percentage': (category_value / total_value * 100) if total_value > 0 else 0,
                'items': sorted(category_items, key=lambda x: x['days_since_movement'], reverse=True)[:10]  # Top 10
            }
        
        return render(request, 'inventory/reports/stock_aging_report.html', {
            'page_title': 'Stock Aging Report',
            'aging_summary': aging_summary,
            'total_value': total_value,
            'total_products': len(aging_data),
            'date_ranges': date_ranges,
        })
        
    except Exception as e:
        messages.error(request, f'Failed to generate stock aging report: {str(e)}')
        return redirect('inventory:inventory_reports')

@login_required
@inventory_permission_required('view')
def inventory_turnover_report(request):
    """
    Inventory turnover analysis for performance metrics.
    Shows how efficiently inventory is being managed.
    """
    try:
        # Get date range (default last 12 months)
        end_date = timezone.now().date()
        start_date = end_date - timedelta(days=365)
        
        if request.GET.get('start_date'):
            start_date = datetime.strptime(request.GET.get('start_date'), '%Y-%m-%d').date()
        if request.GET.get('end_date'):
            end_date = datetime.strptime(request.GET.get('end_date'), '%Y-%m-%d').date()
        
        # Calculate turnover for each product
        products = Product.objects.filter(is_active=True).select_related('category', 'supplier')
        turnover_data = []
        
        for product in products:
            # Calculate average inventory (simplified: current stock)
            avg_inventory = product.current_stock
            
            # Calculate cost of goods sold (from sales movements)
            cogs = product.stock_movements.filter(
                movement_type='sale',
                created_at__date__range=[start_date, end_date]
            ).aggregate(
                total_sold=Sum('quantity')
            )['total_sold'] or 0
            
            cogs_value = cogs * product.cost_price
            
            # Calculate turnover ratio
            if avg_inventory > 0:
                turnover_ratio = cogs / avg_inventory
                days_on_hand = 365 / turnover_ratio if turnover_ratio > 0 else 365
            else:
                turnover_ratio = 0
                days_on_hand = 0 if cogs > 0 else 365
            
            # Classify performance
            if turnover_ratio >= 12:  # Monthly turnover
                performance = 'Excellent'
            elif turnover_ratio >= 6:  # Bi-monthly
                performance = 'Good'
            elif turnover_ratio >= 4:  # Quarterly
                performance = 'Average'
            elif turnover_ratio >= 2:  # Semi-annual
                performance = 'Below Average'
            else:
                performance = 'Poor'
            
            turnover_data.append({
                'product': product,
                'avg_inventory': avg_inventory,
                'cogs': cogs,
                'cogs_value': cogs_value,
                'turnover_ratio': round(turnover_ratio, 2),
                'days_on_hand': round(days_on_hand, 1),
                'performance': performance,
                'stock_value': avg_inventory * product.cost_price,
            })
        
        # Sort by turnover ratio (descending)
        turnover_data.sort(key=lambda x: x['turnover_ratio'], reverse=True)
        
        # Calculate summary statistics
        total_stock_value = sum(item['stock_value'] for item in turnover_data)
        avg_turnover = sum(item['turnover_ratio'] for item in turnover_data) / len(turnover_data) if turnover_data else 0
        
        # Performance distribution
        performance_summary = {}
        for item in turnover_data:
            perf = item['performance']
            if perf not in performance_summary:
                performance_summary[perf] = {'count': 0, 'value': 0}
            performance_summary[perf]['count'] += 1
            performance_summary[perf]['value'] += item['stock_value']
        
        return render(request, 'inventory/reports/inventory_turnover_report.html', {
            'page_title': 'Inventory Turnover Report',
            'turnover_data': turnover_data,
            'total_stock_value': total_stock_value,
            'avg_turnover': round(avg_turnover, 2),
            'performance_summary': performance_summary,
            'start_date': start_date,
            'end_date': end_date,
            'date_range_days': (end_date - start_date).days,
        })
        
    except Exception as e:
        messages.error(request, f'Failed to generate turnover report: {str(e)}')
        return redirect('inventory:inventory_reports')

@login_required
@inventory_permission_required('view')
def abc_analysis_report(request):
    """
    ABC analysis to classify products by value contribution.
    A = High value (70-80% of total), B = Medium (15-20%), C = Low (5-10%)
    """
    try:
        # Get all products with their stock values
        products = Product.objects.filter(is_active=True).select_related('category', 'supplier')
        
        product_data = []
        total_value = Decimal('0')
        
        for product in products:
            stock_value = product.current_stock * product.cost_price
            total_value += stock_value
            
            # Calculate annual sales (last 12 months)
            annual_sales = product.stock_movements.filter(
                movement_type='sale',
                created_at__gte=timezone.now() - timedelta(days=365)
            ).aggregate(
                total_sold=Sum('quantity')
            )['total_sold'] or 0
            
            annual_sales_value = annual_sales * product.selling_price
            
            product_data.append({
                'product': product,
                'stock_value': stock_value,
                'annual_sales': annual_sales,
                'annual_sales_value': annual_sales_value,
            })
        
        # Sort by stock value (descending)
        product_data.sort(key=lambda x: x['stock_value'], reverse=True)
        
        # Calculate cumulative percentages and assign ABC categories
        cumulative_value = Decimal('0')
        for i, item in enumerate(product_data):
            cumulative_value += item['stock_value']
            cumulative_percentage = (cumulative_value / total_value * 100) if total_value > 0 else 0
            
            # Assign ABC category
            if cumulative_percentage <= 80:
                category = 'A'
            elif cumulative_percentage <= 95:
                category = 'B'
            else:
                category = 'C'
            
            item['cumulative_value'] = cumulative_value
            item['cumulative_percentage'] = round(cumulative_percentage, 2)
            item['abc_category'] = category
            item['rank'] = i + 1
        
        # Calculate category summaries
        abc_summary = {}
        for category in ['A', 'B', 'C']:
            category_items = [item for item in product_data if item['abc_category'] == category]
            category_value = sum(item['stock_value'] for item in category_items)
            
            abc_summary[category] = {
                'count': len(category_items),
                'value': category_value,
                'percentage': (category_value / total_value * 100) if total_value > 0 else 0,
                'items': category_items[:20]  # Top 20 for display
            }
        
        return render(request, 'inventory/reports/abc_analysis_report.html', {
            'page_title': 'ABC Analysis Report',
            'abc_summary': abc_summary,
            'total_value': total_value,
            'total_products': len(product_data),
        })
        
    except Exception as e:
        messages.error(request, f'Failed to generate ABC analysis: {str(e)}')
        return redirect('inventory:inventory_reports')

@login_required
@inventory_permission_required('view')
def supplier_performance_report(request):
    """
    Comprehensive supplier performance analysis.
    Evaluates delivery times, quality, reliability, and cost-effectiveness.
    """
    try:
        # Get all active suppliers
        suppliers = Supplier.objects.filter(is_active=True)
        supplier_data = []
        
        for supplier in suppliers:
            # Get purchase orders from last 12 months
            recent_pos = supplier.purchase_orders.filter(
                order_date__gte=timezone.now().date() - timedelta(days=365)
            )
            
            # Calculate metrics
            total_orders = recent_pos.count()
            total_value = recent_pos.aggregate(Sum('total_amount'))['total_amount'] or Decimal('0')
            
            # Delivery performance
            completed_orders = recent_pos.filter(status='received')
            on_time_orders = completed_orders.filter(
                actual_delivery_date__lte=F('expected_delivery_date')
            ).count()
            
            delivery_performance = (on_time_orders / completed_orders.count() * 100) if completed_orders.count() > 0 else 0
            
            # Average lead time
            actual_lead_times = []
            for po in completed_orders:
                if po.actual_delivery_date and po.order_date:
                    lead_time = (po.actual_delivery_date - po.order_date).days
                    actual_lead_times.append(lead_time)
            
            avg_lead_time = sum(actual_lead_times) / len(actual_lead_times) if actual_lead_times else 0
            
            # Quality metrics (from received items)
            quality_issues = 0
            total_items_received = 0
            for po in completed_orders:
                for item in po.items.all():
                    total_items_received += item.quantity_received
                    if item.quality_check_passed is False:
                        quality_issues += item.quantity_received
            
            quality_rate = ((total_items_received - quality_issues) / total_items_received * 100) if total_items_received > 0 else 100
            
            # Cost competitiveness (simplified)
            avg_cost_per_item = total_value / total_items_received if total_items_received > 0 else Decimal('0')
            
            # Overall score (weighted average)
            delivery_weight = 0.4
            quality_weight = 0.4
            reliability_weight = 0.2
            
            overall_score = (
                delivery_performance * delivery_weight +
                quality_rate * quality_weight +
                supplier.reliability_rating * 10 * reliability_weight  # Convert 1-10 to percentage
            )
            
            supplier_data.append({
                'supplier': supplier,
                'total_orders': total_orders,
                'total_value': total_value,
                'delivery_performance': round(delivery_performance, 1),
                'avg_lead_time': round(avg_lead_time, 1),
                'expected_lead_time': supplier.average_lead_time_days,
                'quality_rate': round(quality_rate, 1),
                'quality_issues': quality_issues,
                'total_items_received': total_items_received,
                'avg_cost_per_item': avg_cost_per_item,
                'overall_score': round(overall_score, 1),
                'reliability_rating': supplier.reliability_rating,
            })
        
        # Sort by overall score (descending)
        supplier_data.sort(key=lambda x: x['overall_score'], reverse=True)
        
        return render(request, 'inventory/reports/supplier_performance_report.html', {
            'page_title': 'Supplier Performance Report',
            'supplier_data': supplier_data,
            'total_suppliers': len(supplier_data),
            'report_period': '12 months',
        })
        
    except Exception as e:
        messages.error(request, f'Failed to generate supplier performance report: {str(e)}')
        return redirect('inventory:inventory_reports')

@login_required
@inventory_permission_required('view')
def category_analysis_report(request):
    """
    Category performance analysis for business insights.
    Shows which product categories are performing best.
    """
    try:
        categories = Category.objects.filter(is_active=True)
        category_data = []
        
        for category in categories:
            products = category.products.filter(is_active=True)
            
            # Stock metrics
            total_products = products.count()
            total_stock_value = sum(p.current_stock * p.cost_price for p in products)
            total_stock_units = sum(p.current_stock for p in products)
            
            # Sales metrics (last 6 months)
            six_months_ago = timezone.now() - timedelta(days=180)
            category_sales = StockMovement.objects.filter(
                product__category=category,
                movement_type='sale',
                created_at__gte=six_months_ago
            ).aggregate(
                total_units_sold=Sum('quantity'),
            )['total_units_sold'] or 0
            
            # Calculate average selling price for the category
            avg_selling_price = products.aggregate(Avg('selling_price'))['selling_price__avg'] or Decimal('0')
            sales_value = category_sales * avg_selling_price
            
            # Low stock products
            low_stock_count = products.filter(current_stock__lte=F('reorder_level')).count()
            
            # Profit margins
            avg_cost_price = products.aggregate(Avg('cost_price'))['cost_price__avg'] or Decimal('0')
            avg_margin = avg_selling_price - avg_cost_price
            avg_margin_percent = (avg_margin / avg_selling_price * 100) if avg_selling_price > 0 else 0
            
            # Turnover rate (simplified)
            turnover_rate = (category_sales / total_stock_units) if total_stock_units > 0 else 0
            
            category_data.append({
                'category': category,
                'total_products': total_products,
                'total_stock_value': total_stock_value,
                'total_stock_units': total_stock_units,
                'sales_units': category_sales,
                'sales_value': sales_value,
                'low_stock_count': low_stock_count,
                'avg_selling_price': avg_selling_price,
                'avg_cost_price': avg_cost_price,
                'avg_margin': avg_margin,
                'avg_margin_percent': round(avg_margin_percent, 2),
                'turnover_rate': round(turnover_rate, 2),
            })
        
        # Sort by sales value (descending)
        category_data.sort(key=lambda x: x['sales_value'], reverse=True)
        
        # Calculate totals
        total_sales_value = sum(item['sales_value'] for item in category_data)
        total_stock_value = sum(item['total_stock_value'] for item in category_data)
        
        return render(request, 'inventory/reports/category_analysis_report.html', {
            'page_title': 'Category Analysis Report',
            'category_data': category_data,
            'total_categories': len(category_data),
            'total_sales_value': total_sales_value,
            'total_stock_value': total_stock_value,
            'report_period': '6 months',
        })
        
    except Exception as e:
        messages.error(request, f'Failed to generate category analysis: {str(e)}')
        return redirect('inventory:inventory_reports')

@login_required
@inventory_permission_required('view')
def supplier_comparison_report(request):
    """Compare suppliers"""
    context = {
        'page_title': 'Supplier Comparison Report',
    }
    return render(request, 'inventory/reports/supplier_comparison.html', context)

@login_required
@inventory_permission_required('view')
def purchase_analysis_report(request):
    """Purchase analysis report"""
    context = {
        'page_title': 'Purchase Analysis Report',
    }
    return render(request, 'inventory/reports/purchase_analysis.html', context)

@login_required
@inventory_permission_required('view')
def cost_analysis_report(request):
    """Cost analysis report"""
    context = {
        'page_title': 'Cost Analysis Report',
    }
    return render(request, 'inventory/reports/cost_analysis.html', context)

@login_required
@inventory_permission_required('view')
def margin_analysis_report(request):
    """Margin analysis report"""
    context = {
        'page_title': 'Margin Analysis Report',
    }
    return render(request, 'inventory/reports/margin_analysis.html', context)

@login_required
@inventory_permission_required('view')
def profitability_report(request):
    """Profitability report"""
    context = {
        'page_title': 'Profitability Report',
    }
    return render(request, 'inventory/reports/profitability.html', context)

@login_required
@inventory_permission_required('view')
def tax_compliance_report(request):
    """Tax compliance report"""
    context = {
        'page_title': 'Tax Compliance Report',
    }
    return render(request, 'inventory/reports/tax_compliance.html', context)

@login_required
@inventory_permission_required('view')
def brand_performance_report(request):
    """Brand performance report"""
    brands = Brand.objects.filter(is_active=True).annotate(
        product_count=Count('products', filter=Q(products__is_active=True))
    )
    
    context = {
        'page_title': 'Brand Performance Report',
        'brands': brands,
        'report_date': timezone.now().date(),
    }
    return render(request, 'inventory/reports/brand_performance.html', context)

@login_required
@inventory_permission_required('view')
def executive_summary_report(request):
    """Executive summary report"""
    # Calculate key metrics
    total_products = Product.objects.filter(is_active=True).count()
    total_stock_value = Product.objects.filter(is_active=True).aggregate(
        total=Sum(F('current_stock') * F('cost_price'))
    )['total'] or 0
    
    low_stock_count = Product.objects.filter(
        is_active=True,
        current_stock__lte=F('reorder_level')
    ).count()
    
    pending_po_count = PurchaseOrder.objects.filter(
        status__in=['draft', 'sent', 'acknowledged']
    ).count()
    
    context = {
        'page_title': 'Executive Summary Report',
        'total_products': total_products,
        'total_stock_value': total_stock_value,
        'low_stock_count': low_stock_count,
        'pending_po_count': pending_po_count,
        'report_date': timezone.now().date(),
    }
    return render(request, 'inventory/reports/executive_summary.html', context)

# =====================================
# STOCK TAKE MANAGEMENT VIEWS
# =====================================

class StockTakeListView(LoginRequiredMixin, ListView):
    """
    Manage stock take operations for inventory accuracy.
    """
    model = StockTake
    template_name = 'inventory/stock_take/stock_take_list.html'
    context_object_name = 'stock_takes'
    paginate_by = 20
    
    @method_decorator(stock_take_permission)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    
    def get_queryset(self):
        return StockTake.objects.select_related(
            'location', 'created_by'
        ).order_by('-created_at')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Calculate summary metrics
        stock_takes = self.get_queryset()
        summary = {
            'total_stock_takes': stock_takes.count(),
            'pending': stock_takes.filter(status='pending').count(),
            'in_progress': stock_takes.filter(status='in_progress').count(),
            'completed': stock_takes.filter(status='completed').count(),
        }
        
        context.update({
            'page_title': 'Stock Take Management',
            'summary': summary,
        })
        return context

class StockTakeCreateView(LoginRequiredMixin, CreateView):
    """
    Create new stock take operations.
    """
    model = StockTake
    form_class = StockTakeForm
    template_name = 'inventory/stock_take/stock_take_form.html'
    
    @method_decorator(stock_take_permission)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    
    def form_valid(self, form):
        form.instance.created_by = self.request.user
        response = super().form_valid(form)
        
        # Create stock take items for all products at the location
        if form.instance.location:
            stock_levels = StockLevel.objects.filter(location=form.instance.location)
            for stock_level in stock_levels:
                StockTakeItem.objects.create(
                    stock_take=form.instance,
                    product=stock_level.product,
                    expected_quantity=stock_level.quantity,
                )
        else:
            # All locations - create for all products
            products = Product.objects.filter(is_active=True)
            for product in products:
                StockTakeItem.objects.create(
                    stock_take=form.instance,
                    product=product,
                    expected_quantity=product.current_stock,
                )
        
        messages.success(self.request, 'Stock take created successfully')
        return response
    
    def get_success_url(self):
        return reverse('inventory:stock_take_detail', kwargs={'pk': self.object.pk})

class StockTakeDetailView(LoginRequiredMixin, DetailView):
    """
    Detailed view of stock take with item management.
    """
    model = StockTake
    template_name = 'inventory/stock_take/stock_take_detail.html'
    context_object_name = 'stock_take'
    
    @method_decorator(stock_take_permission)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        stock_take = self.get_object()
        
        items = stock_take.items.select_related('product').order_by('product__name')
        
        # Calculate progress
        total_items = items.count()
        counted_items = items.exclude(actual_quantity__isnull=True).count()
        progress = (counted_items / total_items * 100) if total_items > 0 else 0
        
        # Calculate variances
        variances = []
        total_variance_value = Decimal('0')
        
        for item in items.exclude(actual_quantity__isnull=True):
            variance = item.variance
            variance_value = variance * item.product.cost_price
            total_variance_value += abs(variance_value)
            
            if variance != 0:
                variances.append({
                    'item': item,
                    'variance': variance,
                    'variance_value': variance_value,
                })
        
        context.update({
            'page_title': f'Stock Take: {stock_take.reference}',
            'items': items,
            'total_items': total_items,
            'counted_items': counted_items,
            'progress': round(progress, 1),
            'variances': sorted(variances, key=lambda x: abs(x['variance_value']), reverse=True),
            'total_variance_value': total_variance_value,
        })
        return context

@login_required
@stock_take_permission
@require_POST
def stock_take_complete_view(request, pk):
    """
    Complete a stock take and apply adjustments.
    """
    try:
        stock_take = get_object_or_404(StockTake, pk=pk)
        
        if stock_take.status != 'in_progress':
            messages.error(request, 'Only stock takes in progress can be completed')
            return redirect('inventory:stock_take_detail', pk=pk)
        
        # Check if all items have been counted
        uncounted_items = stock_take.items.filter(actual_quantity__isnull=True).count()
        if uncounted_items > 0:
            messages.error(request, f'{uncounted_items} items still need to be counted')
            return redirect('inventory:stock_take_detail', pk=pk)
        
        with transaction.atomic():
            adjustments_made = 0
            total_variance_value = Decimal('0')
            
            for item in stock_take.items.all():
                if item.variance != 0:
                    # Update stock levels
                    if stock_take.location:
                        # Specific location
                        stock_level, created = StockLevel.objects.get_or_create(
                            product=item.product,
                            location=stock_take.location,
                            defaults={'quantity': 0}
                        )
                        old_quantity = stock_level.quantity
                        stock_level.quantity = item.actual_quantity
                        stock_level.last_counted = timezone.now()
                        stock_level.save()
                    else:
                        # Overall stock
                        old_quantity = item.product.current_stock
                        item.product.current_stock = item.actual_quantity
                        item.product.last_stock_check = timezone.now()
                        item.product.save()
                    
                    # Create stock movement
                    StockMovement.objects.create(
                        product=item.product,
                        movement_type='adjustment',
                        quantity=item.variance,
                        from_location=stock_take.location if item.variance < 0 else None,
                        to_location=stock_take.location if item.variance > 0 else None,
                        previous_stock=old_quantity,
                        new_stock=item.actual_quantity,
                        reference=f"STOCK-TAKE-{stock_take.reference}",
                        notes=f"Stock take adjustment. Expected: {item.expected_quantity}, Actual: {item.actual_quantity}",
                        created_by=request.user,
                    )
                    
                    adjustments_made += 1
                    total_variance_value += abs(item.variance * item.product.cost_price)
            
            # Complete the stock take
            stock_take.status = 'completed'
            stock_take.completed_at = timezone.now()
            stock_take.total_variance_value = total_variance_value
            stock_take.save()
            
            messages.success(
                request,
                f'Stock take completed. {adjustments_made} adjustments made with total variance value of ${total_variance_value:.2f}'
            )
        
    except Exception as e:
        messages.error(request, f'Failed to complete stock take: {str(e)}')
    
    return redirect('inventory:stock_take_detail', pk=pk)

# =====================================
# DATA MANAGEMENT VIEWS
# =====================================

class DataExportView(LoginRequiredMixin, TemplateView):
    """Export inventory data"""
    template_name = 'inventory/data/data_export.html'
    
    @method_decorator(inventory_permission_required('view'))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

class DataImportView(LoginRequiredMixin, TemplateView):
    """Import inventory data"""
    template_name = 'inventory/data/data_import.html'
    
    @method_decorator(inventory_permission_required('edit'))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

@login_required
@inventory_permission_required('admin')
def data_validation_view(request):
    """
    Comprehensive data validation and integrity checking.
    """
    try:
        validation_results = {
            'product_issues': [],
            'stock_issues': [],
            'pricing_issues': [],
            'supplier_issues': [],
        }
        
        # Product validation
        products = Product.objects.select_related('category', 'supplier')
        
        for product in products:
            issues = []
            
            # Check required fields
            if not product.sku:
                issues.append('Missing SKU')
            if not product.name:
                issues.append('Missing name')
            
            # Check pricing
            if product.cost_price <= 0:
                issues.append('Invalid cost price')
            if product.selling_price <= 0:
                issues.append('Invalid selling price')
            if product.selling_price < product.cost_price:
                issues.append('Selling price below cost price')
            
            # Check stock
            if product.current_stock < 0:
                issues.append('Negative stock')
            if product.reorder_level < 0:
                issues.append('Negative reorder level')
            
            # Check relationships
            if not product.category:
                issues.append('Missing category')
            if not product.supplier:
                issues.append('Missing supplier')
            
            if issues:
                validation_results['product_issues'].append({
                    'product': product,
                    'issues': issues,
                })
        
        # Stock level validation
        stock_levels = StockLevel.objects.select_related('product', 'location')
        
        for stock_level in stock_levels:
            issues = []
            
            if stock_level.quantity < 0:
                issues.append('Negative quantity')
            if stock_level.reserved_quantity < 0:
                issues.append('Negative reserved quantity')
            if stock_level.reserved_quantity > stock_level.quantity:
                issues.append('Reserved quantity exceeds total quantity')
            
            if issues:
                validation_results['stock_issues'].append({
                    'stock_level': stock_level,
                    'issues': issues,
                })
        
        # Check for duplicate SKUs
        from django.db.models import Count
        duplicate_skus = Product.objects.values('sku').annotate(
            count=Count('sku')
        ).filter(count__gt=1)
        
        for sku_data in duplicate_skus:
            products = Product.objects.filter(sku=sku_data['sku'])
            validation_results['product_issues'].append({
                'product': f"Duplicate SKU: {sku_data['sku']}",
                'issues': [f"Found {sku_data['count']} products with same SKU"],
                'products': list(products),
            })
        
        # Supplier validation
        suppliers = Supplier.objects.all()
        for supplier in suppliers:
            issues = []
            
            if not supplier.email and not supplier.phone:
                issues.append('No contact information')
            if supplier.average_lead_time_days < 0:
                issues.append('Negative lead time')
            if supplier.minimum_order_amount < 0:
                issues.append('Negative minimum order amount')
            
            if issues:
                validation_results['supplier_issues'].append({
                    'supplier': supplier,
                    'issues': issues,
                })
        
        return render(request, 'inventory/data/data_validation.html', {
            'page_title': 'Data Validation Report',
            'validation_results': validation_results,
            'total_issues': sum(len(issues) for issues in validation_results.values()),
        })
        
    except Exception as e:
        messages.error(request, f'Validation failed: {str(e)}')
        return redirect('inventory:dashboard')

@login_required
@inventory_permission_required('admin')
def data_cleanup_view(request):
    """
    Data cleanup and correction tools.
    """
    if request.method == 'POST':
        cleanup_type = request.POST.get('cleanup_type')
        
        try:
            if cleanup_type == 'fix_negative_stock':
                # Fix negative stock levels
                negative_products = Product.objects.filter(current_stock__lt=0)
                fixed_count = 0
                
                for product in negative_products:
                    product.current_stock = 0
                    product.save()
                    fixed_count += 1
                    
                    # Create adjustment record
                    StockMovement.objects.create(
                        product=product,
                        movement_type='adjustment',
                        quantity=abs(product.current_stock),
                        previous_stock=product.current_stock,
                        new_stock=0,
                        reference=f"CLEANUP-{timezone.now().strftime('%Y%m%d%H%M%S')}",
                        notes="Automated cleanup: Fixed negative stock",
                        created_by=request.user,
                    )
                
                messages.success(request, f'Fixed {fixed_count} products with negative stock')
                
            elif cleanup_type == 'remove_empty_categories':
                # Remove categories with no products
                empty_categories = Category.objects.filter(products__isnull=True, is_active=True)
                count = empty_categories.count()
                empty_categories.update(is_active=False)
                messages.success(request, f'Deactivated {count} empty categories')
                
            elif cleanup_type == 'fix_missing_barcodes':
                # Generate barcodes for products without them
                products_without_barcodes = Product.objects.filter(
                    Q(barcode='') | Q(barcode__isnull=True),
                    is_active=True
                )
                
                generated_count = 0
                for product in products_without_barcodes:
                    timestamp = timezone.now().strftime('%Y%m%d%H%M%S')
                    barcode = f"BT{product.id:06d}{timestamp[-4:]}"
                    
                    # Ensure uniqueness
                    counter = 1
                    original_barcode = barcode
                    while Product.objects.filter(barcode=barcode).exists():
                        barcode = f"{original_barcode}{counter:02d}"
                        counter += 1
                    
                    product.barcode = barcode
                    product.save()
                    generated_count += 1
                
                messages.success(request, f'Generated {generated_count} missing barcodes')
                
            elif cleanup_type == 'sync_stock_totals':
                # Sync product totals with stock levels
                synced_count = 0
                
                for product in Product.objects.filter(is_active=True):
                    total_stock = product.stock_levels.aggregate(
                        total=Sum('quantity')
                    )['total'] or 0
                    
                    if product.current_stock != total_stock:
                        product.current_stock = total_stock
                        product.save()
                        synced_count += 1
                
                messages.success(request, f'Synchronized {synced_count} product stock totals')
                
        except Exception as e:
            messages.error(request, f'Cleanup failed: {str(e)}')
    
    return render(request, 'inventory/data/data_cleanup.html', {
        'page_title': 'Data Cleanup Tools',
    })

@login_required
@inventory_permission_required('view')
def find_duplicates_view(request):
    """
    Find potential duplicate products.
    """
    try:
        # Find products with similar names
        from difflib import SequenceMatcher
        
        products = Product.objects.filter(is_active=True).order_by('name')
        potential_duplicates = []
        
        # Compare each product with others
        for i, product1 in enumerate(products):
            for product2 in products[i+1:]:
                # Calculate similarity
                similarity = SequenceMatcher(None, product1.name.lower(), product2.name.lower()).ratio()
                
                if similarity > 0.8:  # 80% similarity threshold
                    potential_duplicates.append({
                        'product1': product1,
                        'product2': product2,
                        'similarity': round(similarity * 100, 1),
                        'same_supplier': product1.supplier == product2.supplier,
                        'same_category': product1.category == product2.category,
                    })
        
        # Sort by similarity (highest first)
        potential_duplicates.sort(key=lambda x: x['similarity'], reverse=True)
        
        return render(request, 'inventory/data/find_duplicates.html', {
            'page_title': 'Find Duplicate Products',
            'potential_duplicates': potential_duplicates[:50],  # Limit to top 50
            'total_found': len(potential_duplicates),
        })
        
    except Exception as e:
        messages.error(request, f'Duplicate search failed: {str(e)}')
        return redirect('inventory:dashboard')

@login_required
@inventory_permission_required('view')
def import_templates_view(request):
    """Download import templates"""
    template_type = request.GET.get('type', 'products')
    
    if template_type == 'products':
        return product_import_template_view(request)
    
    # Add other template types as needed
    return HttpResponse('Template not found', status=404)

@login_required
@inventory_permission_required('admin')
def data_backup_view(request):
    """Backup inventory data"""
    # Implementation for data backup
    context = {
        'page_title': 'Data Backup',
    }
    return render(request, 'inventory/data/backup.html', context)

@login_required
@inventory_permission_required('admin')
def data_restore_view(request):
    """Restore inventory data"""
    # Implementation for data restore
    context = {
        'page_title': 'Data Restore',
    }
    return render(request, 'inventory/data/restore.html', context)

# =====================================
# MOBILE AND BARCODE FEATURES
# =====================================

class MobileDashboardView(LoginRequiredMixin, TemplateView):
    """
    Mobile-optimized inventory dashboard for warehouse staff.
    """
    template_name = 'inventory/mobile/mobile_dashboard.html'
    
    @method_decorator(inventory_permission_required('view'))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Quick stats for mobile view
        low_stock_count = Product.objects.filter(
            current_stock__lte=F('reorder_level'),
            is_active=True
        ).count()
        
        pending_stock_takes = StockTake.objects.filter(
            status__in=['pending', 'in_progress']
        ).count()
        
        recent_movements = StockMovement.objects.select_related(
            'product'
        ).order_by('-created_at')[:10]
        
        context.update({
            'page_title': 'Mobile Inventory',
            'low_stock_count': low_stock_count,
            'pending_stock_takes': pending_stock_takes,
            'recent_movements': recent_movements,
            'is_mobile': True,
        })
        return context

class MobileStockCheckView(LoginRequiredMixin, TemplateView):
    """
    Quick stock check interface for mobile devices.
    """
    template_name = 'inventory/mobile/mobile_stock_check.html'
    
    @method_decorator(inventory_permission_required('view'))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        search_term = self.request.GET.get('q', '').strip()
        product = None
        
        if search_term:
            # Try to find product by SKU or barcode first
            try:
                product = Product.objects.select_related(
                    'category', 'supplier'
                ).get(
                    Q(sku__iexact=search_term) | Q(barcode=search_term),
                    is_active=True
                )
            except Product.DoesNotExist:
                # Try partial name match
                products = Product.objects.filter(
                    name__icontains=search_term,
                    is_active=True
                ).select_related('category', 'supplier')[:5]
                
                if products.count() == 1:
                    product = products.first()
                elif products.count() > 1:
                    context['multiple_products'] = products
        
        if product:
            # Get stock levels by location
            stock_levels = product.stock_levels.select_related('location').all()
            
            context.update({
                'product': product,
                'stock_levels': stock_levels,
                'stock_status': get_stock_status(product),
            })
        
        context.update({
            'page_title': 'Stock Check',
            'search_term': search_term,
            'is_mobile': True,
        })
        return context

class BarcodeGeneratorView(LoginRequiredMixin, TemplateView):
    """
    Generate barcodes for products that don't have them.
    """
    template_name = 'inventory/barcode/barcode_generator.html'
    
    @method_decorator(inventory_permission_required('edit'))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get products without barcodes
        products_without_barcodes = Product.objects.filter(
            is_active=True,
            barcode__in=['', None]
        ).select_related('category', 'supplier').order_by('name')
        
        context.update({
            'page_title': 'Barcode Generator',
            'products_without_barcodes': products_without_barcodes,
            'total_products': products_without_barcodes.count(),
        })
        return context

class BarcodeScannerView(LoginRequiredMixin, TemplateView):
    """
    Barcode scanner interface for desktop/mobile.
    """
    template_name = 'inventory/barcode/barcode_scanner.html'
    
    @method_decorator(inventory_permission_required('view'))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        context.update({
            'page_title': 'Barcode Scanner',
            'scan_mode': self.request.GET.get('mode', 'lookup'),  # lookup, adjust, transfer
        })
        return context

class MobileBarcodeScannerView(LoginRequiredMixin, TemplateView):
    """Mobile-optimized barcode scanner"""
    template_name = 'inventory/barcodes/mobile_scanner.html'
    
    @method_decorator(inventory_permission_required('view'))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

@csrf_exempt
@require_http_methods(["POST"])
def mobile_sync_api(request):
    """API for mobile app synchronization"""
    try:
        data = json.loads(request.body)
        last_sync = data.get('last_sync')
        
        # Return changes since last sync
        sync_data = {
            'products': [],
            'stock_levels': [],
            'movements': [],
            'timestamp': timezone.now().isoformat(),
        }
        
        return JsonResponse(sync_data)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

@csrf_exempt
@require_http_methods(["GET"])
def mobile_offline_data_api(request):
    """API for mobile offline data"""
    # Return essential data for offline operation
    products = Product.objects.filter(is_active=True).values(
        'id', 'name', 'sku', 'current_stock', 'reorder_level'
    )[:100]  # Limit for mobile
    
    return JsonResponse({
        'products': list(products),
        'timestamp': timezone.now().isoformat(),
    })

@csrf_exempt
@require_http_methods(["POST"])
def mobile_upload_batch_api(request):
    """API for mobile batch uploads"""
    try:
        data = json.loads(request.body)
        batch_data = data.get('batch_data', [])
        
        processed_count = 0
        for item in batch_data:
            # Process each batch item
            # This could be stock adjustments, transfers, etc.
            processed_count += 1
        
        return JsonResponse({
            'success': True,
            'processed_count': processed_count
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

@login_required
@inventory_permission_required('edit')
@require_POST
def mobile_quick_adjust_view(request):
    """
    Quick stock adjustment from mobile device.
    """
    try:
        product_id = request.POST.get('product_id')
        location_id = request.POST.get('location_id')
        new_quantity = int(request.POST.get('new_quantity', 0))
        reason = request.POST.get('reason', 'Mobile adjustment')
        
        product = get_object_or_404(Product, id=product_id)
        
        with transaction.atomic():
            if location_id:
                location = get_object_or_404(Location, id=location_id)
                stock_level, created = StockLevel.objects.get_or_create(
                    product=product,
                    location=location,
                    defaults={'quantity': 0}
                )
                
                old_quantity = stock_level.quantity
                adjustment = new_quantity - old_quantity
                
                stock_level.quantity = new_quantity
                stock_level.last_counted = timezone.now()
                stock_level.save()
                
                # Update product total
                product.current_stock = product.stock_levels.aggregate(
                    total=Sum('quantity')
                )['total'] or 0
                
            else:
                old_quantity = product.current_stock
                adjustment = new_quantity - old_quantity
                product.current_stock = new_quantity
            
            product.last_stock_check = timezone.now()
            product.save()
            
            # Create movement record
            StockMovement.objects.create(
                product=product,
                movement_type='adjustment',
                quantity=adjustment,
                from_location=location if location_id and adjustment < 0 else None,
                to_location=location if location_id and adjustment > 0 else None,
                previous_stock=old_quantity,
                new_stock=new_quantity,
                reference=f"MOBILE-ADJ-{timezone.now().strftime('%Y%m%d%H%M%S')}",
                notes=f"Mobile adjustment: {reason}",
                created_by=request.user
            )
            
        return JsonResponse({
            'success': True,
            'message': f'Stock updated: {product.name} adjusted by {adjustment}',
            'new_quantity': new_quantity,
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@login_required
@inventory_permission_required('edit')
def bulk_generate_barcodes_view(request):
    """
    Generate barcodes for multiple products at once.
    """
    if request.method == 'POST':
        try:
            product_ids = request.POST.getlist('product_ids')
            barcode_prefix = request.POST.get('prefix', 'BT')
            
            if not product_ids:
                messages.error(request, 'No products selected')
                return redirect('inventory:barcode_generator')
            
            generated_count = 0
            with transaction.atomic():
                for product_id in product_ids:
                    try:
                        product = Product.objects.get(id=product_id, is_active=True)
                        
                        if not product.barcode:
                            # Generate unique barcode
                            timestamp = timezone.now().strftime('%Y%m%d%H%M%S')
                            barcode = f"{barcode_prefix}{product.id:06d}{timestamp[-4:]}"
                            
                            # Ensure uniqueness
                            counter = 1
                            original_barcode = barcode
                            while Product.objects.filter(barcode=barcode).exists():
                                barcode = f"{original_barcode}{counter:02d}"
                                counter += 1
                            
                            product.barcode = barcode
                            product.save()
                            generated_count += 1
                            
                    except Product.DoesNotExist:
                        continue
            
            messages.success(request, f'Generated {generated_count} barcodes')
            
        except Exception as e:
            messages.error(request, f'Failed to generate barcodes: {str(e)}')
    
    return redirect('inventory:barcode_generator')

@login_required
@inventory_permission_required('view')
def print_barcode_labels_view(request):
    """
    Generate printable barcode labels.
    """
    try:
        product_ids = request.GET.getlist('product_id')
        
        if not product_ids:
            messages.error(request, 'No products selected for label printing')
            return redirect('inventory:product_list')
        
        products = Product.objects.filter(
            id__in=product_ids,
            is_active=True
        ).exclude(barcode__in=['', None])
        
        if not products:
            messages.error(request, 'No products with barcodes found')
            return redirect('inventory:product_list')
        
        # Generate barcode images
        label_data = []
        for product in products:
            try:
                # Generate barcode using python-barcode or similar
                from barcode import Code128
                from barcode.writer import ImageWriter
                
                code = Code128(product.barcode, writer=ImageWriter())
                barcode_buffer = BytesIO()
                code.write(barcode_buffer)
                barcode_b64 = base64.b64encode(barcode_buffer.getvalue()).decode()
                
                label_data.append({
                    'product': product,
                    'barcode_image': f'data:image/png;base64,{barcode_b64}',
                })
                
            except Exception as e:
                logger.error(f"Failed to generate barcode for {product.sku}: {str(e)}")
                continue
        
        return render(request, 'inventory/barcode/barcode_labels.html', {
            'page_title': 'Barcode Labels',
            'label_data': label_data,
        })
        
    except Exception as e:
        messages.error(request, f'Failed to generate labels: {str(e)}')
        return redirect('inventory:product_list')

@login_required
@inventory_permission_required('view')
def product_qr_code_api(request, product_id):
    """Generate QR code for a product"""
    try:
        product = get_object_or_404(Product, id=product_id, is_active=True)
        
        # Create QR code data
        qr_data = {
            'type': 'product',
            'sku': product.sku,
            'name': product.name,
            'brand': product.brand.name,
            'price': float(product.selling_price),
            'stock': product.total_stock,
            'url': request.build_absolute_uri(
                reverse('inventory:product_detail', args=[product.id])
            )
        }
        
        # Generate QR code
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=4,
        )
        qr.add_data(json.dumps(qr_data))
        qr.make(fit=True)
        
        # Create image
        img = qr.make_image(fill_color="black", back_color="white")
        
        # Convert to base64
        buffer = BytesIO()
        img.save(buffer, format='PNG')
        img_data = base64.b64encode(buffer.getvalue()).decode()
        
        return JsonResponse({
            'success': True,
            'qr_code': img_data,
            'qr_data': qr_data
        })
        
    except Exception as e:
        logger.error(f"Error generating QR code: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
@inventory_permission_required('view')
def barcode_lookup_api(request):
    """
    API endpoint: Look up product by barcode.
    """
    try:
        barcode = request.GET.get('barcode', '').strip()
        
        if not barcode:
            return JsonResponse({'success': False, 'error': 'No barcode provided'})
        
        try:
            product = Product.objects.select_related(
                'category', 'supplier', 'brand'
            ).get(barcode=barcode, is_active=True)
            
            # Get stock levels
            stock_levels = []
            for stock_level in product.stock_levels.select_related('location'):
                stock_levels.append({
                    'location_id': stock_level.location.id,
                    'location_name': stock_level.location.name,
                    'quantity': stock_level.quantity,
                    'available_quantity': stock_level.available_quantity,
                })
            
            data = {
                'product_id': product.id,
                'sku': product.sku,
                'name': product.name,
                'description': product.description,
                'category': product.category.name if product.category else '',
                'supplier': product.supplier.name if product.supplier else '',
                'brand': product.brand.name if product.brand else '',
                'cost_price': float(product.cost_price),
                'selling_price': float(product.selling_price),
                'current_stock': product.current_stock,
                'reorder_level': product.reorder_level,
                'stock_status': get_stock_status(product),
                'stock_levels': stock_levels,
                'last_restocked': product.last_restocked_date.isoformat() if product.last_restocked_date else None,
            }
            
            return JsonResponse({'success': True, 'product': data})
            
        except Product.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': f'Product with barcode "{barcode}" not found'
            })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

# =====================================
# QUICK ACCESS AND UTILITY VIEWS
# =====================================

class QuickAddProductView(LoginRequiredMixin, CreateView):
    """
    Quick product addition for mobile and fast entry.
    """
    model = Product
    fields = ['name', 'sku', 'category', 'supplier', 'cost_price', 'selling_price', 'current_stock']
    template_name = 'inventory/quick/quick_add_product.html'
    
    @method_decorator(inventory_permission_required('edit'))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'page_title': 'Quick Add Product',
            'is_quick_form': True,
        })
        return context
    
    def form_valid(self, form):
        form.instance.created_by = self.request.user
        
        # Generate SKU if not provided
        if not form.instance.sku:
            timestamp = timezone.now().strftime('%Y%m%d%H%M%S')
            form.instance.sku = f"QA-{timestamp}"
        
        messages.success(self.request, f'Product "{form.instance.name}" added successfully')
        return super().form_valid(form)
    
    def get_success_url(self):
        return reverse('inventory:quick_add_product')  # Stay on same page for continuous entry

class CurrencyListView(LoginRequiredMixin, ListView):
    """List all currencies and their exchange rates."""
    model = Currency
    template_name = 'inventory/configuration/currency_list.html'
    context_object_name = 'currencies'

    @method_decorator(inventory_permission_required('view'))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def get_queryset(self):
        return Currency.objects.order_by('code')

class CurrencyCreateView(LoginRequiredMixin, CreateView):
    """Create a new currency."""
    model = Currency
    form_class = CurrencyForm
    template_name = 'inventory/configuration/currency_form.html'
    success_url = reverse_lazy('inventory:currency_list')

    @method_decorator(inventory_permission_required('edit'))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def form_valid(self, form):
        messages.success(self.request, f'Currency "{form.instance.code}" created successfully.')
        return super().form_valid(form)

class CurrencyUpdateView(LoginRequiredMixin, UpdateView):
    """Edit an existing currency."""
    model = Currency
    form_class = CurrencyForm
    template_name = 'inventory/configuration/currency_form.html'
    success_url = reverse_lazy('inventory:currency_list')

    @method_decorator(inventory_permission_required('edit'))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def form_valid(self, form):
        messages.success(self.request, f'Currency "{form.instance.code}" updated successfully.')
        return super().form_valid(form)

class CurrencyDeleteView(LoginRequiredMixin, DeleteView):
    """Delete a currency (admin only)."""
    model = Currency
    template_name = 'inventory/configuration/currency_confirm_delete.html'
    success_url = reverse_lazy('inventory:currency_list')

    @method_decorator(inventory_permission_required('admin'))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

@login_required
@inventory_permission_required('view')
def quick_stock_check_view(request):
    """
    Quick stock level check by SKU or barcode.
    """
    result = None
    search_term = request.GET.get('q', '').strip()
    
    if search_term:
        try:
            product = Product.objects.select_related(
                'category', 'supplier'
            ).get(
                Q(sku__iexact=search_term) | Q(barcode=search_term),
                is_active=True
            )
            
            stock_levels = product.stock_levels.select_related('location').all()
            
            result = {
                'product': product,
                'stock_levels': stock_levels,
                'total_stock': product.current_stock,
                'stock_status': get_stock_status(product),
            }
            
        except Product.DoesNotExist:
            result = {
                'error': f'Product with SKU/Barcode "{search_term}" not found'
            }
        except Exception as e:
            result = {
                'error': f'Error: {str(e)}'
            }
    
    return render(request, 'inventory/quick/quick_stock_check.html', {
        'page_title': 'Quick Stock Check',
        'search_term': search_term,
        'result': result,
    })

@login_required
@inventory_permission_required('view')
def quick_cost_calculator_view(request):
    """
    Quick cost calculation tool.
    """
    calculation_result = None
    
    if request.GET.get('product_id') and request.GET.get('quantity'):
        try:
            product_id = int(request.GET.get('product_id'))
            quantity = int(request.GET.get('quantity', 1))
            
            product = get_object_or_404(Product, id=product_id)
            
            # Calculate costs
            base_cost = product.cost_price * quantity
            
            # Apply overhead factors
            overhead_factors = OverheadFactor.objects.filter(is_active=True)
            total_overhead_rate = sum(factor.percentage for factor in overhead_factors)
            overhead_cost = base_cost * (total_overhead_rate / 100)
            total_cost = base_cost + overhead_cost
            
            # Calculate margins
            selling_price = product.selling_price * quantity
            gross_margin = selling_price - base_cost
            net_margin = selling_price - total_cost
            
            calculation_result = {
                'product': product,
                'quantity': quantity,
                'base_cost': base_cost,
                'overhead_rate': total_overhead_rate,
                'overhead_cost': overhead_cost,
                'total_cost': total_cost,
                'selling_price': selling_price,
                'gross_margin': gross_margin,
                'net_margin': net_margin,
                'gross_margin_percent': (gross_margin / selling_price * 100) if selling_price > 0 else 0,
                'net_margin_percent': (net_margin / selling_price * 100) if selling_price > 0 else 0,
            }
            
        except Exception as e:
            calculation_result = {
                'error': f'Calculation failed: {str(e)}'
            }
    
    # Get recent products for quick selection
    recent_products = Product.objects.filter(is_active=True).order_by('-created_at')[:20]
    
    return render(request, 'inventory/quick/quick_cost_calculator.html', {
        'page_title': 'Quick Cost Calculator',
        'calculation_result': calculation_result,
        'recent_products': recent_products,
    })

# =====================================
# DASHBOARD WIDGET API ENDPOINTS
# =====================================

@login_required
@inventory_permission_required('view')
def low_stock_widget_api(request):
    """
    API endpoint for low stock dashboard widget.
    """
    try:
        low_stock_products = Product.objects.filter(
            current_stock__lte=F('reorder_level'),
            is_active=True
        ).select_related('category', 'supplier').order_by('current_stock')[:10]
        
        widget_data = []
        for product in low_stock_products:
            widget_data.append({
                'id': product.id,
                'name': product.name,
                'sku': product.sku,
                'current_stock': product.current_stock,
                'reorder_level': product.reorder_level,
                'shortage': product.reorder_level - product.current_stock,
                'category': product.category.name if product.category else '',
                'supplier': product.supplier.name if product.supplier else '',
                'url': reverse('inventory:product_detail', args=[product.id]),
            })
        
        return JsonResponse({
            'success': True,
            'data': widget_data,
            'total_low_stock': low_stock_products.count(),
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
@inventory_permission_required('view')
def top_products_widget_api(request):
    """
    API endpoint for top products dashboard widget.
    """
    try:
        # Get top products by stock value
        products = Product.objects.filter(
            is_active=True,
            current_stock__gt=0
        ).annotate(
            stock_value=F('current_stock') * F('cost_price')
        ).order_by('-stock_value')[:10]
        
        widget_data = []
        for product in products:
            widget_data.append({
                'id': product.id,
                'name': product.name,
                'sku': product.sku,
                'current_stock': product.current_stock,
                'stock_value': float(product.current_stock * product.cost_price),
                'url': reverse('inventory:product_detail', args=[product.id]),
            })
        
        return JsonResponse({
            'success': True,
            'data': widget_data,
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
@inventory_permission_required('view')
def global_inventory_search(request):
    """
    Global search across all inventory items.
    """
    try:
        query = request.GET.get('q', '').strip()
        limit = int(request.GET.get('limit', 20))
        
        if len(query) < 2:
            return JsonResponse({'results': []})
        
        # Search products
        products = Product.objects.filter(
            Q(name__icontains=query) |
            Q(sku__icontains=query) |
            Q(barcode__icontains=query) |
            Q(description__icontains=query),
            is_active=True
        ).select_related('category', 'supplier')[:limit]
        
        results = []
        for product in products:
            results.append({
                'type': 'product',
                'id': product.id,
                'title': product.name,
                'subtitle': f"SKU: {product.sku} | Stock: {product.current_stock}",
                'category': product.category.name if product.category else '',
                'url': reverse('inventory:product_detail', args=[product.id]),
                'stock_status': get_stock_status(product),
            })
        
        # Search suppliers if query is longer
        if len(query) >= 3:
            suppliers = Supplier.objects.filter(
                name__icontains=query,
                is_active=True
            )[:5]
            
            for supplier in suppliers:
                results.append({
                    'type': 'supplier',
                    'id': supplier.id,
                    'title': supplier.name,
                    'subtitle': f"Products: {supplier.products.filter(is_active=True).count()}",
                    'url': reverse('inventory:supplier_detail', args=[supplier.id]),
                })
        
        return JsonResponse({
            'success': True,
            'results': results,
            'total_found': len(results),
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

# =====================================
# SYSTEM ADMINISTRATION VIEWS
# =====================================

@login_required
@inventory_permission_required('admin')
def system_health_view(request):
    """System health monitoring"""
    context = {
        'page_title': 'System Health',
    }
    return render(request, 'inventory/admin/system_health.html', context)

@login_required
@inventory_permission_required('admin')
def performance_metrics_view(request):
    """Performance metrics"""
    context = {
        'page_title': 'Performance Metrics',
    }
    return render(request, 'inventory/admin/performance_metrics.html', context)

@login_required
@inventory_permission_required('admin')
def audit_log_view(request):
    """Audit log view"""
    context = {
        'page_title': 'Audit Log',
    }
    return render(request, 'inventory/admin/audit_log.html', context)

@login_required
@inventory_permission_required('admin')
def system_settings_view(request):
    """System settings"""
    context = {
        'page_title': 'System Settings',
    }
    return render(request, 'inventory/admin/system_settings.html', context)

@login_required
@inventory_permission_required('admin')
def system_maintenance_view(request):
    """System maintenance"""
    context = {
        'page_title': 'System Maintenance',
    }
    return render(request, 'inventory/admin/system_maintenance.html', context)

@login_required
@inventory_permission_required('admin')
def user_activity_view(request):
    """User activity monitoring"""
    context = {
        'page_title': 'User Activity',
    }
    return render(request, 'inventory/admin/user_activity.html', context)

@login_required
@inventory_permission_required('admin')
def permission_overview_view(request):
    """Permission overview"""
    context = {
        'page_title': 'Permission Overview',
    }
    return render(request, 'inventory/admin/permission_overview.html', context)

# =====================================
# SUPPORT VIEWS AND BUSINESS ENDPOINTS
# =====================================

@login_required
def inventory_help_view(request):
    """Render static help documentation for inventory module."""
    return render(request, "inventory/help.html")


@login_required
def getting_started_guide(request):
    """Render getting started guide."""
    return render(request, "inventory/getting_started.html")


@login_required
def cost_calculation_guide(request):
    """Render cost calculation guide."""
    return render(request, "inventory/cost_calculation_guide.html")


@login_required
def api_documentation_view(request):
    """Render internal API documentation."""
    return render(request, "inventory/api_documentation.html")


@login_required
def supplier_contact_view(request, pk):
    """Display contact information for a supplier."""
    supplier = get_object_or_404(Supplier, pk=pk, is_active=True)
    return render(request, "inventory/suppliers/contact.html", {"supplier": supplier})


@login_required
def quick_reorder_view(request):
    """List products that need reordering."""
    products = Product.objects.filter(
        is_active=True,
        total_stock__lt=F("reorder_level")
    ).select_related("supplier", "category")
    for product in products:
        product.reorder_qty = max(product.reorder_level * 2 - product.total_stock, 0)
    return render(request, "inventory/quick_reorder.html", {"products": products})


@login_required
def update_exchange_rates_view(request):
    """UI for updating currency exchange rates."""
    if request.method == "POST":
        try:
            updated = _refresh_exchange_rates()
            messages.success(request, f"Updated rates for {len(updated)} currencies.")
        except Exception as exc:
            messages.error(request, f"Failed to update rates: {exc}")
        return redirect("inventory:update_exchange_rates")
    currencies = Currency.objects.filter(is_active=True).order_by("code")
    return render(
        request,
        "inventory/configuration/update_exchange_rates.html",
        {"currencies": currencies},
    )


def _refresh_exchange_rates():
    """Fetch latest exchange rates from public API and update Currency table."""
    response = requests.get(
        "https://api.exchangerate.host/latest?base=USD", timeout=10
    )
    response.raise_for_status()
    data = response.json()
    rates = data.get("rates", {})
    updated = {}
    for currency in Currency.objects.filter(is_active=True):
        if currency.code == "USD":
            currency.exchange_rate_to_usd = Decimal("1")
        elif currency.code in rates:
            rate = Decimal(str(rates[currency.code]))
            currency.exchange_rate_to_usd = Decimal("1") / rate
        else:
            continue
        currency.save(update_fields=["exchange_rate_to_usd", "last_updated"])
        updated[currency.code] = float(currency.exchange_rate_to_usd)
    return updated

# --- API views ---

@login_required
@require_POST
def update_exchange_rates_api(request):
    """Update currency exchange rates via API."""
    try:
        updated = _refresh_exchange_rates()
        return JsonResponse({"status": "success", "updated": updated})
    except Exception as exc:
        return JsonResponse(
            {"status": "error", "message": str(exc)}, status=500
        )


@login_required
def currency_rates_api(request):
    """Return current exchange rates."""
    currencies = Currency.objects.filter(is_active=True).values(
        "code", "name", "exchange_rate_to_usd", "last_updated"
    )
    data = []
    for cur in currencies:
        cur["exchange_rate_to_usd"] = float(cur["exchange_rate_to_usd"])
        cur["last_updated"] = cur["last_updated"].isoformat()
        data.append(cur)
    return JsonResponse({"currencies": data})


@login_required
def currency_convert_api(request):
    """Convert amount between two currencies."""
    amount = request.GET.get("amount")
    from_code = request.GET.get("from")
    to_code = request.GET.get("to")
    if not all([amount, from_code, to_code]):
        return JsonResponse(
            {"status": "error", "message": "Missing parameters"}, status=400
        )
    try:
        amount = Decimal(amount)
        from_currency = Currency.objects.get(code=from_code)
        to_currency = Currency.objects.get(code=to_code)
        usd_amount = amount * from_currency.exchange_rate_to_usd
        target_amount = usd_amount / to_currency.exchange_rate_to_usd
        return JsonResponse(
            {"status": "success", "result": float(target_amount)}
        )
    except Currency.DoesNotExist:
        return JsonResponse(
            {"status": "error", "message": "Unknown currency"}, status=404
        )
    except Exception as exc:
        return JsonResponse(
            {"status": "error", "message": str(exc)}, status=400
        )


@login_required
@require_POST
def bulk_price_update_api(request):
    """Apply percentage change to selling prices of multiple products."""
    try:
        payload = json.loads(request.body or "{}")
        ids = payload.get("product_ids", [])
        percentage = Decimal(str(payload.get("percentage", 0)))
    except (ValueError, TypeError):
        return JsonResponse({"status": "error", "message": "Invalid payload"}, status=400)
    if not ids or percentage == 0:
        return JsonResponse({"status": "error", "message": "Invalid parameters"}, status=400)
    multiplier = Decimal("1") + (percentage / 100)
    with transaction.atomic():
        products = Product.objects.filter(id__in=ids)
        for product in products:
            product.selling_price = product.selling_price * multiplier
            product.save(update_fields=["selling_price"])
    return JsonResponse({"status": "success", "updated": products.count()})


@login_required
def margin_analysis_api(request):
    """Provide margin information for products."""
    product_id = request.GET.get("product_id")
    if product_id:
        product = get_object_or_404(Product, pk=product_id)
        if product.selling_price:
            margin = (
                (product.selling_price - product.total_cost_price_usd)
                / product.selling_price
            ) * 100
        else:
            margin = Decimal("0")
        return JsonResponse(
            {
                "product": product.id,
                "margin_percent": float(margin.quantize(Decimal('0.01'))),
            }
        )
    qs = Product.objects.filter(is_active=True, selling_price__gt=0).annotate(
        margin_percent=100
        * (F("selling_price") - F("total_cost_price_usd"))
        / F("selling_price")
    )
    avg_margin = qs.aggregate(avg=Avg("margin_percent"))["avg"] or 0
    return JsonResponse({"average_margin_percent": float(avg_margin)})


@login_required
def category_performance_api(request):
    """Return performance metrics grouped by category."""
    categories = (
        Category.objects.filter(is_active=True)
        .annotate(
            product_count=Count("products", filter=Q(products__is_active=True)),
            stock_value=Sum(
                F("products__total_stock") * F("products__total_cost_price_usd")
            ),
        )
        .order_by("-stock_value")[:10]
        .values("id", "name", "product_count", "stock_value")
    )
    data = []
    for cat in categories:
        cat["stock_value"] = float(cat["stock_value"] or 0)
        data.append(cat)
    return JsonResponse({"categories": data})


@login_required
def check_stock_availability_api(request):
    """Check if enough stock is available for a product."""
    product_id = request.GET.get("product_id")
    quantity = int(request.GET.get("quantity", "0"))
    product = get_object_or_404(Product, pk=product_id, is_active=True)
    available = product.total_stock >= quantity
    return JsonResponse(
        {
            "product_id": product.id,
            "requested": quantity,
            "available": available,
            "available_quantity": product.total_stock,
        }
    )


@login_required
def generate_barcode_api(request, product_id):
    """Generate barcode image for a product."""
    product = get_object_or_404(Product, pk=product_id)
    code = product.barcode or product.sku
    img_data = BarcodeManager.generate_barcode(code)
    if not img_data:
        return JsonResponse(
            {"status": "error", "message": "Failed to generate barcode"}, status=500
        )
    return JsonResponse({"status": "success", "barcode": img_data})


@login_required
@require_POST
def qr_code_scan_api(request):
    """Lookup product based on scanned QR code data."""
    try:
        payload = json.loads(request.body or "{}")
    except json.JSONDecodeError:
        payload = request.POST
    code = payload.get("code")
    if not code:
        return JsonResponse({"status": "error", "message": "No code provided"}, status=400)
    product = Product.objects.filter(qr_code=code).first()
    if not product:
        return JsonResponse({"status": "error", "message": "Product not found"}, status=404)
    return JsonResponse({"status": "success", "product": {"id": product.id, "name": product.name}})


@login_required
def reorder_recommendations_api(request):
    """Suggest reorder quantities for low stock products."""
    products = (
        Product.objects.filter(is_active=True, total_stock__lt=F("reorder_level"))
        .order_by("total_stock")
        .select_related("supplier")
    )
    recommendations = []
    for product in products:
        qty = max(product.reorder_level * 2 - product.total_stock, 0)
        recommendations.append(
            {
                "product_id": product.id,
                "name": product.name,
                "supplier": product.supplier.name,
                "recommended_quantity": qty,
            }
        )
    return JsonResponse({"recommendations": recommendations})


@login_required
def stock_levels_api(request):
    """Return current stock levels for active products."""
    products = Product.objects.filter(is_active=True).values(
        "id", "name", "total_stock", "reorder_level"
    )
    return JsonResponse({"products": list(products)})


@login_required
def stock_movements_api(request):
    """Return recent stock movement records."""
    movements = (
        StockMovement.objects.select_related("product")
        .order_by("-created_at")[:50]
        .values("id", "product__name", "movement_type", "quantity", "created_at")
    )
    data = [
        {
            "id": m["id"],
            "product": m["product__name"],
            "type": m["movement_type"],
            "quantity": m["quantity"],
            "date": m["created_at"].isoformat(),
        }
        for m in movements
    ]
    return JsonResponse({"movements": data})


@login_required
def stock_trends_api(request):
    """Return incoming/outgoing stock totals per day."""
    days = int(request.GET.get("days", 30))
    start = timezone.now() - timedelta(days=days)
    qs = StockMovement.objects.filter(created_at__gte=start)
    daily = (
        qs.annotate(day=TruncDate("created_at"))
        .values("day")
        .annotate(
            incoming=Sum(
                "quantity", filter=Q(movement_type__in=["in", "purchase", "return"])
            ),
            outgoing=Sum(
                "quantity", filter=Q(movement_type__in=["out", "sale", "adjustment", "damaged"])
            ),
        )
        .order_by("day")
    )
    data = []
    for entry in daily:
        data.append(
            {
                "date": entry["day"].isoformat(),
                "incoming": entry["incoming"] or 0,
                "outgoing": abs(entry["outgoing"] or 0),
            }
        )
    return JsonResponse({"trends": data})


@login_required
def supplier_alerts_widget_api(request):
    """Return count of active reorder alerts grouped by supplier."""
    alerts = (
        ReorderAlert.objects.filter(status__in=["active", "acknowledged"])
        .values("product__supplier__id", "product__supplier__name")
        .annotate(alert_count=Count("id"))
        .order_by("-alert_count")[:5]
    )
    data = [
        {
            "supplier_id": a["product__supplier__id"],
            "supplier": a["product__supplier__name"],
            "alerts": a["alert_count"],
        }
        for a in alerts
    ]
    return JsonResponse({"suppliers": data})


@login_required
def supplier_country_analysis_api(request):
    """Return number of suppliers per country."""
    countries = (
        Supplier.objects.values("country__name")
        .annotate(count=Count("id"))
        .order_by("-count")
    )
    data = [{"country": c["country__name"], "suppliers": c["count"]} for c in countries]
    return JsonResponse({"countries": data})


@login_required
def dashboard_analytics_api(request):
    """Return basic dashboard metrics as JSON."""
    products = Product.objects.filter(is_active=True)
    totals = products.aggregate(
        total_stock_value=Sum(F("total_stock") * F("total_cost_price_usd")),
        low_stock=Count(
            "id",
            filter=Q(total_stock__lte=F("reorder_level"), total_stock__gt=0),
        ),
        out_of_stock=Count("id", filter=Q(total_stock=0)),
    )
    response = {
        "total_products": products.count(),
        "total_categories": Category.objects.filter(is_active=True).count(),
        "total_suppliers": Supplier.objects.filter(is_active=True).count(),
        "stock_value": float(totals["total_stock_value"] or 0),
        "low_stock": totals["low_stock"] or 0,
        "out_of_stock": totals["out_of_stock"] or 0,
    }
    return JsonResponse(response)


@login_required
def cost_trends_widget_api(request):
    """Return purchase cost totals grouped by month."""
    months = int(request.GET.get("months", 6))
    start = timezone.now() - timedelta(days=30 * months)
    qs = (
        StockMovement.objects.filter(
            created_at__gte=start, movement_type__in=["in", "purchase"]
        )
        .annotate(month=TruncDate("created_at"))
        .values("month")
        .annotate(
            total=Sum(F("quantity") * F("product__total_cost_price_usd"))
        )
        .order_by("month")
    )
    data = [
        {"date": entry["month"].isoformat(), "cost": float(entry["total"] or 0)}
        for entry in qs
    ]
    return JsonResponse({"cost_trends": data})

# Initialize logging
logger.info("Inventory management views loaded successfully")
